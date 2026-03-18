#!/usr/bin/env python3
"""
UNIFIED TRAVELER PIPELINE v4.0
==============================
Major Enhancements over v3.0:
  - Multi-asset support: NQ, ES, RTY processed in a single run
  - 16-file intake: 5×15-min Origin, 5×3-min, 5×5-min, 1 measurement
  - Auto-discovery of feed files from directory with pattern validation
  - Unix / ISO timestamp auto-detection and NY conversion
  - Multi-chart OHLC construction (3-min → 6/9/12/18/24-min; 5-min → 10/20/30/45/60/90/120/240-min)
  - 15-min quality-check reconstruction from both 3-min and 5-min data
  - Moving Average normalization (h-series h1–h20, duplicate QC)
  - Raw + Matched traveler reports per asset (v3.0 core logic preserved)
  - Wick analysis: primary & secondary nested swing candles, per-asset Excel tabs
    reporting which MAs and Travelers plot inside candle wicks

Usage:
    python unified_traveler_pipeline_4_0.py \\
        --measurement "Meas 4_8 all.xlsx" \\
        --dt "2026-02-23 18:30" \\
        --lookback 22 \\
        [--dir "."]

Author: Built from unified_traveler_pipeline_3_0.py
Date:   March 2026
Version: 4.0
"""

import pandas as pd
import numpy as np
import datetime as dt
from datetime import datetime, timedelta
import argparse
import sys
import os
import re
import io
import time as time_module
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any

# Excel handling
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 120)
print("🌌 UNIFIED TRAVELER PIPELINE v4.0  —  NQ · ES · RTY  |  Multi-Timeframe  |  Wick Analysis")
print("=" * 120)

pd.set_option("styler.render.max_elements", 2_000_000)

# ============================================================================
# SECTION 1: DEPENDENCY IMPORTS  (same graceful-fallback pattern as v3.0)
# ============================================================================

print("\n📦 Loading dependencies...")

USE_NEW_CALCULATOR = False
apply_full_range_advanced = None

try:
    from custom_range_calculator_cli import apply_full_range_advanced as _calc
    apply_full_range_advanced = _calc
    print("  ✅ custom_range_calculator_cli")
    USE_NEW_CALCULATOR = True
except ImportError:
    if "streamlit" not in sys.modules:
        class _MockST:
            def __getattr__(self, n):
                def _f(*a, **kw):
                    return self if n in ("progress", "empty") else None
                return _f
        sys.modules["streamlit"] = _MockST()
    try:
        from custom_range_calculator_1125_21c import apply_full_range_advanced as _calc
        apply_full_range_advanced = _calc
        print("  ✅ custom_range_calculator_1125_21c (with streamlit mock)")
        USE_NEW_CALCULATOR = True
    except ImportError as e:
        try:
            from custom_range_calculator_0813 import apply_full_range_advanced as _calc
            apply_full_range_advanced = _calc
            print("  ✅ custom_range_calculator_0813 (fallback)")
        except ImportError:
            print("  \u274c No range calculator found.")
            apply_full_range_advanced = None  # will raise clearly at pipeline run

try:
    from a_helpers import (
        clean_timestamp, process_feed, get_input_at_day_start,
        GROUP_1A_TRAVELERS, GROUP_1B_TRAVELERS, GROUP_2A_TRAVELERS, GROUP_2B_TRAVELERS,
        get_input_value, generate_master_traveler_list,
        EPIC_ORIGINS, ANCHOR_ORIGINS,
    )
    print("  ✅ a_helpers")
except ImportError as e:
    print(f"  ❌ a_helpers — {e}")
    sys.exit(1)

try:
    from nested_swing_detector import analyze_swings, parse_timestamp_naive, detect_nested_swings
    print("  ✅ nested_swing_detector")
    HAS_SWING_DETECTOR = True
except ImportError as e:
    print(f"  ⚠️  nested_swing_detector — {e}  (wick analysis will be skipped)")
    HAS_SWING_DETECTOR = False
    def analyze_swings(*a, **kw): return [], []
    def parse_timestamp_naive(x): return pd.to_datetime(x)
    def detect_nested_swings(*a, **kw): return []

try:
    from model_definitions_v21 import MODELS, get_reciprocal_lookup, apply_special_matching
    print("  ✅ model_definitions_v21")
    HAS_MODEL_DEFINITIONS = True
except ImportError:
    HAS_MODEL_DEFINITIONS = False
    MODELS = {}
    print("  ⚠️  model_definitions_v21 not found")

HAS_MODEL_PROCESSOR = False
try:
    import importlib.util
    spec = importlib.util.find_spec("model_processor_v21")
    if spec and spec.origin:
        from model_processor_v21 import get_model_display_info, organize_results_by_category, create_summary_stats
        print("  ✅ model_processor_v21")
        HAS_MODEL_PROCESSOR = True
except Exception:
    pass

if not HAS_MODEL_PROCESSOR and HAS_MODEL_DEFINITIONS:
    def get_model_display_info(m):
        c = MODELS.get(m, {})
        return {"number": c.get("number", 0), "display_name": c.get("display_name", m),
                "description": c.get("description", ""), "check_recip": c.get("check_recip", False)}
    def organize_results_by_category(r): return {}
    def create_summary_stats(r):
        res = r.get("results", {})
        return {"total_models": len(res),
                "models_with_matches": sum(1 for d in res.values() if len(d) > 0),
                "total_matches": sum(len(d) for d in res.values())}

try:
    from bypass_mode_matcher import match_travelers_bypass_mode, process_model_bypass_mode
    print("  ✅ bypass_mode_matcher")
    HAS_BYPASS_MATCHER = True
except ImportError:
    HAS_BYPASS_MATCHER = False
    print("  ⚠️  bypass_mode_matcher not found")

print("\n✅ Critical dependencies loaded.\n")


# ============================================================================
# SECTION 2: CONFIGURATION & PER-ASSET CONSTANTS
# ============================================================================

# --- Session open ---
DEFAULT_SESSION_OPEN_HOUR = 18   # 18:00 NY; detected as 17 if present

# --- Per-asset defaults ---
ASSET_DEFAULTS = {
    "NQ": {"max_spread": 3.0, "radius": 1000, "composite": False},
    "ES": {"max_spread": 1.0, "radius": 215,  "composite": False},
    "RTY": {"max_spread": 1.0, "radius": 100,  "composite": True},   # no MRTY; single-feed composite
}

# --- Chart build specs ---
# From 3-min source: build these intervals (minutes)
BUILD_FROM_3MIN = [6, 9, 12, 18, 24]
# From 5-min source: build these intervals (minutes)
BUILD_FROM_5MIN = [10, 20, 30, 45, 60, 90, 120, 240]
# 15-min QC timeframes (built from both, compared to actual)
QC_TIMEFRAME = 15

# --- Priority origins (mirrored from v3.0) ---
EPIC_ORIGINS_BASE  = {"trinidad", "tobago", "wasp-12b", "macedonia"}
ANCHOR_ORIGINS_BASE = {"spain", "saturn", "jupiter", "kepler-62", "kepler-44"}
EPIC_ORIGINS_WITH_BRACKETS = set(EPIC_ORIGINS_BASE)
for _o in list(EPIC_ORIGINS_BASE):
    for _sfx in ["[1]", "[2]", "[-1]", "[-2]"]:
        EPIC_ORIGINS_WITH_BRACKETS.add(f"{_o}{_sfx}")
PRIORITY_ORIGINS = set(list(EPIC_ORIGINS_WITH_BRACKETS) + list(ANCHOR_ORIGINS_BASE))

DEFAULT_LOOKBACK_DAYS = 20


# ============================================================================
# SECTION 3: FILE DISCOVERY & VALIDATION
# ============================================================================

# Expected file-set: 5 × 15-min, 5 × 3-min, 5 × 5-min
# Asset keys for each timeframe group
_ASSET_PATTERNS = [
    # (asset, feed_role, prefix_pattern_upper)
    ("NQ",  "small", "CME_MINI_MNQ"),
    ("NQ",  "big",   "CME_MINI_NQ"),   # checked AFTER MNQ
    ("ES",  "small", "CME_MINI_MES"),
    ("ES",  "big",   "CME_MINI_ES"),   # checked AFTER MES
    ("RTY", "composite", "CME_MINI_RTY"),
]

def _classify_csv(filepath: Path) -> Optional[Dict]:
    """
    Return dict with keys: asset, role, timeframe  (or None if unrecognised)
    """
    name_upper = filepath.name.upper()

    # Determine asset / role — order matters (M-prefix first)
    asset = role = None
    for (a, r, pat) in _ASSET_PATTERNS:
        if pat in name_upper:
            asset, role = a, r
            break
    if asset is None:
        return None

    # Determine timeframe from ", 15" / ", 3" / ", 5" in name
    # Support comma-space or underscore variants
    name_check = filepath.name  # preserve original case for comma check
    if ", 15" in name_check or "_15 " in name_check or name_check.endswith("_15.csv"):
        tf = 15
    elif ", 3" in name_check and ", 30" not in name_check:
        tf = 3
    elif ", 5" in name_check and ", 50" not in name_check:
        tf = 5
    else:
        return None

    return {"path": filepath, "asset": asset, "role": role, "timeframe": tf}


def discover_feed_files(search_dir: str) -> Dict:
    """
    Scan search_dir for CSV files matching the 15-asset naming convention.
    Returns nested dict:  files[asset][timeframe] = {path, role}
    Also returns a flat list of all discovered entries.
    """
    search_path = Path(search_dir)
    csvs = list(search_path.glob("*.csv")) + list(search_path.glob("*.CSV"))

    discovered: Dict[str, Dict[int, Dict]] = {a: {} for a in ["NQ", "ES", "RTY"]}
    all_entries = []

    for csv_path in csvs:
        info = _classify_csv(csv_path)
        if info is None:
            continue
        a, tf = info["asset"], info["timeframe"]
        # For the same (asset, timeframe) keep the most-recently-modified file
        if tf not in discovered[a] or csv_path.stat().st_mtime > discovered[a][tf]["path"].stat().st_mtime:
            discovered[a][tf] = {"path": csv_path, "role": info["role"]}
            all_entries.append(info)

    return discovered, all_entries


def validate_file_set(discovered: Dict, measurement_path: str) -> bool:
    """
    Validate that all 16 required files are present.
    Prints detailed status.  Returns True if all present.
    """
    print("\n🔍 Validating file set (16 required)...")
    all_ok = True

    # Check measurement
    meas_ok = Path(measurement_path).exists()
    print(f"  {'✅' if meas_ok else '❌'} Measurement: {measurement_path}")
    if not meas_ok:
        all_ok = False

    # Check each asset × timeframe combination
    required = [(a, tf) for a in ["NQ", "ES", "RTY"] for tf in [15, 3, 5]]
    for (asset, tf) in required:
        if tf in discovered.get(asset, {}):
            info = discovered[asset][tf]
            print(f"  ✅ {asset:>3s}  {tf:>2d}-min  [{info['role']:>9s}]  {info['path'].name}")
        else:
            print(f"  ❌ {asset:>3s}  {tf:>2d}-min  — FILE MISSING  "
                  f"(expected CME_MINI_{'M' if asset != 'RTY' else ''}{'NQ' if asset == 'NQ' else asset}..., {tf}")
            all_ok = False

    if all_ok:
        print("  ✅ All 16 files present — ready for takeoff!")
    else:
        print("  ❌ Missing files detected — pipeline cannot continue.")
    return all_ok


# ============================================================================
# SECTION 4: TIMESTAMP NORMALIZATION
# ============================================================================

def detect_timestamp_type(df: pd.DataFrame) -> str:
    """
    Inspect first data row of column A.
    Returns 'unix' (seconds since epoch > 1e9) or 'iso'.
    """
    try:
        val = float(df.iloc[0, 0])
        if val > 1_000_000_000:
            return "unix"
    except (ValueError, TypeError):
        pass
    return "iso"


def insert_ny_datetime_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    Auto-detect Unix vs ISO in column A.
    Insert a 'datetime_ny' column immediately after column A with timezone-naive NY datetimes.
    EDT = UTC−4.  Uses pytz/dateutil via pandas tz_convert for proper DST handling.
    """
    df = df.copy()
    ts_type = detect_timestamp_type(df)
    time_col = df.columns[0]

    if ts_type == "unix":
        print(f"    🕐 Unix timestamps detected in column A → converting to NY (EDT/EST)")
        ny_times = (
            pd.to_datetime(df[time_col], unit="s", utc=True)
            .dt.tz_convert("America/New_York")
            .dt.tz_localize(None)
        )
    else:
        print(f"    🕐 ISO timestamps detected in column A → converting to clean NY datetime")
        raw = pd.to_datetime(df[time_col], utc=True, errors="coerce")
        ny_times = raw.dt.tz_convert("America/New_York").dt.tz_localize(None)

    df.insert(1, "datetime_ny", ny_times)
    print(f"    ✅ datetime_ny column inserted  "
          f"(range: {ny_times.min()}  →  {ny_times.max()})")
    return df


def load_and_normalize_csv(csv_path: Path) -> pd.DataFrame:
    """
    Load a feed CSV, insert NY datetime column, sort by time.
    """
    df = pd.read_csv(str(csv_path))
    df = insert_ny_datetime_column(df)
    df = df.sort_values("datetime_ny").reset_index(drop=True)
    return df


# ============================================================================
# SECTION 5: MULTI-CHART OHLC CONSTRUCTION
# ============================================================================

# Anchor datetime used for resample origin (18:00 on a Monday)
_RESAMPLE_ORIGIN_18 = pd.Timestamp("2000-01-03 18:00:00")
_RESAMPLE_ORIGIN_17 = pd.Timestamp("2000-01-03 17:00:00")


def detect_session_open_hour(df: pd.DataFrame) -> int:
    """
    Look at the datetime_ny column: return 17 if 17:00 bars exist more than 18:00, else 18.
    """
    times = df["datetime_ny"].dt.time
    count_18 = (df["datetime_ny"].dt.hour == 18).sum()
    count_17 = (df["datetime_ny"].dt.hour == 17).sum()
    return 17 if count_17 > count_18 else 18


def build_higher_tf_ohlc(df_source: pd.DataFrame,
                          interval_minutes: int,
                          session_open_hour: int = 18,
                          lookback_days: int = 10) -> pd.DataFrame:
    """
    Resample OHLC (and MAs if present) from df_source to interval_minutes bars.

    Parameters
    ----------
    df_source       : must have 'datetime_ny' and OHLC columns
    interval_minutes: target bar size in minutes
    session_open_hour: 17 or 18 (determines resample origin anchor)
    lookback_days   : limit data to this many calendar days

    Returns DataFrame with columns: datetime_ny, open, high, low, close, [ma cols...]
    """
    df = df_source.copy()
    df = df.set_index("datetime_ny").sort_index()

    # Limit lookback
    cutoff = df.index.max() - pd.Timedelta(days=lookback_days)
    df = df[df.index >= cutoff]

    if df.empty:
        return pd.DataFrame()

    origin = _RESAMPLE_ORIGIN_17 if session_open_hour == 17 else _RESAMPLE_ORIGIN_18
    freq = f"{interval_minutes}min"

    # OHLC aggregation
    agg_dict = {
        "open":  "first",
        "high":  "max",
        "low":   "min",
        "close": "last",
    }

    # Include MA columns (carry 'last' value of the window)
    ma_cols = [c for c in df.columns if c not in ("open", "high", "low", "close", "time",
                                                    "Volume", "volume", "RSI-8")]
    for mc in ma_cols:
        agg_dict[mc] = "last"

    resampled = (
        df.resample(freq, closed="left", label="left", origin=origin)
        .agg(agg_dict)
        .dropna(subset=["open", "close"])
    )

    resampled = resampled.reset_index().rename(columns={"index": "datetime_ny",
                                                          "datetime_ny": "datetime_ny"})
    # Ensure column name is correct after reset_index
    if resampled.columns[0] != "datetime_ny":
        resampled = resampled.rename(columns={resampled.columns[0]: "datetime_ny"})

    return resampled


def build_all_higher_timeframes(df_3min: pd.DataFrame,
                                 df_5min: pd.DataFrame,
                                 df_15min_actual: pd.DataFrame,
                                 asset_id: str,
                                 session_open_hour: int = 18,
                                 lookback_days: int = 10) -> Dict[int, pd.DataFrame]:
    """
    Build all derived timeframe DataFrames.
    Returns dict keyed by interval_minutes.
    Also prints QC report comparing reconstructed 15-min vs actual.
    """
    print(f"\n  📊 Building higher-timeframe OHLC charts for {asset_id}...")
    charts = {}

    # From 3-min
    for mins in BUILD_FROM_3MIN:
        tf_df = build_higher_tf_ohlc(df_3min, mins, session_open_hour, lookback_days)
        charts[mins] = tf_df
        print(f"    {mins:>3d}-min (from 3-min): {len(tf_df):>5d} bars")

    # From 5-min
    for mins in BUILD_FROM_5MIN:
        tf_df = build_higher_tf_ohlc(df_5min, mins, session_open_hour, lookback_days)
        charts[mins] = tf_df
        print(f"    {mins:>3d}-min (from 5-min): {len(tf_df):>5d} bars")

    # QC: reconstruct 15-min from 3-min and 5-min
    qc_from_3 = build_higher_tf_ohlc(df_3min, 15, session_open_hour, lookback_days)
    qc_from_5 = build_higher_tf_ohlc(df_5min, 15, session_open_hour, lookback_days)
    _run_15min_qc(df_15min_actual, qc_from_3, qc_from_5, asset_id)

    # Store actual 15-min for reference
    charts[15] = df_15min_actual

    return charts


def _run_15min_qc(df_actual: pd.DataFrame,
                  df_from_3: pd.DataFrame,
                  df_from_5: pd.DataFrame,
                  asset_id: str,
                  tolerance: float = 0.5):
    """
    Compare reconstructed 15-min bars (from 3-min and from 5-min) against actual 15-min.
    Reports mismatches beyond tolerance on Close.
    """
    print(f"\n  🔬 15-min QC for {asset_id}  (tolerance: ±{tolerance} pts on Close)")

    def _prep(df):
        d = df.copy()
        # Normalize time col
        if "datetime_ny" in d.columns:
            d = d.set_index("datetime_ny")
        elif "time" in d.columns:
            d = d.set_index("time")
        return d[["open", "high", "low", "close"]].sort_index()

    actual  = _prep(df_actual)
    from_3  = _prep(df_from_3)
    from_5  = _prep(df_from_5)

    common_3 = actual.index.intersection(from_3.index)
    common_5 = actual.index.intersection(from_5.index)

    mismatches_3 = mismatches_5 = 0
    for ts in common_3:
        diff = abs(actual.loc[ts, "close"] - from_3.loc[ts, "close"])
        if diff > tolerance:
            mismatches_3 += 1
    for ts in common_5:
        diff = abs(actual.loc[ts, "close"] - from_5.loc[ts, "close"])
        if diff > tolerance:
            mismatches_5 += 1

    pct_ok_3 = 100 * (1 - mismatches_3 / max(len(common_3), 1))
    pct_ok_5 = 100 * (1 - mismatches_5 / max(len(common_5), 1))
    icon_3 = "✅" if mismatches_3 == 0 else "⚠️ "
    icon_5 = "✅" if mismatches_5 == 0 else "⚠️ "

    print(f"    {icon_3} vs 3-min reconstruction: {len(common_3)} bars compared, "
          f"{mismatches_3} mismatches ({pct_ok_3:.1f}% match)")
    print(f"    {icon_5} vs 5-min reconstruction: {len(common_5)} bars compared, "
          f"{mismatches_5} mismatches ({pct_ok_5:.1f}% match)")


# ============================================================================
# SECTION 6: MOVING AVERAGE NORMALIZATION
# ============================================================================

# h-series: extract name from column header (strip whitespace, take first token)
def _extract_h_name(col: str) -> Optional[str]:
    stripped = col.strip()
    if not stripped.lower().startswith("h"):
        return None
    tokens = stripped.split()
    return tokens[0].lower() if tokens else None


def normalize_ma_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Process moving average columns in df:
      1. Map all h-series columns to a canonical name (h1–h20, h4a, h4b, h7a, h7b)
      2. Verify duplicate h-columns carry identical values (QC)
      3. Drop duplicate columns keeping one canonical copy

    Returns (cleaned_df, name_map) where name_map is {original_col: canonical_name}
    """
    df = df.copy()
    name_map: Dict[str, str] = {}     # original col → canonical
    h_groups: Dict[str, List[str]] = {}  # canonical → [original cols...]

    for col in df.columns:
        h_name = _extract_h_name(col)
        if h_name is None:
            continue
        name_map[col] = h_name
        h_groups.setdefault(h_name, []).append(col)

    # QC duplicates
    qc_issues = []
    cols_to_drop = []
    rename_map = {}

    for canonical, orig_cols in h_groups.items():
        if len(orig_cols) == 1:
            rename_map[orig_cols[0]] = canonical
            continue

        # Multiple columns with same canonical name — verify identical values
        reference = df[orig_cols[0]]
        all_equal = all(
            df[c].fillna(0).round(4).equals(reference.fillna(0).round(4))
            for c in orig_cols[1:]
        )

        if not all_equal:
            qc_issues.append(f"⚠️  {canonical}: duplicate columns differ "
                              f"({', '.join(orig_cols)})")

        # Keep first occurrence, drop the rest
        rename_map[orig_cols[0]] = canonical
        for dup_col in orig_cols[1:]:
            cols_to_drop.append(dup_col)

    # Apply renames and drops
    df = df.drop(columns=cols_to_drop, errors="ignore")
    df = df.rename(columns=rename_map)

    if qc_issues:
        for issue in qc_issues:
            print(f"    {issue}")
    else:
        duplicated = [c for c in h_groups if len(h_groups[c]) > 1]
        if duplicated:
            print(f"    ✅ Duplicate MA columns verified identical: {', '.join(duplicated)}")

    return df, rename_map


def get_all_ma_column_names(df: pd.DataFrame) -> List[str]:
    """
    Return list of all MA column names present in df (h-series + numeric-period series).
    Excludes OHLC, time, datetime_ny, Volume, RSI columns.
    """
    exclude = {"time", "datetime_ny", "open", "high", "low", "close",
               "volume", "Volume", "RSI-8", "rsi-8"}
    return [c for c in df.columns if c not in exclude]


# ============================================================================
# SECTION 7: PER-ASSET RAW + MATCHED TRAVELER REPORTS
#            (wraps v3.0 core logic; handles NQ/ES/RTY with per-asset defaults)
# ============================================================================

def _preprocess_15min_feed(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare a 15-min feed DataFrame for the range calculator:
    ensure 'time' column is timezone-naive NY datetime.
    """
    df = df.copy()
    time_col = None
    for cand in ["datetime_ny", "time"]:
        if cand in df.columns:
            time_col = cand
            break
    if time_col is None:
        raise ValueError("Feed has no 'time' or 'datetime_ny' column")

    if time_col != "time":
        # Drop the original raw 'time' column (Unix integers) before renaming
        # datetime_ny → 'time'; otherwise two columns share the name and
        # pd.to_datetime() receives a DataFrame, causing "duplicate keys" error.
        if "time" in df.columns:
            df = df.drop(columns=["time"])
        df = df.rename(columns={time_col: "time"})

    df["time"] = pd.to_datetime(df["time"], errors="coerce")
    try:
        df["time"] = df["time"].dt.tz_localize(None)
    except TypeError:
        pass

    return df.sort_values("time").reset_index(drop=True)


def _generate_travelers_with_calculator(measurement_df, feed1_df, feed2_df,
                                         report_time, window_radius,
                                         input_value_at_start, allowed_origins=None):
    """Wrapper identical to v3.0 — filters feeds then calls apply_full_range_advanced."""
    if allowed_origins:
        base_cols = {"time", "open", "high", "low", "close"}
        if "volume" in feed1_df.columns:
            base_cols.add("volume")

        filtered_cols = set(base_cols)
        for origin in allowed_origins:
            is_bracketed = False
            base_origin = origin
            bracket_idx = None
            if "[" in origin:
                parts = origin.split("[")
                base_origin = parts[0]
                bracket_part = parts[1].rstrip("]")
                if base_origin.lower() in ["wasp-12b", "macedonia"]:
                    is_bracketed = True
                    bracket_idx = bracket_part.lstrip("-")

            if is_bracketed:
                for sfx in [" H", " L", " C"]:
                    for variant in [base_origin, base_origin.upper(),
                                    base_origin.capitalize(), "WASP-12b"]:
                        col = f"{variant}{sfx}[{bracket_idx}]"
                        for df_ref in [feed1_df, feed2_df]:
                            if col in df_ref.columns:
                                filtered_cols.add(col)
            else:
                for sfx in [" H", " L", " C"]:
                    for df_ref in [feed1_df, feed2_df]:
                        col = origin + sfx
                        if col in df_ref.columns:
                            filtered_cols.add(col)

        f1 = feed1_df[[c for c in feed1_df.columns if c in filtered_cols]]
        f2 = feed2_df[[c for c in feed2_df.columns if c in filtered_cols]]
    else:
        f1, f2 = feed1_df, feed2_df

    result_df = apply_full_range_advanced(
        df=measurement_df,
        small_df=f1,
        report_time=report_time,
        window_radius=window_radius,
        day_start_hour=18,
        input_value_at_start=input_value_at_start,
        big_df=f2,
        run_model_g=False,
    )
    return result_df


def generate_asset_traveler_reports(asset_id: str,
                                     df_15min_small: pd.DataFrame,
                                     df_15min_big: Optional[pd.DataFrame],
                                     measurement_path: str,
                                     report_time: datetime,
                                     lookback_days: int = DEFAULT_LOOKBACK_DAYS,
                                     window_radius: int = 600,
                                     process_non_priority: bool = False):
    """
    Generate Priority traveler report for one asset using its 15-min feeds.
    For RTY (composite), df_15min_big is None — the same feed is used for both slots.
    Returns (priority_df, non_priority_df, combined_df)
    """
    print(f"\n  📊 Generating traveler reports for {asset_id}...")

    feed1 = _preprocess_15min_feed(df_15min_small)
    feed2 = _preprocess_15min_feed(df_15min_big if df_15min_big is not None else df_15min_small)

    measurement_df = pd.read_excel(measurement_path)

    # Determine input @ start
    day_start_hour = 18
    base_dt = datetime(report_time.year, report_time.month, report_time.day, day_start_hour)
    if report_time < base_dt:
        base_dt -= timedelta(days=1)

    start_rows = feed1[feed1["time"] == base_dt]
    if not start_rows.empty:
        input_at_start = float(start_rows.iloc[-1]["open"])
    else:
        before = feed1[feed1["time"] <= base_dt]
        input_at_start = float(before.iloc[-1]["open"]) if not before.empty else None

    # Detect origins
    all_origins = set()
    for df_f in [feed1, feed2]:
        for col in df_f.columns:
            if col.endswith(" H") or " H[" in col:
                origin = col[:-2] if col.endswith(" H") else col.replace(" H[", "[")
                all_origins.add(origin)

    priority_origins_lower = {o.lower().strip() for o in PRIORITY_ORIGINS}
    priority_found    = {o for o in all_origins if o.lower().strip() in priority_origins_lower}
    non_priority_found = {o for o in all_origins if o.lower().strip() not in priority_origins_lower}

    # Generate priority report
    priority_df = pd.DataFrame()
    if priority_found:
        _result = _generate_travelers_with_calculator(
            measurement_df, feed1, feed2, report_time,
            window_radius, input_at_start, priority_found
        )
        priority_df = _result if _result is not None else pd.DataFrame()

    # Generate non-priority (optional)
    non_priority_df = pd.DataFrame()
    if process_non_priority and non_priority_found:
        _result = _generate_travelers_with_calculator(
            measurement_df, feed1, feed2, report_time,
            window_radius, input_at_start, non_priority_found
        )
        non_priority_df = _result if _result is not None else pd.DataFrame()

    # Combined
    parts = [d for d in [priority_df, non_priority_df] if not d.empty]
    combined_df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

    print(f"    ✅ Priority: {len(priority_df)} entries  |  "
          f"Non-priority: {len(non_priority_df)} entries  |  "
          f"Combined: {len(combined_df)}")

    return priority_df, non_priority_df, combined_df


def process_strategic_zones_for_asset(combined_df: pd.DataFrame,
                                       report_time: datetime,
                                       max_spread: float = 3.0,
                                       asset_id: str = ""):
    """Run 23 trading models (bypass mode) — identical to v3.0 logic."""
    if not HAS_MODEL_DEFINITIONS or not HAS_BYPASS_MATCHER:
        return None
    if combined_df.empty:
        return None

    print(f"\n  🎯 Strategic Zones for {asset_id} ({len(combined_df)} travelers, spread≤{max_spread})...")
    sorted_models = sorted(MODELS.items(), key=lambda x: x[1]["number"])

    all_results = {
        "results": {}, "timings": {},
        "metadata": {"report_time": report_time, "max_spread": max_spread, "asset": asset_id}
    }

    t0 = time_module.time()
    for idx, (model_name, model_def) in enumerate(sorted_models):
        tm0 = time_module.time()
        pct = (idx + 1) / len(sorted_models) * 100
        print(f"    [{pct:5.1f}%] #{model_def['number']:2d} {model_name:30s}...", end="")
        try:
            all_matches = []
            for day_filter in ["[0]", "[-1]", "[-2]", "[-3]"]:
                cfg = {"pass1_ms": sorted(model_def.get("pass1", set())),
                       "pass2_ms": sorted(model_def.get("pass2", set())),
                       "day_filter": day_filter, "feed_selection": "both"}
                m = process_model_bypass_mode(combined_df, cfg, max_spread=max_spread)
                if not m.empty:
                    all_matches.append(m)

            if all_matches:
                matched = pd.concat(all_matches, ignore_index=True)
                matched["M_#s"] = matched.apply(
                    lambda r: f"{int(r['M1'])}, {int(r['M2'])}", axis=1)
                all_results["results"][model_name] = matched
                print(f" {len(matched):4d} matches")
            else:
                all_results["results"][model_name] = pd.DataFrame()
                print(f"    0 matches")
        except Exception as e:
            print(f" ERROR: {str(e)[:50]}")
            all_results["results"][model_name] = pd.DataFrame()
        all_results["timings"][model_name] = time_module.time() - tm0

    all_results["total_time"] = time_module.time() - t0
    return all_results


# ============================================================================
# SECTION 8: WICK ANALYSIS
# ============================================================================

def _determine_wick_windows(report_time: datetime,
                              df_12min: pd.DataFrame) -> List[Dict]:
    """
    Derive wick analysis windows from actual chart bar timestamps.

    Instead of calendar arithmetic (which fails over weekends / when the 12-min
    chart ends before the report date), we scan backward in df_12min to locate
    the real prior session boundaries.

    Rules:
      - If report_time == 18:00 exactly:
            one window: [first bar of prior session, last bar before 18:00]
      - If report_time > 18:00:
            window 1: [first bar >= current session 18:00, report_time]
            window 2: [first bar of prior session, last bar before 18:00]
    """
    h, m = report_time.hour, report_time.minute
    session_open = report_time.replace(hour=18, minute=0, second=0, microsecond=0)

    tcol = "datetime_ny" if "datetime_ny" in df_12min.columns else "time"
    times = df_12min[tcol].apply(parse_timestamp_naive).sort_values().reset_index(drop=True)

    # ── Locate prior session ─────────────────────────────────────────────────
    # Last bar strictly before session_open
    bars_before = times[times < session_open]
    if bars_before.empty:
        prior_end = session_open - timedelta(minutes=12)
    else:
        prior_end = bars_before.iloc[-1]

    # Find where the prior session starts: walk back up to 7 days for the nearest 18:00
    prior_session_start = None
    for days_back in range(1, 8):
        candidate = (prior_end.to_pydatetime() - timedelta(days=days_back)).replace(
            hour=18, minute=0, second=0, microsecond=0)
        candidate = pd.Timestamp(candidate)
        bars_in_prior = times[(times >= candidate) & (times <= prior_end)]
        if not bars_in_prior.empty:
            prior_session_start = bars_in_prior.iloc[0]
            break

    if prior_session_start is None:
        prior_session_start = bars_before.iloc[0] if not bars_before.empty else pd.Timestamp(session_open)

    prior_window = {
        "label": "Prior session",
        "start": prior_session_start.to_pydatetime() if hasattr(prior_session_start, "to_pydatetime") else prior_session_start,
        "end":   prior_end.to_pydatetime() if hasattr(prior_end, "to_pydatetime") else prior_end,
    }

    if h == 18 and m == 0:
        return [prior_window]
    else:
        # Current session: first bar >= session_open up to report_time
        current_bars = times[(times >= pd.Timestamp(session_open)) & (times <= pd.Timestamp(report_time))]
        current_start = current_bars.iloc[0] if not current_bars.empty else pd.Timestamp(session_open)
        return [
            {"label": "Current session",
             "start": current_start.to_pydatetime() if hasattr(current_start, "to_pydatetime") else current_start,
             "end":   report_time},
            prior_window,
        ]


def _find_secondary_candle(df_12min: pd.DataFrame, swing: Dict) -> Optional[pd.Series]:
    """
    Find the secondary nested swing candle within a swing's time window.

    For a DOWN swing (origin = local high):
        After the swing's deepest low, find the highest-high candle (the counter-bounce).
    For an UP swing (origin = local low):
        After the swing's highest high, find the lowest-low candle (the counter-pullback).
    """
    origin_time  = swing["from_time"]
    extreme_time = swing["to_time"]
    direction    = swing["direction"]
    extreme_price = swing["to_price"]

    tcol = "datetime_ny" if "datetime_ny" in df_12min.columns else "time"
    mask = (df_12min[tcol] > origin_time) & (df_12min[tcol] <= extreme_time)
    window = df_12min[mask].copy().reset_index(drop=True)

    if len(window) < 2:
        return None

    if direction == "Down":
        # Find the row with the minimum low (closest to the extreme)
        low_idx = window["low"].idxmin()
        low_time = window.loc[low_idx, tcol]
        # Look for counter-bounce AFTER the deepest low
        after_low = window[window[tcol] > low_time]
        if after_low.empty:
            return None
        return after_low.loc[after_low["high"].idxmax()]

    else:  # Up swing
        high_idx = window["high"].idxmax()
        high_time = window.loc[high_idx, tcol]
        after_high = window[window[tcol] > high_time]
        if after_high.empty:
            return None
        return after_high.loc[after_high["low"].idxmin()]


def _wick_ranges(candle: pd.Series) -> Dict[str, Optional[Tuple[float, float]]]:
    """
    Return upper and lower wick price ranges for a candle.
    Upper wick: (max(open,close), high)  — only if high > max(open,close)
    Lower wick: (low, min(open,close))   — only if low < min(open,close)
    Returns None for a wick if it is zero.
    """
    o, h, lo, c = candle["open"], candle["high"], candle["low"], candle["close"]
    body_top    = max(o, c)
    body_bottom = min(o, c)

    upper = (body_top, h)  if h  > body_top    else None
    lower = (lo, body_bottom) if lo < body_bottom else None
    return {"upper": upper, "lower": lower}


def _value_in_range(value: float, rng: Tuple[float, float]) -> bool:
    return rng[0] <= value <= rng[1]


def _find_mas_in_wick(candle: pd.Series,
                       ma_cols: List[str],
                       wick_rng: Tuple[float, float]) -> List[str]:
    """Return list of MA column names whose value for this candle falls in the wick range."""
    hits = []
    for col in ma_cols:
        val = candle.get(col, np.nan)
        if pd.isna(val):
            continue
        if _value_in_range(float(val), wick_rng):
            hits.append(col)
    return hits


def _find_travelers_in_wick(travelers_df: pd.DataFrame,
                              wick_rng: Tuple[float, float]) -> pd.DataFrame:
    """
    Return subset of travelers_df whose Output price falls within wick_rng.
    Looks for column 'Output' or 'Output1'.
    """
    if travelers_df is None or travelers_df.empty:
        return pd.DataFrame()

    out_col = "Output" if "Output" in travelers_df.columns else \
              "Output1" if "Output1" in travelers_df.columns else None
    if out_col is None:
        return pd.DataFrame()

    mask = travelers_df[out_col].apply(
        lambda v: not pd.isna(v) and _value_in_range(float(v), wick_rng)
    )
    return travelers_df[mask].copy()


def run_wick_analysis_for_asset(asset_id: str,
                                 df_12min: pd.DataFrame,
                                 travelers_df: pd.DataFrame,
                                 report_time: datetime,
                                 ma_cols: List[str]) -> pd.DataFrame:
    """
    Run full wick analysis for one asset using its 12-min chart.
    Returns a DataFrame of wick analysis results.
    """
    if not HAS_SWING_DETECTOR or df_12min.empty:
        return pd.DataFrame()

    print(f"\n  🕯️  Wick analysis: {asset_id}  (12-min chart, {len(df_12min)} bars)")

    windows = _determine_wick_windows(report_time)
    all_rows = []

    tcol = "datetime_ny" if "datetime_ny" in df_12min.columns else "time"
    df_12min = df_12min.copy()
    df_12min[tcol] = df_12min[tcol].apply(parse_timestamp_naive)

    for win in windows:
        w_start, w_end = win["start"], win["end"]
        print(f"    Window [{win['label']}]: {w_start} → {w_end}")

        # Slice to window
        mask = (df_12min[tcol] >= w_start) & (df_12min[tcol] <= w_end)
        df_window = df_12min[mask].copy().reset_index(drop=True)

        if df_window.empty:
            print(f"      ⚠️  No 12-min data in this window — skipping.")
            continue

        # Detect nested swings within the window
        swings = detect_nested_swings(df_window.rename(columns={tcol: "time"}),
                                       min_swing_size=10,
                                       pullback_tolerance=5)

        if not swings:
            print(f"      ℹ️  No nested swings detected.")
            continue

        print(f"      Found {len(swings)} nested swings.")

        for swing in swings:
            # --- PRIMARY candle ---
            primary_time = parse_timestamp_naive(swing["from_time"])
            primary_mask = df_12min[tcol].apply(
                lambda t: parse_timestamp_naive(t) == primary_time)
            primary_rows = df_12min[primary_mask]
            if primary_rows.empty:
                continue
            primary_candle = primary_rows.iloc[0]

            # Determine role: swing direction tells us the role of MAs/travelers in wick
            # DOWN swing → high wick = resistance; UP swing → low wick = support
            role_upper = "Resistance" if swing["direction"] == "Down" else "Support"
            role_lower = "Support"    if swing["direction"] == "Down" else "Resistance"

            wicks = _wick_ranges(primary_candle)
            for wick_side, wick_rng in [("Upper", wicks["upper"]),
                                         ("Lower", wicks["lower"])]:
                if wick_rng is None:
                    continue
                role = role_upper if wick_side == "Upper" else role_lower

                mas_hit = _find_mas_in_wick(primary_candle, ma_cols, wick_rng)
                travelers_hit = _find_travelers_in_wick(travelers_df, wick_rng)

                if not mas_hit and travelers_hit.empty:
                    continue

                for ma in mas_hit:
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": primary_time,
                        "Candle_Type": "Primary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "MA",
                        "Name":        ma,
                        "Value":       round(float(primary_candle.get(ma, np.nan)), 4),
                        "Candle_O":    primary_candle["open"],
                        "Candle_H":    primary_candle["high"],
                        "Candle_L":    primary_candle["low"],
                        "Candle_C":    primary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

                for _, trav in travelers_hit.iterrows():
                    origin  = trav.get("Origin", "")
                    m_num   = trav.get("M #", trav.get("M1", ""))
                    out_val = trav.get("Output", trav.get("Output1", np.nan))
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": primary_time,
                        "Candle_Type": "Primary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "Traveler",
                        "Name":        f"{origin} M#{m_num}",
                        "Value":       round(float(out_val), 4) if not pd.isna(out_val) else np.nan,
                        "Candle_O":    primary_candle["open"],
                        "Candle_H":    primary_candle["high"],
                        "Candle_L":    primary_candle["low"],
                        "Candle_C":    primary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

            # --- SECONDARY candle ---
            tmp_df = df_12min.copy()
            if tcol != "datetime_ny":
                tmp_df = tmp_df.rename(columns={tcol: "datetime_ny"})
            secondary_candle = _find_secondary_candle(tmp_df, swing)

            if secondary_candle is None:
                continue

            sec_time = parse_timestamp_naive(
                secondary_candle.get("datetime_ny", secondary_candle.get("time", None)))
            wicks_s = _wick_ranges(secondary_candle)

            for wick_side, wick_rng in [("Upper", wicks_s["upper"]),
                                          ("Lower", wicks_s["lower"])]:
                if wick_rng is None:
                    continue
                # Secondary wick role: secondary is a counter-move that fails
                # For DOWN swing secondary (bounce high) → upper wick = resistance
                # For UP swing secondary (pullback low) → lower wick = support
                if swing["direction"] == "Down":
                    role = "Resistance" if wick_side == "Upper" else "Support"
                else:
                    role = "Support" if wick_side == "Lower" else "Resistance"

                mas_hit = _find_mas_in_wick(secondary_candle, ma_cols, wick_rng)
                travelers_hit = _find_travelers_in_wick(travelers_df, wick_rng)

                if not mas_hit and travelers_hit.empty:
                    continue

                for ma in mas_hit:
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": sec_time,
                        "Candle_Type": "Secondary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "MA",
                        "Name":        ma,
                        "Value":       round(float(secondary_candle.get(ma, np.nan)), 4),
                        "Candle_O":    secondary_candle["open"],
                        "Candle_H":    secondary_candle["high"],
                        "Candle_L":    secondary_candle["low"],
                        "Candle_C":    secondary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

                for _, trav in travelers_hit.iterrows():
                    origin  = trav.get("Origin", "")
                    m_num   = trav.get("M #", trav.get("M1", ""))
                    out_val = trav.get("Output", trav.get("Output1", np.nan))
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": sec_time,
                        "Candle_Type": "Secondary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "Traveler",
                        "Name":        f"{origin} M#{m_num}",
                        "Value":       round(float(out_val), 4) if not pd.isna(out_val) else np.nan,
                        "Candle_O":    secondary_candle["open"],
                        "Candle_H":    secondary_candle["high"],
                        "Candle_L":    secondary_candle["low"],
                        "Candle_C":    secondary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

    result_df = pd.DataFrame(all_rows)
    print(f"    ✅ {asset_id} wick analysis: {len(result_df)} wick intersections found.")
    return result_df


# ============================================================================
# SECTION 9A: MATCHED TRAVELERS — COLUMN DERIVATION & STANDARDISED FORMAT
# ============================================================================
# Exact 29-column order that matches the reference MatchedTravelers file
MATCHED_COL_ORDER = [
    "Open", "Output1", "Output2", "Prox", "Origin1", "Origin2", "Group",
    "M1", "M2", "R1", "R2", "M_#s", "Arrival_Order", "Match",
    "Tag1", "Tag2", "Family1", "Family2", "Families",
    "Arrival1", "Arrival2", "Day1", "Day2", "Arrival_Brackets",
    "Model", "Model_Number", "Recent", "Feed1", "Feed2",
]

_ANCHOR_SET = {"spain", "saturn", "jupiter", "kepler-62", "kepler-44"}
_TT_SET     = {"trinidad", "tobago"}


def _build_matched_combined_df(sz_results: Dict,
                                 priority_df: pd.DataFrame,
                                 measurement_df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    """
    Concatenate all model match DataFrames, derive every standardised column,
    reorder to MATCHED_COL_ORDER, and return the finished DataFrame.
    Mirrors the full derivation logic from v3.0 export_split_reports_to_excel.
    """
    all_matched = []
    for mname, mdf in sz_results.get("results", {}).items():
        if mdf.empty:
            continue
        d = mdf.copy()
        model_info = MODELS.get(mname, {})
        if "Model" not in d.columns:
            d["Model"] = model_info.get("display_name", mname)
            d["Model_Number"] = model_info.get("number", 0)
        all_matched.append(d)

    if not all_matched:
        return None

    df = pd.concat(all_matched, ignore_index=True)

    # Rename Input→Output aliases if needed
    if "Input1" in df.columns and "Output1" not in df.columns:
        df["Output1"] = df["Input1"]
    if "Input2" in df.columns and "Output2" not in df.columns:
        df["Output2"] = df["Input2"]

    # Sort by Output1 descending
    if "Output1" in df.columns:
        df = df.sort_values("Output1", ascending=False).reset_index(drop=True)

    # ── Build Tag/Family lookup from RawTraveler priority_df ──────────────
    origin_m_to_tag    = {}
    origin_m_to_family = {}
    if priority_df is not None and not priority_df.empty:
        if "Tag" in priority_df.columns and "Family" in priority_df.columns:
            for _, row in priority_df.iterrows():
                o  = row.get("Origin", "")
                mn = row.get("M #", None)
                if o and mn is not None:
                    try:
                        key = (str(o), int(float(mn)))
                        origin_m_to_tag[key]    = row.get("Tag", "")
                        origin_m_to_family[key] = row.get("Family", "")
                    except Exception:
                        pass

    # M# → Tag fallback from measurement file
    m_to_tag = {}
    if measurement_df is not None and not measurement_df.empty:
        m_col   = next((c for c in measurement_df.columns
                        if c.lower().replace(" ", "") in ["m#", "m", "mnumber"]), None)
        tag_col = next((c for c in measurement_df.columns if "tag" in c.lower()), None)
        if m_col and tag_col:
            for _, row in measurement_df.iterrows():
                try:
                    m_to_tag[int(float(row[m_col]))] = str(row[tag_col])
                except Exception:
                    pass

    # ── Group ─────────────────────────────────────────────────────────────
    def _group(row):
        o1 = str(row.get("Origin1", "")).lower()
        o2 = str(row.get("Origin2", "")).lower()
        if not o1 or not o2:
            return ""
        a1, a2 = o1 in _ANCHOR_SET, o2 in _ANCHOR_SET
        t1, t2 = o1 in _TT_SET,     o2 in _TT_SET
        if a1 and a2 and o1 == o2: return "1. SAA"
        if t1 and t2 and o1 == o2: return "2. STT"
        if t1 and t2:              return "2. TT"
        if t1 and a2:              return "3. TA"
        if a1 and t2:              return "3. AT"
        if a1 and a2:              return "4. AA"
        if a2 and not a1:          return "5. oA"
        if a1 and not a2:          return "6. Ao"
        return "7. oo"

    df["Group"] = df.apply(_group, axis=1)

    # ── Arrival_Order ─────────────────────────────────────────────────────
    def _arr_order(row):
        try:
            m1a, m2a = abs(int(float(row["M1"]))), abs(int(float(row["M2"])))
            if m1a < 40 and m2a >= 40: return "PD"
            if m1a >= 40 and m2a < 40: return "DP"
            if m1a < 40 and m2a < 40:  return "DD"
            return "PP"
        except Exception:
            return ""

    df["Arrival_Order"] = df.apply(_arr_order, axis=1)

    # ── Match ─────────────────────────────────────────────────────────────
    def _match(row):
        try:
            m1, m2 = int(float(row["M1"])), int(float(row["M2"]))
            sign = "same" if (m1 > 0) == (m2 > 0) else "flip"
            m1a, m2a = abs(m1), abs(m2)
            if m1a > m2a:   direct = "up"
            elif m1a < m2a: direct = "down"
            else:            direct = "OPP" if sign == "flip" else "same"
            return f"{sign} {direct}"
        except Exception:
            return ""

    df["Match"] = df.apply(_match, axis=1)

    # ── Tag1 / Tag2 ───────────────────────────────────────────────────────
    def _tag(origin, m_val):
        if pd.isna(m_val): return ""
        m = int(float(m_val))
        key = (str(origin), m)
        if key in origin_m_to_tag:    return origin_m_to_tag[key]
        if m in m_to_tag:             return m_to_tag[m]
        m_abs = abs(m)
        if m_abs == 0:                return "Green"
        if m_abs in {40}:             return "Strength"
        if m_abs in {2,10,22,30,36,39,41,43,50,60,77,107}: return "Alpha"
        if m_abs in {5,14,55,68,96}:  return "Bravo"
        return "Other"

    df["Tag1"] = df.apply(lambda r: _tag(r.get("Origin1",""), r.get("M1")), axis=1)
    df["Tag2"] = df.apply(lambda r: _tag(r.get("Origin2",""), r.get("M2")), axis=1)

    # ── Family1 / Family2 / Families ─────────────────────────────────────
    def _family(origin, m_val):
        if pd.isna(m_val): return ""
        m = int(float(m_val))
        key = (str(origin), m)
        if key in origin_m_to_family: return origin_m_to_family[key]
        m_abs = abs(m)
        if m_abs >= 96:             return "1A"
        if 50 <= m_abs < 96:        return "1B"
        if 20 <= m_abs < 50:        return "2A"
        return "2B"

    df["Family1"]  = df.apply(lambda r: _family(r.get("Origin1",""), r.get("M1")), axis=1)
    df["Family2"]  = df.apply(lambda r: _family(r.get("Origin2",""), r.get("M2")), axis=1)
    df["Families"] = df["Family1"] + ", " + df["Family2"]

    # ── Recent ────────────────────────────────────────────────────────────
    def _recent(row):
        d1 = str(row.get("Day1", ""))
        if "[0]" in d1:  return "Today"
        if "["   in d1:  return "Recent"
        return ""

    df["Recent"] = df.apply(_recent, axis=1)

    # ── Open (within 3 pts of Input @ Start from priority_df) ─────────────
    df["Open"] = None
    if priority_df is not None and not priority_df.empty:
        input_col = next((c for c in priority_df.columns
                          if "input" in c.lower() and "start" in c.lower()), None)
        if input_col:
            starts = []
            if "Feed" in priority_df.columns:
                for feed_type in ("Small", "Big"):
                    rows = priority_df[priority_df["Feed"] == feed_type]
                    vals = rows[input_col].dropna()
                    if not vals.empty:
                        starts.append(float(vals.iloc[0]))
            else:
                vals = priority_df[input_col].dropna().unique()
                starts = sorted(float(v) for v in vals)

            if starts:
                for i, row in df.iterrows():
                    out = row.get("Output1") or row.get("Arrival_Output")
                    if out is None:
                        continue
                    try:
                        out = float(out)
                    except Exception:
                        continue
                    for sv in starts:
                        if abs(out - sv) <= 3:
                            df.at[i, "Open"] = sv
                            break

    # ── Drop temporary / unwanted columns ────────────────────────────────
    for drop_col in ("Input1", "Input2", "Arrival_Output"):
        if drop_col in df.columns:
            df = df.drop(columns=[drop_col])

    # ── Reorder to standard 29-column layout ─────────────────────────────
    ordered   = [c for c in MATCHED_COL_ORDER if c in df.columns]
    remaining = [c for c in df.columns if c not in ordered]
    df = df[ordered + remaining]

    return df


def _write_matched_travelers_excel(df: pd.DataFrame,
                                    fpath: str,
                                    report_time: datetime,
                                    asset_id: str):
    """
    Write the standardised MatchedTravelers Excel file:
      Row 1 — Report Time info
      Row 2 — match count info
      Row 3 — header (amber fill FFC000, black bold)
      Row 4+ — data
      Freeze B4, auto-filter on row 3, all columns 13.0 wide.
    """
    # Write data first (no info rows yet)
    df.to_excel(fpath, index=False, sheet_name="Combined")

    wb = load_workbook(fpath)
    ws = wb["Combined"]

    # Insert 2 info rows at the top
    ws.insert_rows(1, 2)
    ws["A1"] = f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"All Models Combined | {len(df)} total matches"
    ws["A2"].font = Font(italic=True)

    # Header row is now row 3 — apply amber formatting
    amber_fill  = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    black_bold  = Font(bold=True, color="000000")
    center      = Alignment(horizontal="center", vertical="center")
    for cell in ws[3]:
        cell.fill      = amber_fill
        cell.font      = black_bold
        cell.alignment = center

    # Freeze B4 (matches reference file)
    ws.freeze_panes = "B4"

    # Auto-filter on row 3
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Fixed column width 13.0 for all columns (matches reference)
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13.0

    # Format Arrival1 / Arrival2 cells as datetime
    dt_fmt = "YYYY-MM-DD HH:MM:SS"
    arr_cols = [i for i, cell in enumerate(ws[3], start=1)
                if str(cell.value) in ("Arrival1", "Arrival2")]
    for col_idx in arr_cols:
        letter = get_column_letter(col_idx)
        for row in range(4, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = dt_fmt

    wb.save(fpath)


# ============================================================================
# SECTION 9: EXCEL EXPORT
# ============================================================================

HEADER_FILLS = {
    "NQ":  "4472C4",   # blue
    "ES":  "70AD47",   # green
    "RTY": "FFC000",   # amber
    "wick": "C00000",  # red
}


def _fmt_ws_header(ws, header_color: str = "4472C4", freeze: str = "A2"):
    """Apply blue header formatting to row 1 of a worksheet."""
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = freeze
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _auto_col_width(ws, max_width: int = 40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(best + 2, max_width)


def export_all_results(asset_results: Dict,
                        wick_results: Dict[str, pd.DataFrame],
                        report_time: datetime,
                        output_dir: str = ".") -> Dict[str, str]:
    """
    Write Excel outputs:
      - RawTraveler_Priority_{ASSET}_{TS}.xlsx   (per asset)
      - MatchedTravelers_{ASSET}_{TS}.xlsx        (per asset)
      - WickAnalysis_{TS}.xlsx                    (all assets, one tab each)
    """
    print("\n" + "=" * 120)
    print("📥 EXPORTING TO EXCEL")
    print("=" * 120)

    ts = report_time.strftime("%Y%m%d_%H%M")
    created = {}

    # ------------------------------------------------------------------ #
    # Per-asset Raw + Matched
    # ------------------------------------------------------------------ #
    for asset_id, data in asset_results.items():
        priority_df  = data.get("priority_df",  pd.DataFrame())
        sz_results   = data.get("sz_results",   None)
        color        = HEADER_FILLS.get(asset_id, "4472C4")

        # --- RawTraveler Priority ---
        if not priority_df.empty:
            fname = f"RawTraveler_Priority_{asset_id}_{ts}.xlsx"
            fpath = os.path.join(output_dir, fname)
            try:
                priority_df.to_excel(fpath, index=False, sheet_name="Priority_Origins")
                wb = load_workbook(fpath)
                ws = wb["Priority_Origins"]
                _fmt_ws_header(ws, color)
                _auto_col_width(ws)
                wb.save(fpath)
                created[f"RawTraveler_{asset_id}"] = fpath
                print(f"  ✅ {fname}  ({len(priority_df)} rows)")
            except Exception as e:
                print(f"  ❌ Error writing {fname}: {e}")

        # --- MatchedTravelers (standardised v3.0 format) ---
        if sz_results and sz_results.get("results"):
            fname = f"MatchedTravelers_{asset_id}_{ts}.xlsx"
            fpath = os.path.join(output_dir, fname)
            try:
                combined = _build_matched_combined_df(
                    sz_results, priority_df, data.get("measurement_df"))
                if combined is not None and not combined.empty:
                    _write_matched_travelers_excel(
                        combined, fpath, report_time, asset_id)
                    created[f"Matched_{asset_id}"] = fpath
                    print(f"  ✅ {fname}  ({len(combined)} matched rows)")
                else:
                    print(f"  ℹ️  {asset_id}: no matched travelers to export.")
            except Exception as e:
                print(f"  ❌ Error writing {fname}: {e}")
                import traceback; traceback.print_exc()

    # ------------------------------------------------------------------ #
    # WickAnalysis (all assets, one tab per asset)
    # ------------------------------------------------------------------ #
    has_wick = any(not df.empty for df in wick_results.values())
    if has_wick:
        fname = f"WickAnalysis_{ts}.xlsx"
        fpath = os.path.join(output_dir, fname)
        try:
            with pd.ExcelWriter(fpath, engine="openpyxl") as writer:
                for asset_id, wdf in wick_results.items():
                    if wdf.empty:
                        continue
                    sheet_name = f"{asset_id}_Wick"
                    wdf.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply formatting
            wb = load_workbook(fpath)
            for asset_id, wdf in wick_results.items():
                if wdf.empty:
                    continue
                ws = wb[f"{asset_id}_Wick"]
                color = HEADER_FILLS.get(asset_id, "C00000")
                _fmt_ws_header(ws, color)
                _auto_col_width(ws)

                # Highlight Resistance rows orange, Support rows green
                res_fill  = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                sup_fill  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

                role_col_idx = None
                for i, cell in enumerate(ws[1], start=1):
                    if str(cell.value) == "Role":
                        role_col_idx = i
                        break

                if role_col_idx:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        role_val = row[role_col_idx - 1].value
                        fill = res_fill if role_val == "Resistance" else \
                               sup_fill if role_val == "Support" else None
                        if fill:
                            for cell in row:
                                cell.fill = fill

            wb.save(fpath)
            total_wick_rows = sum(len(d) for d in wick_results.values())
            created["WickAnalysis"] = fpath
            print(f"  ✅ {fname}  ({total_wick_rows} wick intersection rows, "
                  f"{sum(1 for d in wick_results.values() if not d.empty)} asset tabs)")
        except Exception as e:
            print(f"  ❌ Error writing {fname}: {e}")
            import traceback; traceback.print_exc()
    else:
        print("  ℹ️  No wick intersections found — WickAnalysis.xlsx skipped.")

    return created


# ============================================================================
# SECTION 10: MAIN PIPELINE ORCHESTRATOR
# ============================================================================

def parse_report_datetime(s: str) -> datetime:
    for fmt in ["%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M", "%Y/%m/%d %H:%M"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse datetime: '{s}'.  Use YYYY-MM-DD HH:MM")


def run_pipeline_v4(search_dir: str,
                    measurement_path: str,
                    report_datetime_str: str,
                    lookback_days: int = DEFAULT_LOOKBACK_DAYS,
                    output_dir: str = ".",
                    process_non_priority: bool = False,
                    per_asset_overrides: Optional[Dict] = None) -> Optional[Dict]:
    """
    Full Pipeline v4.0 execution.

    Steps:
      1.  File discovery & validation
      2.  Load + normalize timestamps (all 15 CSVs)
      3.  Multi-chart construction + QC per asset
      4.  MA normalization per asset
      5.  Raw + Matched traveler reports per asset
      6.  Strategic zones per asset
      7.  Wick analysis per asset (12-min chart)
      8.  Excel export
    """
    print("\n" + "=" * 120)
    print("🚀 PIPELINE v4.0 — EXECUTION START")
    print("=" * 120)
    print(f"  Started: {datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"  Search dir:  {search_dir}")
    print(f"  Measurement: {measurement_path}")

    # ── 1. File discovery ───────────────────────────────────────────────
    discovered, _ = discover_feed_files(search_dir)
    if not validate_file_set(discovered, measurement_path):
        return None

    # ── Parse report time ───────────────────────────────────────────────
    report_time = parse_report_datetime(report_datetime_str)
    print(f"\n  📅 Report time: {report_time:%Y-%m-%d %H:%M}")

    # ── 2. Load & normalize all CSVs ────────────────────────────────────
    print("\n" + "─" * 80)
    print("📂 STEP 2: LOADING AND NORMALIZING TIMESTAMPS")
    print("─" * 80)

    loaded: Dict[str, Dict[int, pd.DataFrame]] = {}
    for asset in ["NQ", "ES", "RTY"]:
        loaded[asset] = {}
        for tf in [15, 3, 5]:
            info = discovered[asset][tf]
            print(f"\n  Loading {asset} {tf}-min: {info['path'].name}")
            df = load_and_normalize_csv(info["path"])
            loaded[asset][tf] = df
            print(f"    {len(df)} rows loaded.")

    # ── 3. Multi-chart construction ─────────────────────────────────────
    print("\n" + "─" * 80)
    print("📊 STEP 3: MULTI-CHART OHLC CONSTRUCTION")
    print("─" * 80)

    charts: Dict[str, Dict[int, pd.DataFrame]] = {}
    for asset in ["NQ", "ES", "RTY"]:
        print(f"\n  🔧 {asset}")
        df3 = loaded[asset][3]
        df5 = loaded[asset][5]
        df15 = loaded[asset][15]

        open_hr = detect_session_open_hour(df3)
        print(f"    Session open detected: {open_hr}:00")

        charts[asset] = build_all_higher_timeframes(
            df3, df5, df15, asset,
            session_open_hour=open_hr,
            lookback_days=lookback_days,
        )

    # ── 4. MA normalization ─────────────────────────────────────────────
    print("\n" + "─" * 80)
    print("〰️  STEP 4: MOVING AVERAGE NORMALIZATION")
    print("─" * 80)

    normalized_ma_cols: Dict[str, List[str]] = {}  # asset → list of canonical MA col names
    for asset in ["NQ", "ES", "RTY"]:
        print(f"\n  {asset} — 3-min feed")
        df3_norm, name_map = normalize_ma_columns(loaded[asset][3])
        loaded[asset][3] = df3_norm   # store normalized version

        # Also normalize the 12-min chart (built from 3-min) so wick analysis has MA values
        if 12 in charts[asset] and not charts[asset][12].empty:
            df12_norm, _ = normalize_ma_columns(charts[asset][12])
            charts[asset][12] = df12_norm

        normalized_ma_cols[asset] = get_all_ma_column_names(df3_norm)
        h_series = sorted([c for c in normalized_ma_cols[asset] if c.startswith("h")])
        other_ma = sorted([c for c in normalized_ma_cols[asset] if not c.startswith("h")])
        print(f"    h-series ({len(h_series)}): {', '.join(h_series[:12])}"
              f"{'...' if len(h_series) > 12 else ''}")
        print(f"    Other MAs ({len(other_ma)}): {', '.join(other_ma[:6])}"
              f"{'...' if len(other_ma) > 6 else ''}")

    # ── 5 & 6. Traveler reports + Strategic zones per asset ─────────────
    print("\n" + "─" * 80)
    print("📈 STEPS 5–6: RAW TRAVELERS + STRATEGIC ZONES (per asset)")
    print("─" * 80)

    asset_results: Dict[str, Dict] = {}

    for asset in ["NQ", "ES", "RTY"]:
        defaults = ASSET_DEFAULTS[asset].copy()
        if per_asset_overrides and asset in per_asset_overrides:
            defaults.update(per_asset_overrides[asset])

        max_spread   = defaults["max_spread"]
        radius       = defaults["radius"]
        is_composite = defaults["composite"]

        print(f"\n{'='*60}")
        print(f"  ASSET: {asset}  {'[COMPOSITE — single-feed RTY]' if is_composite else ''}")
        print(f"  max-spread={max_spread}  radius={radius}")
        print(f"{'='*60}")

        df_15 = loaded[asset][15]
        # For RTY composite, small == big (same feed used for both slots)
        df_15_small = df_15
        df_15_big   = None if is_composite else df_15  # non-composite: same file for now
        # NOTE: For NQ, the "small" feed is MNQ; "big" feed is NQ.
        #       The discover step records role per file. For v4.0 the two 15-min files
        #       per NQ/ES are loaded via the same asset key; we need to handle this correctly.
        # The loaded[asset][15] currently holds whichever file was found last.
        # Since only one 15-min file per (asset, tf) is stored, we need to load both.

        # Actually, v3.0 used separate feed1/feed2 (small/big). For v4.0, the 15-min
        # Origin files are 2 per NQ and 2 per ES (MNQ + NQ, MES + ES), but 1 for RTY.
        # We need to retrieve them by role.
        if asset != "RTY":
            # find small and big
            info_small = discovered[asset].get(15)  # currently only stores 1; need to extend
            # The _classify_csv function should have stored both — let's check
            # We need to re-discover separately for small vs big
            small_candidates = [
                p for p in Path(search_dir).glob("*.csv")
                if _classify_csv(p) and
                   _classify_csv(p)["asset"] == asset and
                   _classify_csv(p)["timeframe"] == 15 and
                   _classify_csv(p)["role"] == "small"
            ]
            big_candidates = [
                p for p in Path(search_dir).glob("*.csv")
                if _classify_csv(p) and
                   _classify_csv(p)["asset"] == asset and
                   _classify_csv(p)["timeframe"] == 15 and
                   _classify_csv(p)["role"] == "big"
            ]
            if small_candidates:
                print(f"  📄 Small feed: {small_candidates[0].name}")
                df_15_small = load_and_normalize_csv(small_candidates[0])
            if big_candidates:
                print(f"  📄 Big feed:   {big_candidates[0].name}")
                df_15_big = load_and_normalize_csv(big_candidates[0])
            if not small_candidates or not big_candidates:
                print(f"  ⚠️  Could not separate small/big 15-min feeds for {asset} — using single feed.")
                df_15_small = df_15
                df_15_big = df_15

        priority_df, non_priority_df, combined_df = generate_asset_traveler_reports(
            asset_id=asset,
            df_15min_small=df_15_small,
            df_15min_big=df_15_big,
            measurement_path=measurement_path,
            report_time=report_time,
            lookback_days=lookback_days,
            window_radius=radius,
            process_non_priority=process_non_priority,
        )

        sz_results = process_strategic_zones_for_asset(
            combined_df, report_time, max_spread=max_spread, asset_id=asset
        )

        asset_results[asset] = {
            "priority_df":     priority_df,
            "non_priority_df": non_priority_df,
            "combined_df":     combined_df,
            "sz_results":      sz_results,
            "measurement_df":  pd.read_excel(measurement_path) if os.path.exists(measurement_path) else None,
        }

    # ── 7. Wick analysis ────────────────────────────────────────────────
    print("\n" + "─" * 80)
    print("🕯️  STEP 7: WICK ANALYSIS")
    print("─" * 80)

    wick_results: Dict[str, pd.DataFrame] = {}

    for asset in ["NQ", "ES", "RTY"]:
        df_12 = charts[asset].get(12, pd.DataFrame())
        if df_12.empty:
            print(f"  ⚠️  {asset}: no 12-min chart available — skipping wick analysis.")
            wick_results[asset] = pd.DataFrame()
            continue

        travelers_df = asset_results[asset].get("priority_df", pd.DataFrame())

        # Use 12-min MA cols (post-normalization if available)
        ma_cols_12 = get_all_ma_column_names(df_12)

        wick_df = run_wick_analysis_for_asset(
            asset_id=asset,
            df_12min=df_12,
            travelers_df=travelers_df,
            report_time=report_time,
            ma_cols=ma_cols_12,
        )
        wick_results[asset] = wick_df

    # ── 8. Export ────────────────────────────────────────────────────────
    print("\n" + "─" * 80)
    print("📥 STEP 8: EXPORT")
    print("─" * 80)

    created_files = export_all_results(
        asset_results=asset_results,
        wick_results=wick_results,
        report_time=report_time,
        output_dir=output_dir,
    )

    # ── Summary ──────────────────────────────────────────────────────────
    print("\n" + "=" * 120)
    print("✅ PIPELINE v4.0 COMPLETE")
    print("=" * 120)
    print(f"  Finished: {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("\n  📁 Created files:")
    for key, path in created_files.items():
        print(f"    {key:<30s}: {path}")
    print()

    return {
        "asset_results": asset_results,
        "wick_results":  wick_results,
        "charts":        charts,
        "created_files": created_files,
    }


# ============================================================================
# SECTION 11: COMMAND-LINE INTERFACE
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Unified Traveler Pipeline v4.0 — NQ + ES + RTY multi-timeframe",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic (auto-discovers files in current directory)
  python unified_traveler_pipeline_4_0.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-02-23 18:30" \\
      --lookback 22

  # Specify search directory
  python unified_traveler_pipeline_4_0.py \\
      --dir "C:/TradingData/2026-03-10" \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-10 18:00"

  # Override per-asset defaults
  python unified_traveler_pipeline_4_0.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-10 18:30" \\
      --max-spread-nq 3.0 --radius-nq 1000 \\
      --max-spread-es 1.0 --radius-es 215 \\
      --max-spread-rty 1.0 --radius-rty 100

Required Files (16 total, auto-discovered in --dir):
  15-min (5):  CME_MINI_MNQ*, CME_MINI_NQ*, CME_MINI_MES*, CME_MINI_ES*, CME_MINI_RTY*  with ", 15"
   3-min (5):  same assets with ", 3"
   5-min (5):  same assets with ", 5"
  + 1 measurement Excel file (specified with --measurement)

Per-Asset Defaults:
  NQ:  max-spread=3.0  radius=1000
  ES:  max-spread=1.0  radius=215
  RTY: max-spread=1.0  radius=100  (composite feed — no MRTY)
""",
    )

    # Required
    parser.add_argument("--measurement", "-m", required=True,
                        help="Path to measurement Excel file")
    parser.add_argument("--dt", "-dt", required=True,
                        help="Report datetime  (YYYY-MM-DD HH:MM)")

    # Optional - directory
    parser.add_argument("--dir", "-d", default=".",
                        help="Directory to scan for feed CSVs (default: current dir)")
    parser.add_argument("--output-dir", "-o", default=".",
                        help="Output directory for Excel files (default: current dir)")

    # Processing options
    parser.add_argument("--lookback", "-l", type=int, default=DEFAULT_LOOKBACK_DAYS,
                        help=f"Lookback days (default: {DEFAULT_LOOKBACK_DAYS})")
    parser.add_argument("--process-non-priority", action="store_true",
                        help="Also process non-priority origins (slow)")

    # Per-asset overrides
    parser.add_argument("--max-spread-nq",  type=float, default=None)
    parser.add_argument("--max-spread-es",  type=float, default=None)
    parser.add_argument("--max-spread-rty", type=float, default=None)
    parser.add_argument("--radius-nq",      type=int,   default=None)
    parser.add_argument("--radius-es",      type=int,   default=None)
    parser.add_argument("--radius-rty",     type=int,   default=None)

    args = parser.parse_args()

    # Build override dict
    overrides = {}
    if args.max_spread_nq is not None:
        overrides.setdefault("NQ", {})["max_spread"] = args.max_spread_nq
    if args.max_spread_es is not None:
        overrides.setdefault("ES", {})["max_spread"] = args.max_spread_es
    if args.max_spread_rty is not None:
        overrides.setdefault("RTY", {})["max_spread"] = args.max_spread_rty
    if args.radius_nq is not None:
        overrides.setdefault("NQ", {})["radius"] = args.radius_nq
    if args.radius_es is not None:
        overrides.setdefault("ES", {})["radius"] = args.radius_es
    if args.radius_rty is not None:
        overrides.setdefault("RTY", {})["radius"] = args.radius_rty

    # Run
    result = run_pipeline_v4(
        search_dir=args.dir,
        measurement_path=args.measurement,
        report_datetime_str=args.dt,
        lookback_days=args.lookback,
        output_dir=args.output_dir,
        process_non_priority=args.process_non_priority,
        per_asset_overrides=overrides if overrides else None,
    )

    return 0 if result else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
