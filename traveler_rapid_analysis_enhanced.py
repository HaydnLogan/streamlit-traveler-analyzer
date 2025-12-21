"""
Traveler Rapid Analysis Dashboard - Enhanced Version
Processes raw data grabs and identifies high-priority trading zones with:
- Direct raw data processing
- Price distance calculations
- Journey start detection
- Historical accuracy tracking
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import sys
from collections import defaultdict

# Traveler Classifications
EPIC_ORIGINS = {"Trinidad", "Tobago", "WASP-12b", "Macedonia"}
ANCHOR_ORIGINS = {"Spain", "Saturn", "Jupiter", "Kepler-62", "Kepler-44"}

# FOGZ (Friends of Green Zero)
FOGZ_TRAVELERS = {0, 40, -40}

# Family classifications
GREEN_FAMILY = {
    2, -2, 10, -10, 21, -21, 22, -22, 30, -30, 36, -36, 39, -39, 41, -41, 43, -43, 50, -50, 60, -60, 77, -77, 107, -107,
    5, -5, 14, -14, 26, -26, 55, -55, 56, -56, 68, -68, 91, -91, 96, -96,
    6, -6, 25, -25, 49, -49, 79, -79, 87, -87,
    3, -3, 37, -37, 63, -63, 99, -99.2, 103, -103,
    1, -1, 73, -73, 111, -111,
    52, -52, 59, -59, 65, -65, 70, -70, 76, -76, 82, -82, 86, -86, 88, -88, 97, -97
}

INDIGO_FAMILY = {
    0, 40, -40,  # Wild
    15, -15, 27, -27, 33, -33, 38, -38, 42, -42, 45, -45, 54, -54, 67, -67, 74, -74, 80, -80, 85, -85, 89, -89, 92, -92, 95, -95, 96.1, -96.1, 97.2, -97.2, 98.2, -98.2, 99.3, -99.3,
    4, -4, 12, -12, 24, -24, 31, -31, 47, -47, 57, -57, 71, -71, 93.5, -93.5, 101, -101,
    62, -62, 78.01, -78.01, 83, -83, 86.5, -86.5, 90.5, -90.5, 95.5, -95.5, 97.1, -97.1, 98.1, -98.1, 99.1, -99.1
}

X0_TAGS = {"X01p", "X02p", "X03p", "X04p", "X05p", "X06p", "X07p", "X08p", "X09p",
           "X01d", "X02d", "X03d", "X04d", "X05d", "X06d", "X07d", "X08d", "X09d"}
XD0_TAGS = {"XD01p", "XD02p", "XD03p", "XD04p", "XD05p", "XD06p", "XD07p", "XD08p", "XD09p", "XD010p",
            "XD01d", "XD02d"}

def load_price_data(csv_file):
    """Load price data from CSV feed"""
    df = pd.read_csv(csv_file)
    df['time'] = pd.to_datetime(df['time'])
    return df

def detect_tight_ranges(price_df, lookback_periods=10, range_threshold=50):
    """
    Detect tight consolidation ranges that could signal journey starts
    Returns periods where range was tight followed by breakout
    """
    if len(price_df) < lookback_periods:
        return []
    
    tight_ranges = []
    
    for i in range(lookback_periods, len(price_df)):
        window = price_df.iloc[i-lookback_periods:i]
        range_size = window['high'].max() - window['low'].min()
        
        if range_size < range_threshold:
            # Check for breakout
            current_bar = price_df.iloc[i]
            prev_high = window['high'].max()
            prev_low = window['low'].min()
            
            if current_bar['high'] > prev_high or current_bar['low'] < prev_low:
                tight_ranges.append({
                    'breakout_time': current_bar['time'],
                    'breakout_price': current_bar['close'],
                    'range_size': range_size,
                    'prev_high': prev_high,
                    'prev_low': prev_low,
                    'direction': 'UP' if current_bar['high'] > prev_high else 'DOWN'
                })
    
    return tight_ranges

def load_raw_data_grab(filepath):
    """
    Load raw data grab file and build traveler pairs
    Raw format has individual M# entries, need to build combinations
    """
    df = pd.read_excel(filepath, sheet_name=0)
    
    # Group by Output to find traveler pairs at same location
    pairs_list = []
    
    for output in df['Output'].unique():
        output_data = df[df['Output'] == output]
        
        # Get all M# values at this output
        m_numbers = output_data['M #'].unique()
        origins = output_data['Origin'].unique()
        
        # Create pairs
        for i, m1 in enumerate(m_numbers):
            for m2 in m_numbers[i+1:]:
                # Get data for each M#
                m1_data = output_data[output_data['M #'] == m1].iloc[0]
                m2_data = output_data[output_data['M #'] == m2].iloc[0]
                
                pairs_list.append({
                    'Output': output,
                    'Arrival_Output': output,
                    'M_#s': f"{m1}, {m2}",
                    'Outputs': f"{output:.2f}, {output:.2f}",
                    'Origins': f"{m1_data['Origin']}, {m2_data['Origin']}",
                    'Feed': m1_data['Feed'],
                    'Tags': f"{m1_data.get('Tag', '')}, {m2_data.get('Tag', '')}",
                    'Arrival_DateTime': m1_data['Arrival'],
                    'Day': m1_data['Day'],
                    'Family': m1_data.get('Family', '')
                })
    
    return pd.DataFrame(pairs_list), 'raw'

def load_processed_data(filepath):
    """Load processed swing analysis format"""
    df = pd.read_excel(filepath, sheet_name=0, header=2)
    if 'M_#s' in df.columns and 'Outputs' in df.columns:
        return df, 'processed'
    raise ValueError("Not a processed file format")

def load_data_grab(filepath):
    """Load data grab file - handles both raw and processed formats"""
    try:
        return load_processed_data(filepath)
    except:
        return load_raw_data_grab(filepath)

def calculate_proximity(outputs):
    """Calculate proximity between two outputs"""
    if pd.isna(outputs) or ',' not in str(outputs):
        return None
    parts = str(outputs).split(',')
    if len(parts) == 2:
        try:
            return abs(float(parts[0].strip()) - float(parts[1].strip()))
        except:
            return None
    return None

def parse_m_numbers(m_str):
    """Parse M# string into tuple of numbers"""
    if pd.isna(m_str):
        return None
    parts = str(m_str).replace(' ', '').split(',')
    if len(parts) == 2:
        try:
            return (float(parts[0]), float(parts[1]))
        except:
            return None
    return None

def parse_origins(origin_str):
    """Parse origin string into tuple"""
    if pd.isna(origin_str):
        return None
    parts = str(origin_str).split(',')
    if len(parts) == 2:
        return (parts[0].strip(), parts[1].strip())
    return None

def score_zone(row, feed_counts, current_price=None, journey_starts=None):
    """
    Score a zone with enhanced criteria
    """
    score = 0
    reasons = []
    
    m_nums = parse_m_numbers(row.get('M_#s', ''))
    origins = parse_origins(row.get('Origins', ''))
    output = row.get('Display_Output', row.get('Arrival_Output', row.get('Output', 0)))
    
    # M# 0 presence (100 points)
    if m_nums and 0 in m_nums:
        score += 100
        reasons.append("M# 0 (FOGZ)")
    
    # FOGZ presence (80 points)
    if m_nums and any(m in FOGZ_TRAVELERS for m in m_nums):
        score += 80
        reasons.append("FOGZ traveler")
    
    # Anchor origins (50 points per anchor)
    if origins:
        anchor_count = sum(1 for o in origins if o in ANCHOR_ORIGINS)
        if anchor_count > 0:
            score += anchor_count * 50
            reasons.append(f"{anchor_count} Anchor origin(s)")
    
    # Epic origins (30 points per epic)
    if origins:
        epic_count = sum(1 for o in origins if o in EPIC_ORIGINS)
        if epic_count > 0:
            score += epic_count * 30
            reasons.append(f"{epic_count} Epic origin(s)")
    
    # Same family pair (40 points)
    if m_nums:
        both_green = all(m in GREEN_FAMILY for m in m_nums)
        both_indigo = all(m in INDIGO_FAMILY for m in m_nums)
        if both_green or both_indigo:
            score += 40
            family = "Green" if both_green else "Indigo"
            reasons.append(f"Same {family} family")
    
    # X0 or XD0 tags (30 points)
    tags = str(row.get('Tags', row.get('Tag', '')))
    if any(tag in tags for tag in X0_TAGS) or any(tag in tags for tag in XD0_TAGS):
        score += 30
        reasons.append("X0/XD0 tag")
    
    # Multi-feed confirmation (60 points)
    outputs_str = row.get('Outputs', '')
    if pd.notna(outputs_str):
        for out_val in str(outputs_str).split(','):
            try:
                out_float = float(out_val.strip())
                if out_float in feed_counts and feed_counts[out_float] > 1:
                    score += 60
                    reasons.append("Multi-feed match")
                    break
            except:
                pass
    
    # Proximity bonus (tighter = better, max 20 points)
    prox = calculate_proximity(row.get('Outputs', ''))
    if prox is not None and prox < 10:
        prox_score = int((10 - prox) * 2)
        score += prox_score
        reasons.append(f"Tight prox ({prox:.2f})")
    
    # Price distance bonus (closer = better, max 50 points)
    if current_price is not None and pd.notna(output):
        distance = abs(float(output) - current_price)
        if distance < 100:
            distance_score = int((100 - distance) / 2)
            score += distance_score
            reasons.append(f"Near price ({distance:.2f} pts)")
    
    # Journey start bonus (100 points)
    if journey_starts and pd.notna(output):
        for js in journey_starts:
            if abs(float(output) - js['breakout_price']) < 10:
                score += 100
                reasons.append(f"Journey start {js['direction']}")
                break
    
    return score, "; ".join(reasons)

def identify_gandalf_zones(df):
    """Identify Gandalf zones: (0, +/-40) combinations that appear on only ONE feed"""
    fogz_matches = df[df['M_#s'].str.contains('0, 40|0, -40', na=False, regex=True)].copy()
    
    if len(fogz_matches) == 0:
        return []
    
    output_data = []
    for _, row in fogz_matches.iterrows():
        outputs_str = str(row.get('Outputs', ''))
        if ',' in outputs_str:
            outputs = [float(x.strip()) for x in outputs_str.split(',')]
            for out in outputs:
                output_data.append({
                    'Output': out,
                    'Feed': row['Feed'],
                    'M_#s': row['M_#s'],
                    'Origins': row.get('Origins', ''),
                    'Full_Row': row
                })
    
    output_df = pd.DataFrame(output_data)
    feed_counts = output_df.groupby('Output')['Feed'].nunique()
    gandalf_outputs = feed_counts[feed_counts == 1].index.tolist()
    
    gandalf_zones = []
    for out in gandalf_outputs:
        zone_data = output_df[output_df['Output'] == out].iloc[0]
        gandalf_zones.append({
            'Output': out,
            'Feed': zone_data['Feed'],
            'M_#s': zone_data['M_#s'],
            'Origins': zone_data['Origins'],
            'Type': 'GANDALF ZONE',
            'Description': f'{zone_data["Feed"]} feed only - Price should not stay past this level'
        })
    
    return gandalf_zones

def track_historical_accuracy(df, price_df, lookback_hours=24):
    """
    Track which patterns preceded actual price moves
    Returns accuracy statistics for different pattern types
    """
    if price_df is None or len(price_df) == 0:
        return None
    
    # Get price range for lookback period
    recent_time = price_df['time'].max() - timedelta(hours=lookback_hours)
    recent_prices = price_df[price_df['time'] >= recent_time].copy()
    
    if len(recent_prices) < 2:
        return None
    
    price_high = recent_prices['high'].max()
    price_low = recent_prices['low'].min()
    
    # Analyze which zones were hit
    stats = {
        'total_zones': len(df),
        'zones_hit': 0,
        'fogz_hit': 0,
        'fogz_total': 0,
        'anchor_hit': 0,
        'anchor_total': 0,
        'multi_feed_hit': 0,
        'multi_feed_total': 0,
        'journey_starts': []
    }
    
    for _, row in df.iterrows():
        output = row.get('Display_Output', row.get('Arrival_Output', row.get('Output')))
        if pd.isna(output):
            continue
        
        output_float = float(output)
        zone_hit = price_low <= output_float <= price_high
        
        if zone_hit:
            stats['zones_hit'] += 1
        
        # FOGZ patterns
        m_nums = parse_m_numbers(row.get('M_#s', ''))
        if m_nums and any(m in FOGZ_TRAVELERS for m in m_nums):
            stats['fogz_total'] += 1
            if zone_hit:
                stats['fogz_hit'] += 1
        
        # Anchor patterns
        origins = parse_origins(row.get('Origins', ''))
        if origins and any(o in ANCHOR_ORIGINS for o in origins):
            stats['anchor_total'] += 1
            if zone_hit:
                stats['anchor_hit'] += 1
    
    # Calculate accuracy rates
    stats['hit_rate'] = stats['zones_hit'] / stats['total_zones'] if stats['total_zones'] > 0 else 0
    stats['fogz_accuracy'] = stats['fogz_hit'] / stats['fogz_total'] if stats['fogz_total'] > 0 else 0
    stats['anchor_accuracy'] = stats['anchor_hit'] / stats['anchor_total'] if stats['anchor_total'] > 0 else 0
    
    return stats

def create_dashboard(df, report_time, output_file, current_price=None, price_df=None):
    """Create the enhanced rapid analysis dashboard"""
    
    # Extract individual outputs for multi-feed counting
    output_feed_map = {}
    for _, row in df.iterrows():
        outputs_str = str(row.get('Outputs', ''))
        feed = row.get('Feed', '')
        if ',' in outputs_str:
            outputs = [float(x.strip()) for x in outputs_str.split(',')]
            for out in outputs:
                if out not in output_feed_map:
                    output_feed_map[out] = set()
                output_feed_map[out].add(feed)
    
    output_feed_counts = {out: len(feeds) for out, feeds in output_feed_map.items()}
    
    # Detect journey starts from price data
    journey_starts = []
    if price_df is not None:
        journey_starts = detect_tight_ranges(price_df)
    
    # Use Arrival_Output as the primary output for sorting/display
    df['Display_Output'] = df.get('Arrival_Output', df.get('Output', ''))
    
    # Calculate price distances
    if current_price is not None:
        df['Distance'] = df['Display_Output'].apply(
            lambda x: abs(float(x) - current_price) if pd.notna(x) else None
        )
    else:
        df['Distance'] = None
    
    # Score all zones
    df['Score'], df['Reasons'] = zip(*df.apply(
        lambda row: score_zone(row, output_feed_counts, current_price, journey_starts), 
        axis=1
    ))
    
    # Sort by score
    df_sorted = df.sort_values('Score', ascending=False)
    
    # Get top zones
    top_zones = df_sorted.head(100).copy()
    
    # Identify Gandalf zones
    gandalf_zones = identify_gandalf_zones(df)
    
    # Track historical accuracy
    historical_stats = track_historical_accuracy(df, price_df) if price_df is not None else None
    
    # Create workbook
    wb = Workbook()
    
    # === PRIORITY ZONES SHEET ===
    ws_priority = wb.active
    ws_priority.title = "Priority Zones"
    
    # Header
    ws_priority['A1'] = 'TRAVELER RAPID ANALYSIS DASHBOARD - ENHANCED'
    ws_priority['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_priority['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws_priority.merge_cells('A1:M1')
    
    info_text = f'Report Time: {report_time}'
    if current_price:
        info_text += f' | Current Price: {current_price:.2f}'
    ws_priority['A2'] = info_text
    ws_priority['A2'].font = Font(size=11, italic=True)
    ws_priority.merge_cells('A2:M2')
    
    ws_priority['A3'] = f'Total Matches: {len(df)} | Top {len(top_zones)} High-Priority Zones | Journey Starts: {len(journey_starts)}'
    ws_priority['A3'].font = Font(size=10)
    ws_priority.merge_cells('A3:M3')
    
    # Column headers
    headers = ['Rank', 'Score', 'Output', 'Distance', 'M_#s', 'Origins', 'Feed', 'Tag', 'Proximity', 'Arrival', 'Day', 'Journey?', 'Priority Reasons']
    for col, header in enumerate(headers, 1):
        cell = ws_priority.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for idx, (_, row) in enumerate(top_zones.iterrows(), 1):
        r = idx + 5
        ws_priority.cell(r, 1, idx)
        ws_priority.cell(r, 2, row['Score'])
        ws_priority.cell(r, 3, row.get('Display_Output', ''))
        
        dist = row.get('Distance')
        ws_priority.cell(r, 4, f"{dist:.2f}" if pd.notna(dist) else "")
        
        ws_priority.cell(r, 5, row.get('M_#s', ''))
        ws_priority.cell(r, 6, row.get('Origins', ''))
        ws_priority.cell(r, 7, row.get('Feed', ''))
        ws_priority.cell(r, 8, row.get('Tags', row.get('Tag', '')))
        
        prox = calculate_proximity(row.get('Outputs', ''))
        ws_priority.cell(r, 9, f"{prox:.2f}" if prox else "")
        
        arrival = row.get('Arrival_DateTime', row.get('Arrival', ''))
        ws_priority.cell(r, 10, str(arrival) if pd.notna(arrival) else "")
        ws_priority.cell(r, 11, row.get('Day', ''))
        
        # Check if near journey start
        is_journey = "✓" if "Journey start" in str(row.get('Reasons', '')) else ""
        ws_priority.cell(r, 12, is_journey)
        
        ws_priority.cell(r, 13, row.get('Reasons', ''))
        
        # Color coding
        if row['Score'] >= 250:
            fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Dark green
        elif row['Score'] >= 200:
            fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        elif row['Score'] >= 150:
            fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        else:
            fill = None
        
        if fill:
            for col in range(1, 14):
                ws_priority.cell(r, col).fill = fill
    
    # Set column widths
    widths = [8, 10, 12, 10, 12, 25, 10, 10, 10, 20, 10, 10, 60]
    for i, width in enumerate(widths, 1):
        ws_priority.column_dimensions[chr(64+i)].width = width
    
    # === JOURNEY STARTS SHEET ===
    if journey_starts:
        ws_journey = wb.create_sheet("Journey Starts")
        
        ws_journey['A1'] = 'DETECTED JOURNEY STARTS'
        ws_journey['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_journey['A1'].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        ws_journey.merge_cells('A1:G1')
        
        ws_journey['A2'] = 'Tight range breakouts that may signal trend starts/continuations'
        ws_journey['A2'].font = Font(size=10, italic=True)
        ws_journey.merge_cells('A2:G2')
        
        headers_j = ['Breakout Time', 'Breakout Price', 'Direction', 'Range Size', 'Prev High', 'Prev Low', 'FOGZ Match?']
        for col, header in enumerate(headers_j, 1):
            cell = ws_journey.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        
        for idx, js in enumerate(journey_starts, 1):
            r = idx + 4
            ws_journey.cell(r, 1, str(js['breakout_time']))
            ws_journey.cell(r, 2, js['breakout_price'])
            ws_journey.cell(r, 3, js['direction'])
            ws_journey.cell(r, 4, f"{js['range_size']:.2f}")
            ws_journey.cell(r, 5, js['prev_high'])
            ws_journey.cell(r, 6, js['prev_low'])
            
            # Check for FOGZ match
            fogz_match = ""
            for _, zone in df.iterrows():
                if '0,' in str(zone.get('M_#s', '')) or ', 0' in str(zone.get('M_#s', '')):
                    zone_output = zone.get('Display_Output', zone.get('Output'))
                    if pd.notna(zone_output) and abs(float(zone_output) - js['breakout_price']) < 10:
                        fogz_match = "✓ FOGZ"
                        break
            ws_journey.cell(r, 7, fogz_match)
        
        for col in 'ABCDEFG':
            ws_journey.column_dimensions[col].width = 18
    
    # === GANDALF ZONES SHEET ===
    if gandalf_zones:
        ws_gandalf = wb.create_sheet("Gandalf Zones")
        
        ws_gandalf['A1'] = 'GANDALF ZONES - "You Shall Not Pass"'
        ws_gandalf['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_gandalf['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        ws_gandalf.merge_cells('A1:G1')
        
        ws_gandalf['A2'] = 'Single-feed (0, +/-40) matches - Price should not stay past these levels'
        ws_gandalf['A2'].font = Font(size=10, italic=True)
        ws_gandalf.merge_cells('A2:G2')
        
        headers_g = ['Output', 'Distance', 'Feed', 'M_#s', 'Origins', 'Type', 'Description']
        for col, header in enumerate(headers_g, 1):
            cell = ws_gandalf.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        
        for idx, zone in enumerate(gandalf_zones, 1):
            r = idx + 4
            ws_gandalf.cell(r, 1, zone['Output'])
            
            dist = abs(zone['Output'] - current_price) if current_price else None
            ws_gandalf.cell(r, 2, f"{dist:.2f}" if dist else "")
            
            ws_gandalf.cell(r, 3, zone['Feed'])
            ws_gandalf.cell(r, 4, zone['M_#s'])
            ws_gandalf.cell(r, 5, zone['Origins'])
            ws_gandalf.cell(r, 6, zone['Type'])
            ws_gandalf.cell(r, 7, zone['Description'])
            
            # Highlight closest zones
            if dist and dist < 50:
                for col in range(1, 8):
                    ws_gandalf.cell(r, col).fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            else:
                for col in range(1, 8):
                    ws_gandalf.cell(r, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        ws_gandalf.column_dimensions['A'].width = 12
        ws_gandalf.column_dimensions['B'].width = 10
        ws_gandalf.column_dimensions['C'].width = 10
        ws_gandalf.column_dimensions['D'].width = 12
        ws_gandalf.column_dimensions['E'].width = 25
        ws_gandalf.column_dimensions['F'].width = 15
        ws_gandalf.column_dimensions['G'].width = 60
    
    # === HISTORICAL ACCURACY SHEET ===
    if historical_stats:
        ws_history = wb.create_sheet("Historical Accuracy")
        
        ws_history['A1'] = 'PATTERN ACCURACY TRACKING'
        ws_history['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_history['A1'].fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        ws_history.merge_cells('A1:D1')
        
        ws_history['A3'] = 'Pattern Type'
        ws_history['B3'] = 'Zones Hit'
        ws_history['C3'] = 'Total Zones'
        ws_history['D3'] = 'Accuracy %'
        
        for col in 'ABCD':
            ws_history[f'{col}3'].font = Font(bold=True)
            ws_history[f'{col}3'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        stats_data = [
            ['All Zones', historical_stats['zones_hit'], historical_stats['total_zones'], 
             f"{historical_stats['hit_rate']*100:.1f}%"],
            ['FOGZ Patterns', historical_stats['fogz_hit'], historical_stats['fogz_total'], 
             f"{historical_stats['fogz_accuracy']*100:.1f}%"],
            ['Anchor Origins', historical_stats['anchor_hit'], historical_stats['anchor_total'], 
             f"{historical_stats['anchor_accuracy']*100:.1f}%"],
        ]
        
        for i, row_data in enumerate(stats_data, 4):
            for j, val in enumerate(row_data, 1):
                ws_history.cell(i, j, val)
        
        for col in 'ABCD':
            ws_history.column_dimensions[col].width = 20
    
    # === FOGZ ZONES SHEET ===
    fogz_data = df[df['M_#s'].str.contains('0, 40|0, -40', na=False, regex=True)].copy()
    if len(fogz_data) > 0:
        ws_fogz = wb.create_sheet("FOGZ (0,±40)")
        
        ws_fogz['A1'] = 'FOGZ M# 0 with ±40 Combinations'
        ws_fogz['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_fogz['A1'].fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        ws_fogz.merge_cells('A1:I1')
        
        headers_f = ['Arrival Output', 'Distance', 'M_#s', 'Outputs', 'Origins', 'Feed', 'Tags', 'Arrival DateTime', 'Day']
        for col, header in enumerate(headers_f, 1):
            cell = ws_fogz.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        for idx, (_, row) in enumerate(fogz_data.iterrows(), 1):
            r = idx + 3
            ws_fogz.cell(r, 1, row.get('Display_Output', ''))
            
            dist = row.get('Distance')
            ws_fogz.cell(r, 2, f"{dist:.2f}" if pd.notna(dist) else "")
            
            ws_fogz.cell(r, 3, row.get('M_#s', ''))
            ws_fogz.cell(r, 4, row.get('Outputs', ''))
            ws_fogz.cell(r, 5, row.get('Origins', ''))
            ws_fogz.cell(r, 6, row.get('Feed', ''))
            ws_fogz.cell(r, 7, row.get('Tags', row.get('Tag', '')))
            arrival = row.get('Arrival_DateTime', row.get('Arrival', ''))
            ws_fogz.cell(r, 8, str(arrival) if pd.notna(arrival) else "")
            ws_fogz.cell(r, 9, row.get('Day', ''))
        
        for col in 'ABCDEFGHI':
            ws_fogz.column_dimensions[col].width = 15
    
    # Save
    wb.save(output_file)
    return len(top_zones), len(gandalf_zones), len(journey_starts), historical_stats

def main(data_grab_file, output_file=None, price_csv=None, current_price=None):
    """Main analysis function"""
    print(f"Loading data from: {data_grab_file}")
    df, file_type = load_data_grab(data_grab_file)
    print(f"Detected file type: {file_type}")
    
    # Load price data if provided
    price_df = None
    if price_csv:
        print(f"Loading price data from: {price_csv}")
        price_df = load_price_data(price_csv)
        if current_price is None:
            current_price = price_df['close'].iloc[-1]
            print(f"Using latest close as current price: {current_price:.2f}")
    
    # Extract report time
    try:
        xl = pd.ExcelFile(data_grab_file)
        sheet_name = xl.sheet_names[0]
        report_time = sheet_name
    except:
        report_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_file = f"/mnt/user-data/outputs/traveler_analysis_{timestamp}.xlsx"
    
    print(f"Processing {len(df)} matches...")
    
    top_count, gandalf_count, journey_count, hist_stats = create_dashboard(
        df, report_time, output_file, current_price, price_df
    )
    
    print(f"\nAnalysis complete!")
    print(f"Top zones identified: {top_count}")
    print(f"Gandalf zones found: {gandalf_count}")
    print(f"Journey starts detected: {journey_count}")
    
    if hist_stats:
        print(f"\nHistorical Accuracy (last 24h):")
        print(f"  Overall hit rate: {hist_stats['hit_rate']*100:.1f}%")
        print(f"  FOGZ accuracy: {hist_stats['fogz_accuracy']*100:.1f}%")
        print(f"  Anchor accuracy: {hist_stats['anchor_accuracy']*100:.1f}%")
    
    print(f"\nDashboard saved to: {output_file}")
    return output_file

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Traveler Rapid Analysis Dashboard')
    parser.add_argument('data_file', help='Data grab Excel file (raw or processed)')
    parser.add_argument('--output', '-o', help='Output file path')
    parser.add_argument('--price-csv', '-p', help='Price CSV file for journey start detection')
    parser.add_argument('--current-price', '-c', type=float, help='Current price for distance calculations')
    
    args = parser.parse_args()
    
    main(args.data_file, args.output, args.price_csv, args.current_price)
