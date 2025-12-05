"""
Excel Exporter v21 - Enhanced Excel Export with Highlighting

This module handles Excel export for all model results with:
- Report Time in sheet headers
- Report datetime in filename
- Highlighting (yellow for [0], blue for [-1][-2][-3])
"""

import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def export_all_models_to_excel(all_results, report_time):
    """
    Export all model results to Excel with highlighting.
    
    Parameters:
    -----------
    all_results : dict
        Dictionary from model_processor with 'results', 'timings', etc.
    report_time : datetime
        Report datetime to include in headers and filename
    
    Returns:
    --------
    tuple : (output_bytes, filename)
    """
    output = io.BytesIO()
    results = all_results['results']
    timings = all_results.get('timings', {})
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write each model to its own sheet
        for model_name, df in results.items():
            if len(df) > 0:
                # Create sheet name (Excel limit: 31 chars)
                sheet_name = model_name[:31]
                
                # Write dataframe
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                
                # Get worksheet to add header and formatting
                worksheet = writer.sheets[sheet_name]
                
                # Add Report Time header (row 1)
                report_str = report_time.strftime('%Y-%m-%d %H:%M')
                worksheet['A1'] = f"Report Time: {report_str}"
                worksheet['A1'].font = worksheet['A1'].font.copy(bold=True)
                
                # Add timing info if available (row 2)
                if model_name in timings:
                    timing = timings[model_name]
                    worksheet['A2'] = f"Generated in {timing:.2f}s | {len(df)} matches"
                    worksheet['A2'].font = worksheet['A2'].font.copy(italic=True)
        
        # Create Combined sheet with all results
        all_tables = [df for df in results.values() if len(df) > 0]
        if all_tables:
            combined_df = pd.concat(all_tables, ignore_index=True)
            combined_df = combined_df.sort_values('Arrival_Output', ascending=False)
            
            # Write combined sheet
            combined_df.to_excel(writer, sheet_name='Combined', index=False, startrow=2)
            
            # Add headers to combined sheet
            combined_ws = writer.sheets['Combined']
            combined_ws['A1'] = f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')}"
            combined_ws['A1'].font = combined_ws['A1'].font.copy(bold=True)
            combined_ws['A2'] = f"All Models Combined | {len(combined_df)} total matches"
            combined_ws['A2'].font = combined_ws['A2'].font.copy(italic=True)
    
    # Reopen workbook to apply highlighting
    output.seek(0)
    workbook = load_workbook(output)
    
    # Define fill colors
    yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    blue_fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
    
    # Apply highlighting to each sheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Find Arrival_Brackets column (data starts at row 4 due to headers)
        header_row = 4  # Row 1-2 are custom headers, row 3 is blank, row 4 is column headers
        
        # Find column index for Arrival_Brackets
        arrival_brackets_col = None
        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            if cell.value == 'Arrival_Brackets':
                arrival_brackets_col = col_idx
                break
        
        if arrival_brackets_col:
            # Apply highlighting based on Arrival_Brackets values
            for row in worksheet.iter_rows(min_row=header_row+1, max_row=worksheet.max_row):
                cell = row[arrival_brackets_col - 1]  # -1 because enumerate starts at 1
                value = str(cell.value) if cell.value else ''
                
                # Check for [0] - yellow
                if '[0]' in value:
                    cell.fill = yellow_fill
                # Check for [-1], [-2], [-3] - blue
                elif any(x in value for x in ['[-1]', '[-2]', '[-3]']):
                    cell.fill = blue_fill
    
    # Save modified workbook
    output_final = io.BytesIO()
    workbook.save(output_final)
    output_final.seek(0)
    
    # Create filename with report datetime
    report_dt_str = report_time.strftime('%Y%m%d_%H%M')
    filename = f"swing_analysis_{report_dt_str}.xlsx"
    
    return output_final.getvalue(), filename


def create_download_button(all_results, report_time):
    """
    Create Streamlit download button for Excel export.
    
    Parameters:
    -----------
    all_results : dict
        Dictionary from model_processor
    report_time : datetime
        Report datetime
    
    Returns:
    --------
    None (creates Streamlit widget)
    """
    import streamlit as st
    
    try:
        excel_bytes, filename = export_all_models_to_excel(all_results, report_time)
        
        # Count how many models have results
        num_models = sum(1 for df in all_results['results'].values() if len(df) > 0)
        total_matches = sum(len(df) for df in all_results['results'].values())
        
        st.download_button(
            label=f"📥 Download All Model Results (Excel)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_v21"
        )
        
        st.caption(
            f"Excel file contains {num_models} model sheets + Combined sheet "
            f"({total_matches} total matches) | "
            f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')} | "
            f"🟡 Yellow = [0] (Today) | 🔵 Blue = [-1][-2][-3] (Recent)"
        )
        
    except Exception as e:
        st.error(f"Error creating Excel export: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
