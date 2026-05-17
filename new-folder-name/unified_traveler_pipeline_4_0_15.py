#!/usr/bin/env python3
"""
UNIFIED TRAVELER PIPELINE v4.0.15
==================================
Major Enhancements over v3.0:
  - Multi-asset support: NQ, ES, RTY processed in a single run
  - Quick mode (NEW DEFAULT): 5×15-min + measurement only → Raw + Matched reports
  - Full mode (--full): 16-file intake with OHLC construction, MA normalization, wick analysis
  - Auto-discovery of feed files from directory with pattern validation
  - Unix / ISO timestamp auto-detection and NY conversion
  - Multi-chart OHLC construction (3-min → 6/9/12/18/24-min; 5-min → 10/20/30/45/60/90/120/240-min)
    [full mode only]
  - 15-min quality-check reconstruction from both 3-min and 5-min data [full mode only]
  - Moving Average normalization (h-series h1–h20, duplicate QC) [full mode only]
  - Raw + Matched traveler reports per asset (v3.0 core logic preserved)
  - Wick analysis: primary & secondary nested swing candles, per-asset Excel tabs
    reporting which MAs and Travelers plot inside candle wicks [full mode only]

Usage (quick — default, 6 files):
    python unified_traveler_pipeline_4_0_15.py \\
        --measurement "Meas 4_8 all.xlsx" \\
        --dt "2026-02-23 18:30" \\
        --lookback 22 \\
        [--dir "."]

Usage (full — 16 files, wick analysis):
    python unified_traveler_pipeline_4_0_15.py \\
        --measurement "Meas 4_8 all.xlsx" \\
        --dt "2026-02-23 18:30" \\
        --lookback 22 \\
        --full

Author: Built from unified_traveler_pipeline_3_0.py
Date:   May 2026
Version: 4.0.15
"""

import pandas as pd
import numpy as np
import datetime as dt
from datetime import datetime, timedelta
import argparse
import sys
import os
import re
import io
import json
import time as time_module
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any

# Excel handling
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("=" * 120)
print("🌌 UNIFIED TRAVELER PIPELINE v4.0.15  —  NQ · ES · RTY  |  Quick Mode (default)  |  Full Mode (--full)  |  Backtest/Forecast")
print("=" * 120)

pd.set_option("styler.render.max_elements", 2_000_000)

# ============================================================================
# SECTION 1: DEPENDENCY IMPORTS  (same graceful-fallback pattern as v3.0)
# ============================================================================

print("\n📦 Loading dependencies...")

USE_NEW_CALCULATOR = False
apply_full_range_advanced = None

try:
    from custom_range_calculator_cli import apply_full_range_advanced as _calc
    apply_full_range_advanced = _calc
    print("  ✅ custom_range_calculator_cli")
    USE_NEW_CALCULATOR = True
except ImportError:
    if "streamlit" not in sys.modules:
        class _MockST:
            def __getattr__(self, n):
                def _f(*a, **kw):
                    return self if n in ("progress", "empty") else None
                return _f
        sys.modules["streamlit"] = _MockST()
    try:
        from custom_range_calculator_1125_21c import apply_full_range_advanced as _calc
        apply_full_range_advanced = _calc
        print("  ✅ custom_range_calculator_1125_21c (with streamlit mock)")
        USE_NEW_CALCULATOR = True
    except ImportError:
        print("  ❌ No range calculator found.")
        apply_full_range_advanced = None  # will raise clearly at pipeline run

try:
    from a_helpers import (
        clean_timestamp, process_feed, get_input_at_day_start,
        GROUP_1A_TRAVELERS, GROUP_1B_TRAVELERS, GROUP_2A_TRAVELERS, GROUP_2B_TRAVELERS,
        get_input_value, generate_master_traveler_list,
        EPIC_ORIGINS, ANCHOR_ORIGINS,
    )
    print("  ✅ a_helpers")
except ImportError as e:
    print(f"  ❌ a_helpers — {e}")
    sys.exit(1)

try:
    from nested_swing_detector import analyze_swings, parse_timestamp_naive, detect_nested_swings
    print("  ✅ nested_swing_detector")
    HAS_SWING_DETECTOR = True
except ImportError as e:
    print(f"  ⚠️  nested_swing_detector — {e}  (wick analysis will be skipped)")
    HAS_SWING_DETECTOR = False
    def analyze_swings(*a, **kw): return [], []
    def parse_timestamp_naive(x): return pd.to_datetime(x)
    def detect_nested_swings(*a, **kw): return []

try:
    from model_definitions_v21 import MODELS, get_reciprocal_lookup, apply_special_matching
    print("  ✅ model_definitions_v21")
    HAS_MODEL_DEFINITIONS = True
except ImportError:
    HAS_MODEL_DEFINITIONS = False
    MODELS = {}
    print("  ⚠️  model_definitions_v21 not found")

HAS_MODEL_PROCESSOR = False
try:
    import importlib.util
    spec = importlib.util.find_spec("model_processor_v21")
    if spec and spec.origin:
        from model_processor_v21 import get_model_display_info, organize_results_by_category, create_summary_stats
        print("  ✅ model_processor_v21")
        HAS_MODEL_PROCESSOR = True
except Exception:
    pass

if not HAS_MODEL_PROCESSOR and HAS_MODEL_DEFINITIONS:
    def get_model_display_info(m):
        c = MODELS.get(m, {})
        return {"number": c.get("number", 0), "display_name": c.get("display_name", m),
                "description": c.get("description", ""), "check_recip": c.get("check_recip", False)}
    def organize_results_by_category(r): return {}
    def create_summary_stats(r):
        res = r.get("results", {})
        return {"total_models": len(res),
                "models_with_matches": sum(1 for d in res.values() if len(d) > 0),
                "total_matches": sum(len(d) for d in res.values())}

try:
    from bypass_mode_matcher import match_travelers_bypass_mode, process_model_bypass_mode
    print("  ✅ bypass_mode_matcher")
    HAS_BYPASS_MATCHER = True
except ImportError:
    HAS_BYPASS_MATCHER = False
    print("  ⚠️  bypass_mode_matcher not found")

print("\n✅ Critical dependencies loaded.\n")


# ============================================================================
# SECTION 2: CONFIGURATION & PER-ASSET CONSTANTS
# ============================================================================

# --- Session open ---
DEFAULT_SESSION_OPEN_HOUR = 18   # 18:00 NY; detected as 17 if present

# --- Per-asset defaults ---
ASSET_DEFAULTS = {
    "NQ": {"max_spread": 3.0, "radius": 1000, "composite": False},
    "ES": {"max_spread": 1.0, "radius": 215,  "composite": False},
    "RTY": {"max_spread": 0.5, "radius": 100,  "composite": True},   # no MRTY; single-feed composite
}

# --- Chart build specs ---
# From 3-min source: build these intervals (minutes)
BUILD_FROM_3MIN = [6, 9, 12, 18, 24]
# From 5-min source: build these intervals (minutes)
BUILD_FROM_5MIN = [10, 20, 30, 45, 60, 90, 120, 240]
# 15-min QC timeframes (built from both, compared to actual)
QC_TIMEFRAME = 15

# --- Priority origins (mirrored from v3.0) ---
EPIC_ORIGINS_BASE  = {"trinidad", "tobago", "wasp-12b", "macedonia"}
ANCHOR_ORIGINS_BASE = {"spain", "saturn", "jupiter", "kepler-62", "kepler-44"}
EPIC_ORIGINS_WITH_BRACKETS = set(EPIC_ORIGINS_BASE)
for _o in list(EPIC_ORIGINS_BASE):
    for _sfx in ["[1]", "[2]", "[-1]", "[-2]"]:
        EPIC_ORIGINS_WITH_BRACKETS.add(f"{_o}{_sfx}")
PRIORITY_ORIGINS = set(list(EPIC_ORIGINS_WITH_BRACKETS) + list(ANCHOR_ORIGINS_BASE))

DEFAULT_LOOKBACK_DAYS = 20

# --- Backtest / Forecast constants ---
BACKTEST_SWING_THRESHOLDS = {"NQ": 60.0, "ES": 20.0, "RTY": 16.0}  # min pts for a valid swing
DEFAULT_MA_PROXIMITY      = 5.0    # pts: MA must be within this of swing extreme to be tagged
DEFAULT_TV_PROXIMITY      = 3.0    # pts: traveler Output1 must be within this of swing extreme
DEFAULT_VELOCITY_LOOKBACK = 12     # bars (5-min): lookback for MA velocity calc (~60 min)
DEFAULT_CATALOG_FILENAME  = "backtest_catalog.json"
# Group rank order for filtering (lower index = higher rank)
_GROUP_RANK_ORDER = ["1. SAA", "2. STT", "2. TT", "3. TA", "3. AT", "4. AA",
                     "5. oA", "6. Ao", "7. oo"]


# ============================================================================
# SECTION 3: FILE DISCOVERY & VALIDATION
# ============================================================================

# Expected file-set: 5 × 15-min, 5 × 3-min, 5 × 5-min
# Asset keys for each timeframe group
_ASSET_PATTERNS = [
    # (asset, feed_role, prefix_pattern_upper)
    ("NQ",  "small", "CME_MINI_MNQ"),
    ("NQ",  "big",   "CME_MINI_NQ"),   # checked AFTER MNQ
    ("ES",  "small", "CME_MINI_MES"),
    ("ES",  "big",   "CME_MINI_ES"),   # checked AFTER MES
    ("RTY", "composite", "CME_MINI_RTY"),
]

def _classify_csv(filepath: Path) -> Optional[Dict]:
    """
    Return dict with keys: asset, role, timeframe  (or None if unrecognised)
    """
    name_upper = filepath.name.upper()

    # Determine asset / role — order matters (M-prefix first)
    asset = role = None
    for (a, r, pat) in _ASSET_PATTERNS:
        if pat in name_upper:
            asset, role = a, r
            break
    if asset is None:
        return None

    # Determine timeframe from ", 15" / ", 3" / ", 5" in name
    # Support comma-space or underscore variants
    name_check = filepath.name  # preserve original case for comma check
    if ", 15" in name_check or "_15 " in name_check or name_check.endswith("_15.csv"):
        tf = 15
    elif ", 3" in name_check and ", 30" not in name_check:
        tf = 3
    elif ", 5" in name_check and ", 50" not in name_check:
        tf = 5
    else:
        return None

    return {"path": filepath, "asset": asset, "role": role, "timeframe": tf}


def discover_feed_files(search_dir: str) -> Dict:
    """
    Scan search_dir for CSV files matching the 15-asset naming convention.
    Returns nested dict:  files[asset][timeframe] = {path, role}
    Also returns a flat list of all discovered entries.
    """
    search_path = Path(search_dir)
    csvs = list(search_path.glob("*.csv")) + list(search_path.glob("*.CSV"))

    discovered: Dict[str, Dict[int, Dict]] = {a: {} for a in ["NQ", "ES", "RTY"]}
    all_entries = []

    for csv_path in csvs:
        info = _classify_csv(csv_path)
        if info is None:
            continue
        a, tf = info["asset"], info["timeframe"]
        # For the same (asset, timeframe) keep the most-recently-modified file
        if tf not in discovered[a] or csv_path.stat().st_mtime > discovered[a][tf]["path"].stat().st_mtime:
            discovered[a][tf] = {"path": csv_path, "role": info["role"]}
            all_entries.append(info)

    return discovered, all_entries


def _extract_contract_suffix(filepath: Path) -> Optional[str]:
    """
    Extract the CME quarterly contract letter + year from a filename.
    e.g.  CME_MINI_NQH2026, 15.csv  →  'H2026'
          CME_MINI_MNQM2026__3.csv  →  'M2026'
    Returns None if no contract suffix pattern is found (e.g. continuous contracts).
    """
    # Match a contract letter (H M U Z) followed by a 4-digit year
    m = re.search(r'([HMUZ])(\d{4})', filepath.name, re.IGNORECASE)
    if m:
        return f"{m.group(1).upper()}{m.group(2)}"
    return None


def _validate_contract_suffixes(discovered: Dict) -> bool:
    """
    For each non-composite asset, check that all discovered files share the same
    contract letter+year.  Prints warnings on mismatch; RTY composite is exempt.
    Returns True if no mismatches found (or if only continuous contracts are present).
    """
    print("\n  🔎 Contract suffix consistency check...")
    all_match = True

    for asset in ["NQ", "ES", "RTY"]:
        info_map = discovered.get(asset, {})
        if not info_map:
            continue

        # Collect (timeframe, suffix) pairs for this asset
        suffix_map: Dict[int, Optional[str]] = {}
        for tf, info in info_map.items():
            suffix_map[tf] = _extract_contract_suffix(info["path"])

        # Gather unique non-None suffixes
        unique_suffixes = {s for s in suffix_map.values() if s is not None}

        if len(unique_suffixes) == 0:
            # All continuous — nothing to validate
            print(f"    ℹ️  {asset}: no contract suffix detected (continuous contract) — skipping check")
        elif len(unique_suffixes) == 1:
            suffix = next(iter(unique_suffixes))
            print(f"    ✅ {asset}: all files share contract suffix {suffix}")
        else:
            # Mismatch detected
            all_match = False
            detail = "  |  ".join(
                f"{tf}-min → {suffix_map[tf] or 'n/a'}"
                for tf in sorted(suffix_map)
            )
            print(f"    ⚠️  {asset}: CONTRACT SUFFIX MISMATCH — {detail}")
            print(f"       Traveler origins and MA charts may be from different contracts.")
            print(f"       This can produce incorrect traveler output values.")

    return all_match


def validate_file_set(discovered: Dict, measurement_path: str,
                      quick_mode: bool = True) -> bool:
    """
    Validate that required files are present.
    quick_mode=True  → only 5×15-min + measurement (6 files)
    quick_mode=False → all 16 files (5×15-min, 5×3-min, 5×5-min + measurement)
    Also checks contract suffix consistency across timeframes per asset.
    Prints detailed status.  Returns True if all present (contract warnings do not abort).
    """
    required_tfs = [15] if quick_mode else [15, 3, 5]
    n_required = 1 + 5 * len(required_tfs)
    print(f"\n🔍 Validating file set ({'QUICK — ' if quick_mode else 'FULL — '}{n_required} required)...")
    all_ok = True

    # Check measurement
    meas_ok = Path(measurement_path).exists()
    print(f"  {'✅' if meas_ok else '❌'} Measurement: {measurement_path}")
    if not meas_ok:
        all_ok = False

    # Check each asset × required timeframe combination
    required = [(a, tf) for a in ["NQ", "ES", "RTY"] for tf in required_tfs]
    for (asset, tf) in required:
        if tf in discovered.get(asset, {}):
            info = discovered[asset][tf]
            print(f"  ✅ {asset:>3s}  {tf:>2d}-min  [{info['role']:>9s}]  {info['path'].name}")
        else:
            print(f"  ❌ {asset:>3s}  {tf:>2d}-min  — FILE MISSING  "
                  f"(expected CME_MINI_{'M' if asset != 'RTY' else ''}{'NQ' if asset == 'NQ' else asset}..., {tf}")
            all_ok = False

    # Contract suffix consistency (warning only — does not abort)
    _validate_contract_suffixes(discovered)

    if all_ok:
        label = "6 files" if quick_mode else "16 files"
        print(f"  ✅ All {label} present — ready for takeoff!")
    else:
        print("  ❌ Missing files detected — pipeline cannot continue.")
    return all_ok


def _get_contract_id(discovered: Dict, asset: str) -> str:
    """
    Derive a short contract identifier string from discovered feed filenames.

    Examples:
      CME_MINI_NQM2026, 15.csv   →  'NQM'
      CME_MINI_MNQH2026__15.csv  →  'NQH'
      CME_MINI_NQ1!__5.csv       →  'NQ1_'  (composite; '!' sanitised)
      No files found             →  asset   ('NQ')

    Uses the 15-min feed if available, else any available timeframe.
    """
    info_map = discovered.get(asset, {})
    if not info_map:
        return asset

    # Prefer the 15-min feed for the identifier; fall back to any available
    tf_pref = [15, 3, 5]
    path = None
    for tf in tf_pref:
        if tf in info_map:
            path = info_map[tf]["path"]
            break
    if path is None:
        return asset

    name = path.stem.upper()           # e.g. CME_MINI_NQM2026__15
    # Match contract letter + 4-digit year
    m = re.search(r'([HMUZ])(\d{4})', name)
    if m:
        letter = m.group(1)
        # Extract asset base from name
        for base in ["MNQ", "NQ", "MES", "ES", "RTY"]:
            if base in name:
                # Use the non-M-prefixed base for the label
                clean_base = base.lstrip("M") if base.startswith("M") and base != "MES" else base
                clean_base = "ES" if base == "MES" else clean_base
                clean_base = "NQ" if base in ("MNQ", "NQ") else clean_base
                clean_base = base if base == "RTY" else clean_base
                return f"{clean_base}{letter}"

    # Check for composite (NQ1!, ES1!, etc.)
    if "1_" in name or "1!" in name or "NQ1" in name:
        return f"{asset}1_"

    return asset




def detect_timestamp_type(df: pd.DataFrame) -> str:
    """
    Inspect first data row of column A.
    Returns 'unix' (seconds since epoch > 1e9) or 'iso'.
    """
    try:
        val = float(df.iloc[0, 0])
        if val > 1_000_000_000:
            return "unix"
    except (ValueError, TypeError):
        pass
    return "iso"


def insert_ny_datetime_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    Auto-detect Unix vs ISO in column A.
    Insert a 'datetime_ny' column immediately after column A with timezone-naive NY datetimes.
    EDT = UTC−4.  Uses pytz/dateutil via pandas tz_convert for proper DST handling.
    """
    df = df.copy()
    ts_type = detect_timestamp_type(df)
    time_col = df.columns[0]

    if ts_type == "unix":
        print(f"    🕐 Unix timestamps detected in column A → converting to NY (EDT/EST)")
        ny_times = (
            pd.to_datetime(df[time_col], unit="s", utc=True)
            .dt.tz_convert("America/New_York")
            .dt.tz_localize(None)
        )
    else:
        print(f"    🕐 ISO timestamps detected in column A → converting to clean NY datetime")
        raw = pd.to_datetime(df[time_col], utc=True, errors="coerce")
        ny_times = raw.dt.tz_convert("America/New_York").dt.tz_localize(None)

    df.insert(1, "datetime_ny", ny_times)
    print(f"    ✅ datetime_ny column inserted  "
          f"(range: {ny_times.min()}  →  {ny_times.max()})")
    return df


def load_and_normalize_csv(csv_path: Path) -> pd.DataFrame:
    """
    Load a feed CSV, insert NY datetime column, sort by time.
    """
    df = pd.read_csv(str(csv_path))
    df = insert_ny_datetime_column(df)
    df = df.sort_values("datetime_ny").reset_index(drop=True)
    return df


# ============================================================================
# SECTION 5: MULTI-CHART OHLC CONSTRUCTION
# ============================================================================

# Anchor datetime used for resample origin (18:00 on a Monday)
_RESAMPLE_ORIGIN_18 = pd.Timestamp("2000-01-03 18:00:00")
_RESAMPLE_ORIGIN_17 = pd.Timestamp("2000-01-03 17:00:00")


def detect_session_open_hour(df: pd.DataFrame) -> int:
    """
    Look at the datetime_ny column: return 17 if 17:00 bars exist more than 18:00, else 18.
    """
    times = df["datetime_ny"].dt.time
    count_18 = (df["datetime_ny"].dt.hour == 18).sum()
    count_17 = (df["datetime_ny"].dt.hour == 17).sum()
    return 17 if count_17 > count_18 else 18


def build_higher_tf_ohlc(df_source: pd.DataFrame,
                          interval_minutes: int,
                          session_open_hour: int = 18,
                          lookback_days: int = 10) -> pd.DataFrame:
    """
    Resample OHLC (and MAs if present) from df_source to interval_minutes bars.

    Parameters
    ----------
    df_source       : must have 'datetime_ny' and OHLC columns
    interval_minutes: target bar size in minutes
    session_open_hour: 17 or 18 (determines resample origin anchor)
    lookback_days   : limit data to this many calendar days

    Returns DataFrame with columns: datetime_ny, open, high, low, close, [ma cols...]
    """
    df = df_source.copy()
    df = df.set_index("datetime_ny").sort_index()

    # Limit lookback
    cutoff = df.index.max() - pd.Timedelta(days=lookback_days)
    df = df[df.index >= cutoff]

    if df.empty:
        return pd.DataFrame()

    origin = _RESAMPLE_ORIGIN_17 if session_open_hour == 17 else _RESAMPLE_ORIGIN_18
    freq = f"{interval_minutes}min"

    # OHLC aggregation
    agg_dict = {
        "open":  "first",
        "high":  "max",
        "low":   "min",
        "close": "last",
    }

    # Include MA columns (carry 'last' value of the window)
    ma_cols = [c for c in df.columns if c not in ("open", "high", "low", "close", "time",
                                                    "Volume", "volume", "RSI-8")]
    for mc in ma_cols:
        agg_dict[mc] = "last"

    resampled = (
        df.resample(freq, closed="left", label="left", origin=origin)
        .agg(agg_dict)
        .dropna(subset=["open", "close"])
    )

    resampled = resampled.reset_index().rename(columns={"index": "datetime_ny",
                                                          "datetime_ny": "datetime_ny"})
    # Ensure column name is correct after reset_index
    if resampled.columns[0] != "datetime_ny":
        resampled = resampled.rename(columns={resampled.columns[0]: "datetime_ny"})

    return resampled


def build_all_higher_timeframes(df_3min: pd.DataFrame,
                                 df_5min: pd.DataFrame,
                                 df_15min_actual: pd.DataFrame,
                                 asset_id: str,
                                 session_open_hour: int = 18,
                                 lookback_days: int = 10) -> Dict[int, pd.DataFrame]:
    """
    Build all derived timeframe DataFrames.
    Returns dict keyed by interval_minutes.
    Also prints QC report comparing reconstructed 15-min vs actual.
    """
    print(f"\n  📊 Building higher-timeframe OHLC charts for {asset_id}...")
    charts = {}

    # From 3-min
    for mins in BUILD_FROM_3MIN:
        tf_df = build_higher_tf_ohlc(df_3min, mins, session_open_hour, lookback_days)
        charts[mins] = tf_df
        print(f"    {mins:>3d}-min (from 3-min): {len(tf_df):>5d} bars")

    # From 5-min
    for mins in BUILD_FROM_5MIN:
        tf_df = build_higher_tf_ohlc(df_5min, mins, session_open_hour, lookback_days)
        charts[mins] = tf_df
        print(f"    {mins:>3d}-min (from 5-min): {len(tf_df):>5d} bars")

    # QC: reconstruct 15-min from 3-min and 5-min
    qc_from_3 = build_higher_tf_ohlc(df_3min, 15, session_open_hour, lookback_days)
    qc_from_5 = build_higher_tf_ohlc(df_5min, 15, session_open_hour, lookback_days)
    _run_15min_qc(df_15min_actual, qc_from_3, qc_from_5, asset_id)

    # Store actual 15-min for reference
    charts[15] = df_15min_actual

    return charts


def _run_15min_qc(df_actual: pd.DataFrame,
                  df_from_3: pd.DataFrame,
                  df_from_5: pd.DataFrame,
                  asset_id: str,
                  tolerance: float = 0.5):
    """
    Compare reconstructed 15-min bars (from 3-min and from 5-min) against actual 15-min.
    Reports mismatches beyond tolerance on Close.
    """
    print(f"\n  🔬 15-min QC for {asset_id}  (tolerance: ±{tolerance} pts on Close)")

    def _prep(df):
        d = df.copy()
        # Normalize time col
        if "datetime_ny" in d.columns:
            d = d.set_index("datetime_ny")
        elif "time" in d.columns:
            d = d.set_index("time")
        return d[["open", "high", "low", "close"]].sort_index()

    actual  = _prep(df_actual)
    from_3  = _prep(df_from_3)
    from_5  = _prep(df_from_5)

    common_3 = actual.index.intersection(from_3.index)
    common_5 = actual.index.intersection(from_5.index)

    mismatches_3 = mismatches_5 = 0
    for ts in common_3:
        diff = abs(actual.loc[ts, "close"] - from_3.loc[ts, "close"])
        if diff > tolerance:
            mismatches_3 += 1
    for ts in common_5:
        diff = abs(actual.loc[ts, "close"] - from_5.loc[ts, "close"])
        if diff > tolerance:
            mismatches_5 += 1

    pct_ok_3 = 100 * (1 - mismatches_3 / max(len(common_3), 1))
    pct_ok_5 = 100 * (1 - mismatches_5 / max(len(common_5), 1))
    icon_3 = "✅" if mismatches_3 == 0 else "⚠️ "
    icon_5 = "✅" if mismatches_5 == 0 else "⚠️ "

    print(f"    {icon_3} vs 3-min reconstruction: {len(common_3)} bars compared, "
          f"{mismatches_3} mismatches ({pct_ok_3:.1f}% match)")
    print(f"    {icon_5} vs 5-min reconstruction: {len(common_5)} bars compared, "
          f"{mismatches_5} mismatches ({pct_ok_5:.1f}% match)")


# ============================================================================
# SECTION 6: MOVING AVERAGE NORMALIZATION
# ============================================================================

# h-series: extract name from column header (strip whitespace, take first token)
def _extract_h_name(col: str) -> Optional[str]:
    stripped = col.strip()
    if not stripped.lower().startswith("h"):
        return None
    tokens = stripped.split()
    return tokens[0].lower() if tokens else None


def normalize_ma_columns(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """
    Process moving average columns in df:
      1. Map all h-series columns to a canonical name (h1–h20, h4a, h4b, h7a, h7b)
      2. Verify duplicate h-columns carry identical values (QC)
      3. Drop duplicate columns keeping one canonical copy

    Returns (cleaned_df, name_map) where name_map is {original_col: canonical_name}
    """
    df = df.copy()
    name_map: Dict[str, str] = {}     # original col → canonical
    h_groups: Dict[str, List[str]] = {}  # canonical → [original cols...]

    for col in df.columns:
        h_name = _extract_h_name(col)
        if h_name is None:
            continue
        name_map[col] = h_name
        h_groups.setdefault(h_name, []).append(col)

    # QC duplicates
    qc_issues = []
    cols_to_drop = []
    rename_map = {}

    for canonical, orig_cols in h_groups.items():
        if len(orig_cols) == 1:
            rename_map[orig_cols[0]] = canonical
            continue

        # Multiple columns with same canonical name — verify identical values
        reference = df[orig_cols[0]]
        all_equal = all(
            df[c].fillna(0).round(4).equals(reference.fillna(0).round(4))
            for c in orig_cols[1:]
        )

        if not all_equal:
            qc_issues.append(f"⚠️  {canonical}: duplicate columns differ "
                              f"({', '.join(orig_cols)})")

        # Keep first occurrence, drop the rest
        rename_map[orig_cols[0]] = canonical
        for dup_col in orig_cols[1:]:
            cols_to_drop.append(dup_col)

    # Apply renames and drops
    df = df.drop(columns=cols_to_drop, errors="ignore")
    df = df.rename(columns=rename_map)

    if qc_issues:
        for issue in qc_issues:
            print(f"    {issue}")
    else:
        duplicated = [c for c in h_groups if len(h_groups[c]) > 1]
        if duplicated:
            print(f"    ✅ Duplicate MA columns verified identical: {', '.join(duplicated)}")

    return df, rename_map


def get_all_ma_column_names(df: pd.DataFrame) -> List[str]:
    """
    Return list of all MA column names present in df (h-series + numeric-period series).
    Excludes OHLC, time, datetime_ny, Volume, RSI columns.
    """
    exclude = {"time", "datetime_ny", "open", "high", "low", "close",
               "volume", "Volume", "RSI-8", "rsi-8"}
    return [c for c in df.columns if c not in exclude]


# ============================================================================
# SECTION 7: PER-ASSET RAW + MATCHED TRAVELER REPORTS
#            (wraps v3.0 core logic; handles NQ/ES/RTY with per-asset defaults)
# ============================================================================

def _preprocess_15min_feed(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare a 15-min feed DataFrame for the range calculator:
    ensure 'time' column is timezone-naive NY datetime.
    """
    df = df.copy()
    time_col = None
    for cand in ["datetime_ny", "time"]:
        if cand in df.columns:
            time_col = cand
            break
    if time_col is None:
        raise ValueError("Feed has no 'time' or 'datetime_ny' column")

    if time_col != "time":
        # Drop the original raw 'time' column (Unix integers) before renaming
        # datetime_ny → 'time'; otherwise two columns share the name and
        # pd.to_datetime() receives a DataFrame, causing "duplicate keys" error.
        if "time" in df.columns:
            df = df.drop(columns=["time"])
        df = df.rename(columns={time_col: "time"})

    df["time"] = pd.to_datetime(df["time"], errors="coerce")
    try:
        df["time"] = df["time"].dt.tz_localize(None)
    except TypeError:
        pass

    return df.sort_values("time").reset_index(drop=True)


def _generate_travelers_with_calculator(measurement_df, feed1_df, feed2_df,
                                         report_time, window_radius,
                                         input_value_at_start, allowed_origins=None,
                                         lookback_days: int = DEFAULT_LOOKBACK_DAYS):
    """Wrapper — filters feeds by origin and lookback, then calls apply_full_range_advanced."""
    # ── Lookback trim: slice feeds to the requested window before passing to calculator
    # (apply_full_range_advanced has no scope/lookback parameter — filtering is done here)
    if lookback_days:
        cutoff = report_time - pd.Timedelta(days=lookback_days)
        feed1_df = feed1_df[feed1_df["time"] >= cutoff].copy()
        feed2_df = feed2_df[feed2_df["time"] >= cutoff].copy()

    if allowed_origins:
        base_cols = {"time", "open", "high", "low", "close"}
        if "volume" in feed1_df.columns:
            base_cols.add("volume")

        filtered_cols = set(base_cols)
        for origin in allowed_origins:
            is_bracketed = False
            base_origin = origin
            bracket_idx = None
            if "[" in origin:
                parts = origin.split("[")
                base_origin = parts[0]
                bracket_part = parts[1].rstrip("]")
                if base_origin.lower() in ["wasp-12b", "macedonia"]:
                    is_bracketed = True
                    bracket_idx = bracket_part.lstrip("-")

            if is_bracketed:
                for sfx in [" H", " L", " C"]:
                    for variant in [base_origin, base_origin.upper(),
                                    base_origin.capitalize(), "WASP-12b"]:
                        col = f"{variant}{sfx}[{bracket_idx}]"
                        for df_ref in [feed1_df, feed2_df]:
                            if col in df_ref.columns:
                                filtered_cols.add(col)
            else:
                for sfx in [" H", " L", " C"]:
                    for df_ref in [feed1_df, feed2_df]:
                        col = origin + sfx
                        if col in df_ref.columns:
                            filtered_cols.add(col)

        f1 = feed1_df[[c for c in feed1_df.columns if c in filtered_cols]]
        f2 = feed2_df[[c for c in feed2_df.columns if c in filtered_cols]]
    else:
        f1, f2 = feed1_df, feed2_df

    result_df = apply_full_range_advanced(
        df=measurement_df,
        small_df=f1,
        report_time=report_time,
        window_radius=window_radius,
        day_start_hour=18,
        input_value_at_start=input_value_at_start,
        big_df=f2,
        run_model_g=False,
    )
    return result_df


def generate_asset_traveler_reports(asset_id: str,
                                     df_15min_small: pd.DataFrame,
                                     df_15min_big: Optional[pd.DataFrame],
                                     measurement_path: str,
                                     report_time: datetime,
                                     lookback_days: int = DEFAULT_LOOKBACK_DAYS,
                                     window_radius: int = 600,
                                     process_non_priority: bool = False):
    """
    Generate Priority traveler report for one asset using its 15-min feeds.
    For RTY (composite), df_15min_big is None — the same feed is used for both slots.
    Returns (priority_df, non_priority_df, combined_df)
    """
    print(f"\n  📊 Generating traveler reports for {asset_id}...")

    feed1 = _preprocess_15min_feed(df_15min_small)
    feed2 = _preprocess_15min_feed(df_15min_big if df_15min_big is not None else df_15min_small)

    measurement_df = pd.read_excel(measurement_path)

    # Determine input @ start
    day_start_hour = 18
    base_dt = datetime(report_time.year, report_time.month, report_time.day, day_start_hour)
    if report_time < base_dt:
        base_dt -= timedelta(days=1)

    start_rows = feed1[feed1["time"] == base_dt]
    if not start_rows.empty:
        input_at_start = float(start_rows.iloc[-1]["open"])
    else:
        before = feed1[feed1["time"] <= base_dt]
        input_at_start = float(before.iloc[-1]["open"]) if not before.empty else None

    # Detect origins
    all_origins = set()
    for df_f in [feed1, feed2]:
        for col in df_f.columns:
            if col.endswith(" H") or " H[" in col:
                origin = col[:-2] if col.endswith(" H") else col.replace(" H[", "[")
                all_origins.add(origin)

    priority_origins_lower = {o.lower().strip() for o in PRIORITY_ORIGINS}
    priority_found    = {o for o in all_origins if o.lower().strip() in priority_origins_lower}
    non_priority_found = {o for o in all_origins if o.lower().strip() not in priority_origins_lower}

    # Generate priority report
    priority_df = pd.DataFrame()
    if priority_found:
        _result = _generate_travelers_with_calculator(
            measurement_df, feed1, feed2, report_time,
            window_radius, input_at_start, priority_found,
            lookback_days=lookback_days
        )
        priority_df = _result if _result is not None else pd.DataFrame()

    # Generate non-priority (optional)
    non_priority_df = pd.DataFrame()
    if process_non_priority and non_priority_found:
        _result = _generate_travelers_with_calculator(
            measurement_df, feed1, feed2, report_time,
            window_radius, input_at_start, non_priority_found,
            lookback_days=lookback_days
        )
        non_priority_df = _result if _result is not None else pd.DataFrame()

    # Combined
    parts = [d for d in [priority_df, non_priority_df] if not d.empty]
    combined_df = pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

    print(f"    ✅ Priority: {len(priority_df)} entries  |  "
          f"Non-priority: {len(non_priority_df)} entries  |  "
          f"Combined: {len(combined_df)}")

    return priority_df, non_priority_df, combined_df


def _bracket_index(val) -> int:
    """
    Extract the integer index from a Day bracket string.
    Examples:  '[0]' → 0,  '[-1]' → -1,  '[-3]' → -3,  'anything else' → 0
    """
    m = re.search(r'\[(-?\d+)\]', str(val))
    return int(m.group(1)) if m else 0


def _filter_day_bracket_order(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove rows where Day2's bracket is *newer* (less negative) than Day1's bracket.

    Rule: bracket(Day2) must be ≤ bracket(Day1).
      OK  → Day1=[-1], Day2=[-1]  (same day, equal)
      OK  → Day1=[-1], Day2=[-2]  (Day2 is older)
      BAD → Day1=[-1], Day2=[0]   (Day2 is newer — drop)
      BAD → Day1=[-3], Day2=[-2]  (Day2 is newer — drop)
    """
    if df.empty:
        return df
    if "Day1" not in df.columns or "Day2" not in df.columns:
        return df

    day1_idx = df["Day1"].apply(_bracket_index)
    day2_idx = df["Day2"].apply(_bracket_index)
    return df[day2_idx <= day1_idx].reset_index(drop=True)


def process_strategic_zones_for_asset(combined_df: pd.DataFrame,
                                       report_time: datetime,
                                       max_spread: float = 3.0,
                                       asset_id: str = ""):
    """Run 23 trading models (bypass mode) — identical to v3.0 logic."""
    if not HAS_MODEL_DEFINITIONS or not HAS_BYPASS_MATCHER:
        return None
    if combined_df.empty:
        return None

    print(f"\n  🎯 Strategic Zones for {asset_id} ({len(combined_df)} travelers, spread≤{max_spread})...")
    sorted_models = sorted(MODELS.items(), key=lambda x: x[1]["number"])

    all_results = {
        "results": {}, "timings": {},
        "metadata": {"report_time": report_time, "max_spread": max_spread, "asset": asset_id}
    }

    t0 = time_module.time()
    for idx, (model_name, model_def) in enumerate(sorted_models):
        tm0 = time_module.time()
        pct = (idx + 1) / len(sorted_models) * 100
        print(f"    [{pct:5.1f}%] #{model_def['number']:2d} {model_name:30s}...", end="")
        try:
            all_matches = []
            for day_filter in ["[0]", "[-1]", "[-2]", "[-3]"]:
                cfg = {"pass1_ms": sorted(model_def.get("pass1", set())),
                       "pass2_ms": sorted(model_def.get("pass2", set())),
                       "day_filter": day_filter, "feed_selection": "both",
                       "check_recip": model_def.get("check_recip", False)}
                m = process_model_bypass_mode(combined_df, cfg, max_spread=max_spread)
                # ── Bracket-order guard ──────────────────────────────────────
                # Day2 must never be *newer* than Day1.
                # e.g. Day1=[-3], Day2=[0] is invalid — drop it.
                m = _filter_day_bracket_order(m)
                if not m.empty:
                    all_matches.append(m)

            if all_matches:
                matched = pd.concat(all_matches, ignore_index=True)
                matched["M_#s"] = matched.apply(
                    lambda r: f"{int(r['M1'])}, {int(r['M2'])}", axis=1)
                all_results["results"][model_name] = matched
                print(f" {len(matched):4d} matches")
            else:
                all_results["results"][model_name] = pd.DataFrame()
                print(f"    0 matches")
        except Exception as e:
            print(f" ERROR: {str(e)[:50]}")
            all_results["results"][model_name] = pd.DataFrame()
        all_results["timings"][model_name] = time_module.time() - tm0

    all_results["total_time"] = time_module.time() - t0
    return all_results


# ============================================================================
# SECTION 8: WICK ANALYSIS
# ============================================================================

def _determine_wick_windows(report_time: datetime,
                              df_12min: pd.DataFrame) -> List[Dict]:
    """
    Derive wick analysis windows from actual chart bar timestamps.

    Instead of calendar arithmetic (which fails over weekends / when the 12-min
    chart ends before the report date), we scan backward in df_12min to locate
    the real prior session boundaries.

    Rules:
      - If report_time == 18:00 exactly:
            one window: [first bar of prior session, last bar before 18:00]
      - If report_time > 18:00:
            window 1: [first bar >= current session 18:00, report_time]
            window 2: [first bar of prior session, last bar before 18:00]
    """
    h, m = report_time.hour, report_time.minute
    session_open = report_time.replace(hour=18, minute=0, second=0, microsecond=0)

    tcol = "datetime_ny" if "datetime_ny" in df_12min.columns else "time"
    times = df_12min[tcol].apply(parse_timestamp_naive).sort_values().reset_index(drop=True)

    # ── Locate prior session ─────────────────────────────────────────────────
    # Last bar strictly before session_open
    bars_before = times[times < session_open]
    if bars_before.empty:
        prior_end = session_open - timedelta(minutes=12)
    else:
        prior_end = bars_before.iloc[-1]

    # Find where the prior session starts: walk back up to 7 days for the nearest 18:00
    prior_session_start = None
    for days_back in range(1, 8):
        candidate = (prior_end.to_pydatetime() - timedelta(days=days_back)).replace(
            hour=18, minute=0, second=0, microsecond=0)
        candidate = pd.Timestamp(candidate)
        bars_in_prior = times[(times >= candidate) & (times <= prior_end)]
        if not bars_in_prior.empty:
            prior_session_start = bars_in_prior.iloc[0]
            break

    if prior_session_start is None:
        prior_session_start = bars_before.iloc[0] if not bars_before.empty else pd.Timestamp(session_open)

    prior_window = {
        "label": "Prior session",
        "start": prior_session_start.to_pydatetime() if hasattr(prior_session_start, "to_pydatetime") else prior_session_start,
        "end":   prior_end.to_pydatetime() if hasattr(prior_end, "to_pydatetime") else prior_end,
    }

    if h == 18 and m == 0:
        return [prior_window]
    else:
        # Current session: first bar >= session_open up to report_time
        current_bars = times[(times >= pd.Timestamp(session_open)) & (times <= pd.Timestamp(report_time))]
        current_start = current_bars.iloc[0] if not current_bars.empty else pd.Timestamp(session_open)
        return [
            {"label": "Current session",
             "start": current_start.to_pydatetime() if hasattr(current_start, "to_pydatetime") else current_start,
             "end":   report_time},
            prior_window,
        ]


def _find_secondary_candle(df_12min: pd.DataFrame, swing: Dict) -> Optional[pd.Series]:
    """
    Find the secondary nested swing candle within a swing's time window.

    For a DOWN swing (origin = local high):
        After the swing's deepest low, find the highest-high candle (the counter-bounce).
    For an UP swing (origin = local low):
        After the swing's highest high, find the lowest-low candle (the counter-pullback).
    """
    origin_time  = swing["from_time"]
    extreme_time = swing["to_time"]
    direction    = swing["direction"]
    extreme_price = swing["to_price"]

    tcol = "datetime_ny" if "datetime_ny" in df_12min.columns else "time"
    mask = (df_12min[tcol] > origin_time) & (df_12min[tcol] <= extreme_time)
    window = df_12min[mask].copy().reset_index(drop=True)

    if len(window) < 2:
        return None

    if direction == "Down":
        # Find the row with the minimum low (closest to the extreme)
        low_idx = window["low"].idxmin()
        low_time = window.loc[low_idx, tcol]
        # Look for counter-bounce AFTER the deepest low
        after_low = window[window[tcol] > low_time]
        if after_low.empty:
            return None
        return after_low.loc[after_low["high"].idxmax()]

    else:  # Up swing
        high_idx = window["high"].idxmax()
        high_time = window.loc[high_idx, tcol]
        after_high = window[window[tcol] > high_time]
        if after_high.empty:
            return None
        return after_high.loc[after_high["low"].idxmin()]


def _wick_ranges(candle: pd.Series) -> Dict[str, Optional[Tuple[float, float]]]:
    """
    Return upper and lower wick price ranges for a candle.
    Upper wick: (max(open,close), high)  — only if high > max(open,close)
    Lower wick: (low, min(open,close))   — only if low < min(open,close)
    Returns None for a wick if it is zero.
    """
    o, h, lo, c = candle["open"], candle["high"], candle["low"], candle["close"]
    body_top    = max(o, c)
    body_bottom = min(o, c)

    upper = (body_top, h)  if h  > body_top    else None
    lower = (lo, body_bottom) if lo < body_bottom else None
    return {"upper": upper, "lower": lower}


def _value_in_range(value: float, rng: Tuple[float, float]) -> bool:
    return rng[0] <= value <= rng[1]


def _find_mas_in_wick(candle: pd.Series,
                       ma_cols: List[str],
                       wick_rng: Tuple[float, float]) -> List[str]:
    """Return list of MA column names whose value for this candle falls in the wick range."""
    hits = []
    for col in ma_cols:
        val = candle.get(col, np.nan)
        if pd.isna(val):
            continue
        if _value_in_range(float(val), wick_rng):
            hits.append(col)
    return hits


def _find_travelers_in_wick(travelers_df: pd.DataFrame,
                              wick_rng: Tuple[float, float]) -> pd.DataFrame:
    """
    Return subset of travelers_df whose Output price falls within wick_rng.
    Looks for column 'Output' or 'Output1'.
    """
    if travelers_df is None or travelers_df.empty:
        return pd.DataFrame()

    out_col = "Output" if "Output" in travelers_df.columns else \
              "Output1" if "Output1" in travelers_df.columns else None
    if out_col is None:
        return pd.DataFrame()

    mask = travelers_df[out_col].apply(
        lambda v: not pd.isna(v) and _value_in_range(float(v), wick_rng)
    )
    return travelers_df[mask].copy()


def run_wick_analysis_for_asset(asset_id: str,
                                 df_12min: pd.DataFrame,
                                 travelers_df: pd.DataFrame,
                                 report_time: datetime,
                                 ma_cols: List[str]) -> pd.DataFrame:
    """
    Run full wick analysis for one asset using its 12-min chart.
    Returns a DataFrame of wick analysis results.
    """
    if not HAS_SWING_DETECTOR or df_12min.empty:
        return pd.DataFrame()

    print(f"\n  🕯️  Wick analysis: {asset_id}  (12-min chart, {len(df_12min)} bars)")

    windows = _determine_wick_windows(report_time, df_12min)
    all_rows = []

    tcol = "datetime_ny" if "datetime_ny" in df_12min.columns else "time"
    df_12min = df_12min.copy()
    df_12min[tcol] = df_12min[tcol].apply(parse_timestamp_naive)

    for win in windows:
        w_start, w_end = win["start"], win["end"]
        print(f"    Window [{win['label']}]: {w_start} → {w_end}")

        # Slice to window
        mask = (df_12min[tcol] >= w_start) & (df_12min[tcol] <= w_end)
        df_window = df_12min[mask].copy().reset_index(drop=True)

        if df_window.empty:
            print(f"      ⚠️  No 12-min data in this window — skipping.")
            continue

        # Detect nested swings within the window
        swings = detect_nested_swings(df_window.rename(columns={tcol: "time"}),
                                       min_swing_size=10,
                                       pullback_tolerance=5)

        if not swings:
            print(f"      ℹ️  No nested swings detected.")
            continue

        print(f"      Found {len(swings)} nested swings.")

        for swing in swings:
            # --- PRIMARY candle ---
            primary_time = parse_timestamp_naive(swing["from_time"])
            primary_mask = df_12min[tcol].apply(
                lambda t: parse_timestamp_naive(t) == primary_time)
            primary_rows = df_12min[primary_mask]
            if primary_rows.empty:
                continue
            primary_candle = primary_rows.iloc[0]

            # Determine role: swing direction tells us the role of MAs/travelers in wick
            # DOWN swing → high wick = resistance; UP swing → low wick = support
            role_upper = "Resistance" if swing["direction"] == "Down" else "Support"
            role_lower = "Support"    if swing["direction"] == "Down" else "Resistance"

            wicks = _wick_ranges(primary_candle)
            for wick_side, wick_rng in [("Upper", wicks["upper"]),
                                         ("Lower", wicks["lower"])]:
                if wick_rng is None:
                    continue
                role = role_upper if wick_side == "Upper" else role_lower

                mas_hit = _find_mas_in_wick(primary_candle, ma_cols, wick_rng)
                travelers_hit = _find_travelers_in_wick(travelers_df, wick_rng)

                if not mas_hit and travelers_hit.empty:
                    continue

                for ma in mas_hit:
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": primary_time,
                        "Candle_Type": "Primary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "MA",
                        "Name":        ma,
                        "Value":       round(float(primary_candle.get(ma, np.nan)), 4),
                        "Candle_O":    primary_candle["open"],
                        "Candle_H":    primary_candle["high"],
                        "Candle_L":    primary_candle["low"],
                        "Candle_C":    primary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

                for _, trav in travelers_hit.iterrows():
                    origin  = trav.get("Origin", "")
                    m_num   = trav.get("M #", trav.get("M1", ""))
                    out_val = trav.get("Output", trav.get("Output1", np.nan))
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": primary_time,
                        "Candle_Type": "Primary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "Traveler",
                        "Name":        f"{origin} M#{m_num}",
                        "Value":       round(float(out_val), 4) if not pd.isna(out_val) else np.nan,
                        "Candle_O":    primary_candle["open"],
                        "Candle_H":    primary_candle["high"],
                        "Candle_L":    primary_candle["low"],
                        "Candle_C":    primary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

            # --- SECONDARY candle ---
            tmp_df = df_12min.copy()
            if tcol != "datetime_ny":
                tmp_df = tmp_df.rename(columns={tcol: "datetime_ny"})
            secondary_candle = _find_secondary_candle(tmp_df, swing)

            if secondary_candle is None:
                continue

            sec_time = parse_timestamp_naive(
                secondary_candle.get("datetime_ny", secondary_candle.get("time", None)))
            wicks_s = _wick_ranges(secondary_candle)

            for wick_side, wick_rng in [("Upper", wicks_s["upper"]),
                                          ("Lower", wicks_s["lower"])]:
                if wick_rng is None:
                    continue
                # Secondary wick role: secondary is a counter-move that fails
                # For DOWN swing secondary (bounce high) → upper wick = resistance
                # For UP swing secondary (pullback low) → lower wick = support
                if swing["direction"] == "Down":
                    role = "Resistance" if wick_side == "Upper" else "Support"
                else:
                    role = "Support" if wick_side == "Lower" else "Resistance"

                mas_hit = _find_mas_in_wick(secondary_candle, ma_cols, wick_rng)
                travelers_hit = _find_travelers_in_wick(travelers_df, wick_rng)

                if not mas_hit and travelers_hit.empty:
                    continue

                for ma in mas_hit:
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": sec_time,
                        "Candle_Type": "Secondary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "MA",
                        "Name":        ma,
                        "Value":       round(float(secondary_candle.get(ma, np.nan)), 4),
                        "Candle_O":    secondary_candle["open"],
                        "Candle_H":    secondary_candle["high"],
                        "Candle_L":    secondary_candle["low"],
                        "Candle_C":    secondary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

                for _, trav in travelers_hit.iterrows():
                    origin  = trav.get("Origin", "")
                    m_num   = trav.get("M #", trav.get("M1", ""))
                    out_val = trav.get("Output", trav.get("Output1", np.nan))
                    all_rows.append({
                        "Asset":       asset_id,
                        "Window":      win["label"],
                        "Candle_Time": sec_time,
                        "Candle_Type": "Secondary",
                        "Swing_Dir":   swing["direction"],
                        "Swing_Size":  round(swing["swing_size"], 2),
                        "Wick_Side":   wick_side,
                        "Role":        role,
                        "Type":        "Traveler",
                        "Name":        f"{origin} M#{m_num}",
                        "Value":       round(float(out_val), 4) if not pd.isna(out_val) else np.nan,
                        "Candle_O":    secondary_candle["open"],
                        "Candle_H":    secondary_candle["high"],
                        "Candle_L":    secondary_candle["low"],
                        "Candle_C":    secondary_candle["close"],
                        "Wick_Low":    round(wick_rng[0], 4),
                        "Wick_High":   round(wick_rng[1], 4),
                    })

    result_df = pd.DataFrame(all_rows)
    print(f"    ✅ {asset_id} wick analysis: {len(result_df)} wick intersections found.")
    return result_df


# ============================================================================
# SECTION 9A: MATCHED TRAVELERS — COLUMN DERIVATION & STANDARDISED FORMAT
# ============================================================================
# Exact 29-column order that matches the reference MatchedTravelers file
MATCHED_COL_ORDER = [
    "Open", "Ref", "Output1", "Output2", "Prox", "Origin1", "Origin2", "Group",
    "M1", "M2", "R1", "R2", "M_#s", "Arrival_Order", "Match",
    "Tag1", "Tag2", "Family1", "Family2", "Families",
    "Arrival1", "Arrival2", "Day1", "Day2", "Arrival_Brackets",
    "Model", "Model_Number", "Recent", "Feed1", "Feed2",
]

# ── Per-column widths for MatchedTravelers sheets (template-derived) ─────────
# Columns not listed here use the Excel default width (~8.43).
# Letter positions map to MATCHED_COL_ORDER (Ref inserted after Open):
#   A=Open  B=Ref  C=Output1  D=Output2  E=Prox  F=Origin1  G=Origin2  H=Group
#   I=M1  J=M2  K=R1  L=R2  M=M_#s  N=Arrival_Order  O=Match
#   P=Tag1  Q=Tag2  R=Family1  S=Family2  T=Families
#   U=Arrival1  V=Arrival2  W=Day1  X=Day2  Y=Arrival_Brackets
#   Z=Model  AA=Model_Number  AB=Recent  AC=Feed1  AD=Feed2
MATCHED_COL_WIDTHS = {
    "A":  13.0,        # Open
    "B":  8.0,         # Ref  (Output1 − Open)
    "C":  9.88671875,  # Output1
    "E":  6.0,         # Prox
    "F":  13.0,        # Origin1
    "H":  7.6640625,   # Group
    "I":  5.21875,     # M1
    "M":  7.5546875,   # M_#s
    "N":  3.6640625,   # Arrival_Order
    "O":  13.0,        # Match
    "W":  5.77734375,  # Day1
    "Y":  9.6640625,   # Arrival_Brackets
    "Z":  13.0,        # Model
    "AA": 4.33203125,  # Model_Number
    "AB": 7.44140625,  # Recent
    "AC": 13.0,        # Feed1
    "AD": 7.33203125,  # Feed2
}

# ── Filter metadata: description and conditions for every standard filter ─────
# Used in row 2 of each filtered tab and in the Filter Summary tab.
FILTER_META = {
    "TP 1":   {"desc": "Test Pair 1",       "cond": "Today, (|27|,|57|)"},
    "TP 2":   {"desc": "Test Pair 2",       "cond": "Today, (|27|,|40|)"},
    "TCF 1":  {"desc": "Test Cross Feed 1", "cond": "Today; Model 4; Small Output1 = Big Output 1"},
    "Tco 1a": {"desc": "Test Combo 1a",     "cond": "Today; Model 7; Grp: 1. SAA; Tag1 X0"},
    "Tco 1b": {"desc": "Test Combo 1b",     "cond": "Today; Model 7; Grp: 2 or 3; Tag1 X0"},
    "Tco 1c": {"desc": "Test Combo 1c",     "cond": "Today; Model 7; Grp: 1. SAA; Tag1: X1 and X2"},
    "Tco 1d": {"desc": "Test Combo 1d",     "cond": "Today; Model 7; Grp: 1. SAA. No tag filter"},
    "Tco 2d": {"desc": "Test Combo 2d",     "cond": "Model 7; Arrival_Brackets: [0], [-1]"},
}

_ANCHOR_SET = {"spain", "saturn", "jupiter", "kepler-62", "kepler-44"}
_TT_SET     = {"trinidad", "tobago"}


def _build_matched_combined_df(sz_results: Dict,
                                 priority_df: pd.DataFrame,
                                 measurement_df: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    """
    Concatenate all model match DataFrames, derive every standardised column,
    reorder to MATCHED_COL_ORDER, and return the finished DataFrame.
    Mirrors the full derivation logic from v3.0 export_split_reports_to_excel.
    """
    all_matched = []
    for mname, mdf in sz_results.get("results", {}).items():
        if mdf.empty:
            continue
        d = mdf.copy()
        model_info = MODELS.get(mname, {})
        if "Model" not in d.columns:
            d["Model"] = model_info.get("display_name", mname)
            d["Model_Number"] = model_info.get("number", 0)
        all_matched.append(d)

    if not all_matched:
        return None

    df = pd.concat(all_matched, ignore_index=True)

    # Rename Input→Output aliases if needed
    if "Input1" in df.columns and "Output1" not in df.columns:
        df["Output1"] = df["Input1"]
    if "Input2" in df.columns and "Output2" not in df.columns:
        df["Output2"] = df["Input2"]

    # Sort by Output1 descending
    if "Output1" in df.columns:
        df = df.sort_values("Output1", ascending=False).reset_index(drop=True)

    # ── Build Tag/Family lookup from RawTraveler priority_df ──────────────
    origin_m_to_tag    = {}
    origin_m_to_family = {}
    if priority_df is not None and not priority_df.empty:
        if "Tag" in priority_df.columns and "Family" in priority_df.columns:
            for _, row in priority_df.iterrows():
                o  = row.get("Origin", "")
                mn = row.get("M #", None)
                if o and mn is not None:
                    try:
                        key = (str(o), int(float(mn)))
                        origin_m_to_tag[key]    = row.get("Tag", "")
                        origin_m_to_family[key] = row.get("Family", "")
                    except Exception:
                        pass

    # M# → Tag fallback from measurement file
    m_to_tag = {}
    if measurement_df is not None and not measurement_df.empty:
        m_col   = next((c for c in measurement_df.columns
                        if c.lower().replace(" ", "") in ["m#", "m", "mnumber"]), None)
        tag_col = next((c for c in measurement_df.columns if "tag" in c.lower()), None)
        if m_col and tag_col:
            for _, row in measurement_df.iterrows():
                try:
                    m_to_tag[int(float(row[m_col]))] = str(row[tag_col])
                except Exception:
                    pass

    # ── Group ─────────────────────────────────────────────────────────────
    def _group(row):
        o1 = str(row.get("Origin1", "")).lower()
        o2 = str(row.get("Origin2", "")).lower()
        if not o1 or not o2:
            return ""
        a1, a2 = o1 in _ANCHOR_SET, o2 in _ANCHOR_SET
        t1, t2 = o1 in _TT_SET,     o2 in _TT_SET
        if a1 and a2 and o1 == o2: return "1. SAA"
        if t1 and t2 and o1 == o2: return "2. STT"
        if t1 and t2:              return "2. TT"
        if t1 and a2:              return "3. TA"
        if a1 and t2:              return "3. AT"
        if a1 and a2:              return "4. AA"
        if a2 and not a1:          return "5. oA"
        if a1 and not a2:          return "6. Ao"
        return "7. oo"

    df["Group"] = df.apply(_group, axis=1)

    # ── Arrival_Order ─────────────────────────────────────────────────────
    def _arr_order(row):
        try:
            m1a, m2a = abs(int(float(row["M1"]))), abs(int(float(row["M2"])))
            if m1a < 40 and m2a >= 40: return "PD"
            if m1a >= 40 and m2a < 40: return "DP"
            if m1a < 40 and m2a < 40:  return "DD"
            return "PP"
        except Exception:
            return ""

    df["Arrival_Order"] = df.apply(_arr_order, axis=1)

    # ── Match ─────────────────────────────────────────────────────────────
    def _match(row):
        try:
            m1, m2 = int(float(row["M1"])), int(float(row["M2"]))
            sign = "same" if (m1 > 0) == (m2 > 0) else "flip"
            m1a, m2a = abs(m1), abs(m2)
            if m1a > m2a:   direct = "up"
            elif m1a < m2a: direct = "down"
            else:            direct = "OPP" if sign == "flip" else "same"
            return f"{sign} {direct}"
        except Exception:
            return ""

    df["Match"] = df.apply(_match, axis=1)

    # ── Tag1 / Tag2 ───────────────────────────────────────────────────────
    def _tag(origin, m_val):
        if pd.isna(m_val): return ""
        m = int(float(m_val))
        key = (str(origin), m)
        if key in origin_m_to_tag:    return origin_m_to_tag[key]
        if m in m_to_tag:             return m_to_tag[m]
        m_abs = abs(m)
        if m_abs == 0:                return "Green"
        if m_abs in {40}:             return "Strength"
        if m_abs in {2,10,22,30,36,39,41,43,50,60,77,107}: return "Alpha"
        if m_abs in {5,14,55,68,96}:  return "Bravo"
        return "Other"

    df["Tag1"] = df.apply(lambda r: _tag(r.get("Origin1",""), r.get("M1")), axis=1)
    df["Tag2"] = df.apply(lambda r: _tag(r.get("Origin2",""), r.get("M2")), axis=1)

    # ── Family1 / Family2 / Families ─────────────────────────────────────
    def _family(origin, m_val):
        if pd.isna(m_val): return ""
        m = int(float(m_val))
        key = (str(origin), m)
        if key in origin_m_to_family: return origin_m_to_family[key]
        m_abs = abs(m)
        if m_abs >= 96:             return "1A"
        if 50 <= m_abs < 96:        return "1B"
        if 20 <= m_abs < 50:        return "2A"
        return "2B"

    df["Family1"]  = df.apply(lambda r: _family(r.get("Origin1",""), r.get("M1")), axis=1)
    df["Family2"]  = df.apply(lambda r: _family(r.get("Origin2",""), r.get("M2")), axis=1)
    df["Families"] = df["Family1"] + ", " + df["Family2"]

    # ── Recent ────────────────────────────────────────────────────────────
    def _recent(row):
        d1 = str(row.get("Day1", ""))
        if "[0]" in d1:  return "Today"
        if "["   in d1:  return "Recent"
        return ""

    df["Recent"] = df.apply(_recent, axis=1)

    # ── Open (within 3 pts of Input @ Start from priority_df) ─────────────
    df["Open"] = None
    if priority_df is not None and not priority_df.empty:
        input_col = next((c for c in priority_df.columns
                          if "input" in c.lower() and "start" in c.lower()), None)
        if input_col:
            starts = []
            if "Feed" in priority_df.columns:
                for feed_type in ("Small", "Big"):
                    rows = priority_df[priority_df["Feed"] == feed_type]
                    vals = rows[input_col].dropna()
                    if not vals.empty:
                        starts.append(float(vals.iloc[0]))
            else:
                vals = priority_df[input_col].dropna().unique()
                starts = sorted(float(v) for v in vals)

            if starts:
                for i, row in df.iterrows():
                    out = row.get("Output1") or row.get("Arrival_Output")
                    if out is None:
                        continue
                    try:
                        out = float(out)
                    except Exception:
                        continue
                    for sv in starts:
                        if abs(out - sv) <= 3:
                            df.at[i, "Open"] = sv
                            break

    # ── Drop temporary / unwanted columns ────────────────────────────────
    for drop_col in ("Input1", "Input2", "Arrival_Output"):
        if drop_col in df.columns:
            df = df.drop(columns=[drop_col])

    # ── Ref = Output1 − Open ──────────────────────────────────────────────
    # Populated only where both Output1 and Open are available.
    if "Output1" in df.columns:
        def _calc_ref(row):
            try:
                o1   = float(row["Output1"])
                open_ = float(row["Open"]) if row.get("Open") is not None else None
                if open_ is None:
                    return None
                return round(o1 - open_, 4)
            except (TypeError, ValueError):
                return None
        df["Ref"] = df.apply(_calc_ref, axis=1)
    else:
        df["Ref"] = None

    # ── Reorder to standard 29-column layout ─────────────────────────────
    ordered   = [c for c in MATCHED_COL_ORDER if c in df.columns]
    remaining = [c for c in df.columns if c not in ordered]
    df = df[ordered + remaining]

    return df


def _write_matched_travelers_excel(df: pd.DataFrame,
                                    fpath: str,
                                    report_time: datetime,
                                    asset_id: str):
    """
    Write the standardised MatchedTravelers Excel file:
      Row 1 — Report Time info
      Row 2 — match count info
      Row 3 — header (amber fill FFC000, black bold, centered)
      Row 4+ — data
      Freeze B4, auto-filter on row 3, per-column widths from MATCHED_COL_WIDTHS.

    Color scheme (v4.0.15):
      Origin1/Origin2 cells — per-origin background matching a_helpers:
        Spain/Saturn    → Neon Green  #39FF14
        Jupiter         → Light Blue  #D1ECF1
        Kepler-62/44    → Red Orange  #FF4D00
        Trinidad/Tobago → Gold        #F0CB59
        Wasp (any)      → Light Gray  #C0C0C0
        Macedonia (any) → Magenta     #E022D7
      M1/M2/R1/R2 font — positive → red #FF0000, negative → blue #0070C0
      Arrival_Brackets  — [0] → light yellow #FFF9C4; [-1]/[-2]/[-3] → light blue #BBDEFB

    Filtered tabs (after Combined):
      TP 1   — Recent=Today, |M1|=27, |M2|=57
      TP 2   — Recent=Today, |M1|=27, |M2|=40
      TCF 1  — Model='Lrg Disc PD', Recent=Today, Feed1≠Feed2, Output1≈Output2
      Tco 1a — Model='Recips PD', Recent=Today, Group='1. SAA', Tag1 contains 'X0'
      Tco 1b — Model='Recips PD', Recent=Today, Group starts '2.'/'3.', Tag1 contains 'X0'
      Tco 1c — Model='Recips PD', Recent=Today, Group='1. SAA', Tag1 contains 'X1' or 'X2'
      Tco 1d — Model='Recips PD', Recent=Today, Group='1. SAA', no tag filter
      Tco 2d — Model='Recips PD', Arrival_Brackets contains '[0]' and '[-1]'
      Each filtered tab has its description in row 2.
      All filters (with or without results) appear in the 'Filter Summary' tab.
    """
    # ── Shared style objects ──────────────────────────────────────────────────
    amber_fill   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    black_bold   = Font(bold=True, color="000000")
    center       = Alignment(horizontal="center", vertical="center")
    yellow_fill  = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    ltblue_fill  = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    ltred_fill   = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    red_font     = Font(color="FF0000")
    blue_font    = Font(color="0070C0")

    # Per-origin fills (matches a_helpers color scheme)
    _ORIGIN_FILLS = {
        "spain":     PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid"),
        "saturn":    PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid"),
        "jupiter":   PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
        "kepler-62": PatternFill(start_color="FF4D00", end_color="FF4D00", fill_type="solid"),
        "kepler-44": PatternFill(start_color="FF4D00", end_color="FF4D00", fill_type="solid"),
        "trinidad":  PatternFill(start_color="F0CB59", end_color="F0CB59", fill_type="solid"),
        "tobago":    PatternFill(start_color="F0CB59", end_color="F0CB59", fill_type="solid"),
    }
    _WASP_FILL     = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    _MACE_FILL     = PatternFill(start_color="E022D7", end_color="E022D7", fill_type="solid")

    def _origin_fill(val):
        o = str(val).lower().strip() if val else ""
        if o in _ORIGIN_FILLS:
            return _ORIGIN_FILLS[o]
        if "wasp" in o:
            return _WASP_FILL
        if "macedonia" in o:
            return _MACE_FILL
        return None

    def _apply_data_colors(ws, header_row: int = 3):
        """Apply per-row coloring to data rows starting at header_row+1."""
        # Locate target columns by header name
        col_map = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                col_map[str(cell.value)] = col_idx

        o1_col  = col_map.get("Origin1")
        o2_col  = col_map.get("Origin2")
        m1_col  = col_map.get("M1")
        m2_col  = col_map.get("M2")
        r1_col  = col_map.get("R1")
        r2_col  = col_map.get("R2")
        ab_col  = col_map.get("Arrival_Brackets")
        ref_col = col_map.get("Ref")

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            # Origin colors
            for o_col in (o1_col, o2_col):
                if o_col:
                    cell = row[o_col - 1]
                    fill = _origin_fill(cell.value)
                    if fill:
                        cell.fill = fill

            # M1/M2/R1/R2 font: positive=red, negative=blue
            for m_col in (m1_col, m2_col, r1_col, r2_col):
                if m_col:
                    cell = row[m_col - 1]
                    try:
                        v = float(cell.value)
                        cell.font = red_font if v > 0 else (blue_font if v < 0 else cell.font)
                    except (TypeError, ValueError):
                        pass

            # Arrival_Brackets: [0]=light yellow, [-1]/[-2]/[-3]=light blue
            if ab_col:
                cell = row[ab_col - 1]
                val  = str(cell.value) if cell.value else ""
                if "[0]" in val:
                    cell.fill = yellow_fill
                elif any(x in val for x in ("[-1]", "[-2]", "[-3]")):
                    cell.fill = ltblue_fill

            # Ref = Output1 − Open:
            #   ≥ +3  → light red   (meaningfully above open)
            #   ≤ −3  → light blue  (meaningfully below open)
            #   else  → yellow      (within ±3 of open)
            if ref_col:
                cell = row[ref_col - 1]
                try:
                    v = float(cell.value)
                    if v >= 3:
                        cell.fill = ltred_fill
                    elif v <= -3:
                        cell.fill = ltblue_fill
                    else:
                        cell.fill = yellow_fill
                except (TypeError, ValueError):
                    pass

    def _write_filtered_tab(wb, tab_name: str, fdf: pd.DataFrame,
                             description: str = ""):
        """Write a single filtered-results tab with amber headers and data colors."""
        ws = wb.create_sheet(title=tab_name)
        # Row 1 — title (bold)
        ws["A1"] = f"{tab_name} — {len(fdf)} result(s)"
        ws["A1"].font = Font(bold=True, size=12)
        ws.row_dimensions[1].height = 15.6
        # Row 2 — filter description
        ws["A2"] = description
        ws["A2"].font = Font(italic=False)
        # Row 3 — amber column headers
        for ci, col_name in enumerate(fdf.columns, start=1):
            cell = ws.cell(row=3, column=ci, value=col_name)
            cell.fill      = amber_fill
            cell.font      = black_bold
            cell.alignment = center
        # Row 4+ — data
        for ri, row_data in enumerate(fdf.itertuples(index=False), start=4):
            for ci, val in enumerate(row_data, start=1):
                ws.cell(row=ri, column=ci, value=val)
        # Freeze C4 (Open + Ref visible while scrolling), autofilter row 3, per-column widths
        ws.freeze_panes = "C4"
        if fdf.shape[1] > 0:
            ws.auto_filter.ref = f"A3:{get_column_letter(fdf.shape[1])}3"
        for i in range(1, fdf.shape[1] + 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = MATCHED_COL_WIDTHS.get(letter, 8.43)
        _apply_data_colors(ws, header_row=3)

    def _add_filtered_tabs(wb, combined_df: pd.DataFrame):
        """Build all standard filtered tabs from the Combined data, then a summary tab."""
        d = combined_df.copy()

        # Coerce numeric columns safely
        for col in ("M1", "M2", "Output1", "Output2"):
            if col in d.columns:
                d[col] = pd.to_numeric(d[col], errors="coerce")

        def _str(col):
            return d[col].fillna("").astype(str) if col in d.columns else pd.Series([""] * len(d))

        recent_s   = _str("Recent")
        model_s    = _str("Model")
        group_s    = _str("Group")
        tag1_s     = _str("Tag1")
        feed1_s    = _str("Feed1")
        feed2_s    = _str("Feed2")
        ab_s       = _str("Arrival_Brackets")

        m1 = d.get("M1", pd.Series(dtype=float))
        m2 = d.get("M2", pd.Series(dtype=float))
        o1 = d.get("Output1", pd.Series(dtype=float))
        o2 = d.get("Output2", pd.Series(dtype=float))

        today = (recent_s == "Today")

        filters = [
            ("TP 1",   today & (m1.abs() == 27) & (m2.abs() == 57)),
            ("TP 2",   today & (m1.abs() == 27) & (m2.abs() == 40)),
            ("TCF 1",  today & (model_s == "Lrg Disc PD")
                             & (feed1_s != feed2_s)
                             & ((o1 - o2).abs() < 0.01)),
            ("Tco 1a", today & (model_s == "Recips PD")
                             & (group_s == "1. SAA")
                             & tag1_s.str.contains("X0", na=False)),
            ("Tco 1b", today & (model_s == "Recips PD")
                             & (group_s.str.startswith("2.") | group_s.str.startswith("3."))
                             & tag1_s.str.contains("X0", na=False)),
            ("Tco 1c", today & (model_s == "Recips PD")
                             & (group_s == "1. SAA")
                             & tag1_s.str.contains(r"X1|X2", na=False)),
            ("Tco 1d", today & (model_s == "Recips PD")
                             & (group_s == "1. SAA")),
            # Tco 2d: Model 7 (Recips PD), Arrival_Brackets contains [0] AND [-1]
            ("Tco 2d", (model_s == "Recips PD")
                             & ab_s.str.contains(r"\[0\]",  na=False, regex=True)
                             & ab_s.str.contains(r"\[-1\]", na=False, regex=True)),
        ]

        # Track result counts for summary tab
        filter_counts = {}
        for tab_name, mask in filters:
            filtered = d[mask].reset_index(drop=True)
            qty = len(filtered)
            filter_counts[tab_name] = qty
            meta = FILTER_META.get(tab_name, {})
            desc_row2 = f"{meta.get('desc', tab_name)}: {meta.get('cond', '')}"
            if not filtered.empty:
                _write_filtered_tab(wb, tab_name, filtered, description=desc_row2)

        # ── Filter Summary tab (all filters, always created) ────────────────
        nr = wb.create_sheet(title="Filter Summary")
        headers = ("Qty", "Filter", "Description", "Conditions", "Status")
        for ci, hdr in enumerate(headers, start=1):
            cell = nr.cell(row=1, column=ci, value=hdr)
            cell.fill = amber_fill
            cell.font = black_bold
        nr.freeze_panes = "A2"
        nr.auto_filter.ref = "A1:E1"
        nr.column_dimensions["A"].width = 6.0
        nr.column_dimensions["B"].width = 10.6640625
        nr.column_dimensions["C"].width = 18.0
        nr.column_dimensions["D"].width = 47.88671875
        nr.column_dimensions["E"].width = 22.0

        for ri, (tab_name, _) in enumerate(filters, start=2):
            meta  = FILTER_META.get(tab_name, {})
            qty   = filter_counts.get(tab_name, 0)
            status = f"{qty} result{'s' if qty != 1 else ''}" if qty > 0 else "No results found"
            nr.cell(row=ri, column=1, value=qty)
            nr.cell(row=ri, column=2, value=tab_name)
            nr.cell(row=ri, column=3, value=meta.get("desc", ""))
            nr.cell(row=ri, column=4, value=meta.get("cond", ""))
            nr.cell(row=ri, column=5, value=status)

    # ── Write base data to Combined sheet ────────────────────────────────────
    df.to_excel(fpath, index=False, sheet_name="Combined")

    wb = load_workbook(fpath)
    ws = wb["Combined"]

    # Insert 2 info rows at the top
    ws.insert_rows(1, 2)
    ws["A1"] = f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = f"All Models Combined | {len(df)} total matches"
    ws["A2"].font = Font(italic=True)

    # Header row is now row 3 — apply amber formatting
    for cell in ws[3]:
        cell.fill      = amber_fill
        cell.font      = black_bold
        cell.alignment = center

    # Freeze C4 (Open + Ref visible while scrolling), auto-filter on row 3, per-column widths from MATCHED_COL_WIDTHS
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = MATCHED_COL_WIDTHS.get(letter, 8.43)

    # Format Arrival1 / Arrival2 cells as datetime
    dt_fmt = "YYYY-MM-DD HH:MM:SS"
    arr_cols = [i for i, cell in enumerate(ws[3], start=1)
                if str(cell.value) in ("Arrival1", "Arrival2")]
    for col_idx in arr_cols:
        letter = get_column_letter(col_idx)
        for row in range(4, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = dt_fmt

    # Apply data-row coloring to Combined sheet
    _apply_data_colors(ws, header_row=3)

    # Add filtered result tabs
    _add_filtered_tabs(wb, df)

    wb.save(fpath)


# ============================================================================
# SECTION 9: EXCEL EXPORT
# ============================================================================

HEADER_FILLS = {
    "NQ":  "4472C4",   # blue
    "ES":  "70AD47",   # green
    "RTY": "FFC000",   # amber
    "wick": "C00000",  # red
}


def _fmt_ws_header(ws, header_color: str = "4472C4", freeze: str = "A2"):
    """Apply blue header formatting to row 1 of a worksheet."""
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = freeze
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _auto_col_width(ws, max_width: int = 40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(best + 2, max_width)


def export_all_results(asset_results: Dict,
                        wick_results: Dict[str, pd.DataFrame],
                        report_time: datetime,
                        output_dir: str = ".",
                        contract_ids: Optional[Dict[str, str]] = None) -> Dict[str, str]:
    """
    Write Excel outputs:
      - RawTraveler_Priority_{CONTRACT}_{TS}.xlsx   (per asset)
      - MatchedTravelers_{CONTRACT}_{TS}.xlsx        (per asset)
      - {CONTRACT_PREFIX}WickAnalysis_{TS}.xlsx      (all assets, one tab each)

    contract_ids: dict mapping asset → short contract label, e.g. {"NQ": "NQM", "ES": "ESH"}
                  If None, falls back to bare asset name (old behaviour).
    """
    print("\n" + "=" * 120)
    print("📥 EXPORTING TO EXCEL")
    print("=" * 120)

    ts = report_time.strftime("%Y%m%d_%H%M")
    created = {}
    if contract_ids is None:
        contract_ids = {}

    # ------------------------------------------------------------------ #
    # Per-asset Raw + Matched
    # ------------------------------------------------------------------ #
    for asset_id, data in asset_results.items():
        priority_df  = data.get("priority_df",  pd.DataFrame())
        sz_results   = data.get("sz_results",   None)
        color        = HEADER_FILLS.get(asset_id, "4472C4")
        cid          = contract_ids.get(asset_id, asset_id)   # e.g. "NQM"

        # --- RawTraveler Priority ---
        if not priority_df.empty:
            fname = f"RawTraveler_Priority_{cid}_{ts}.xlsx"
            fpath = os.path.join(output_dir, fname)
            try:
                priority_df.to_excel(fpath, index=False, sheet_name="Priority_Origins")
                wb = load_workbook(fpath)
                ws = wb["Priority_Origins"]
                _fmt_ws_header(ws, color)
                _auto_col_width(ws)
                wb.save(fpath)
                created[f"RawTraveler_{asset_id}"] = fpath
                print(f"  ✅ {fname}  ({len(priority_df)} rows)")
            except Exception as e:
                print(f"  ❌ Error writing {fname}: {e}")

        # --- MatchedTravelers (standardised v3.0 format) ---
        if sz_results and sz_results.get("results"):
            fname = f"MatchedTravelers_{cid}_{ts}.xlsx"
            fpath = os.path.join(output_dir, fname)
            try:
                combined = _build_matched_combined_df(
                    sz_results, priority_df, data.get("measurement_df"))
                if combined is not None and not combined.empty:
                    _write_matched_travelers_excel(
                        combined, fpath, report_time, asset_id)
                    created[f"Matched_{asset_id}"] = fpath
                    print(f"  ✅ {fname}  ({len(combined)} matched rows)")
                else:
                    print(f"  ℹ️  {asset_id}: no matched travelers to export.")
            except Exception as e:
                print(f"  ❌ Error writing {fname}: {e}")
                import traceback; traceback.print_exc()

    # ------------------------------------------------------------------ #
    # WickAnalysis (all assets, one tab per asset)
    # Prefix with contract/composite label when known
    # ------------------------------------------------------------------ #
    has_wick = any(not df.empty for df in wick_results.values())
    if has_wick:
        # Build a prefix from contract IDs (e.g. "NQM_ESH_RTY")
        cid_list = [contract_ids.get(a, a) for a in ["NQ", "ES", "RTY"]
                    if not wick_results.get(a, pd.DataFrame()).empty]
        cid_prefix = "_".join(cid_list) + "_" if cid_list else ""
        fname = f"{cid_prefix}WickAnalysis_{ts}.xlsx"
        fpath = os.path.join(output_dir, fname)
        try:
            with pd.ExcelWriter(fpath, engine="openpyxl") as writer:
                for asset_id, wdf in wick_results.items():
                    if wdf.empty:
                        continue
                    sheet_name = f"{asset_id}_Wick"
                    wdf.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply formatting
            wb = load_workbook(fpath)
            for asset_id, wdf in wick_results.items():
                if wdf.empty:
                    continue
                ws = wb[f"{asset_id}_Wick"]
                color = HEADER_FILLS.get(asset_id, "C00000")
                _fmt_ws_header(ws, color)
                _auto_col_width(ws)

                # Highlight Resistance rows orange, Support rows green
                res_fill  = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                sup_fill  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

                role_col_idx = None
                for i, cell in enumerate(ws[1], start=1):
                    if str(cell.value) == "Role":
                        role_col_idx = i
                        break

                if role_col_idx:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        role_val = row[role_col_idx - 1].value
                        fill = res_fill if role_val == "Resistance" else \
                               sup_fill if role_val == "Support" else None
                        if fill:
                            for cell in row:
                                cell.fill = fill

            wb.save(fpath)
            total_wick_rows = sum(len(d) for d in wick_results.values())
            created["WickAnalysis"] = fpath
            print(f"  ✅ {fname}  ({total_wick_rows} wick intersection rows, "
                  f"{sum(1 for d in wick_results.values() if not d.empty)} asset tabs)")
        except Exception as e:
            print(f"  ❌ Error writing {fname}: {e}")
            import traceback; traceback.print_exc()
    else:
        print("  ℹ️  No wick intersections found — WickAnalysis.xlsx skipped.")

    return created


# ============================================================================
# SECTION 10: MAIN PIPELINE ORCHESTRATOR
# ============================================================================

def parse_report_datetime(s: str) -> datetime:
    for fmt in ["%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M", "%Y/%m/%d %H:%M"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse datetime: '{s}'.  Use YYYY-MM-DD HH:MM")


def run_pipeline_v4(search_dir: str,
                    measurement_path: str,
                    report_datetime_str: str,
                    lookback_days: int = DEFAULT_LOOKBACK_DAYS,
                    output_dir: str = ".",
                    process_non_priority: bool = False,
                    per_asset_overrides: Optional[Dict] = None,
                    mode: str = "standard",
                    end_datetime_str: Optional[str] = None,
                    ohlc_paths: Optional[Dict[str, str]] = None,
                    catalog_dir: str = ".",
                    ma_proximity: float = DEFAULT_MA_PROXIMITY,
                    tv_proximity: float = DEFAULT_TV_PROXIMITY,
                    swing_overrides: Optional[Dict] = None,
                    quick_mode: bool = True) -> Optional[Dict]:
    """
    Full Pipeline v4.0.15 execution.

    quick_mode=True  (default):
      Requires only 5×15-min CSVs + measurement file (6 total).
      Steps:
        1.  File discovery & validation (15-min only)
        2.  Load + normalize timestamps (15-min only)
        3.  Raw + Matched traveler reports per asset
        4.  Excel export

    quick_mode=False  (--full):
      Requires all 16 files (5×15-min, 5×3-min, 5×5-min + measurement).
      Steps:
        1.  File discovery & validation (all 16)
        2.  Load + normalize timestamps (all 15 CSVs)
        3.  Multi-chart construction + QC per asset
        4.  MA normalization per asset
        5.  Raw + Matched traveler reports per asset
        6.  Strategic zones per asset
        7.  Wick analysis per asset (12-min chart)
        8.  Excel export  (contract ID embedded in all filenames)
        9.  Backtest  (mode='backtest' or 'both')
        10. Forecast   (mode='forecast' or 'both')

    Parameters
    ----------
    quick_mode      : True (default) → 15-min only, no wick/MA; False → full 16-file run
    mode            : 'standard' | 'backtest' | 'forecast' | 'both'
    end_datetime_str: session end for backtest (YYYY-MM-DD HH:MM);
                      defaults to report_time + 23h (next RTH close)
    ohlc_paths      : dict {asset: path_to_5min_csv} for backtest/forecast;
                      if None, uses the 5-min CSVs already discovered in search_dir
    catalog_dir     : directory for backtest_catalog.json (default: current dir)
    ma_proximity    : pts — MA must be within this of extreme to be tagged
    tv_proximity    : pts — traveler Output1 must be within this of extreme
    """
    print("\n" + "=" * 120)
    print("🚀 PIPELINE v4.0.15 — EXECUTION START")
    print("=" * 120)
    print(f"  Started: {datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"  Search dir:  {search_dir}")
    print(f"  Measurement: {measurement_path}")
    print(f"  Mode:        {mode.upper()}")
    print(f"  Run style:   {'QUICK (15-min + measurement only)' if quick_mode else 'FULL (16 files, OHLC + MA + Wick)'}")

    # ── 1. File discovery ───────────────────────────────────────────────
    discovered, _ = discover_feed_files(search_dir)
    if not validate_file_set(discovered, measurement_path, quick_mode=quick_mode):
        return None

    # Derive contract IDs from discovered filenames
    contract_ids: Dict[str, str] = {
        asset: _get_contract_id(discovered, asset)
        for asset in ["NQ", "ES", "RTY"]
    }
    print(f"\n  🏷️  Contract IDs: " +
          "  |  ".join(f"{a} → {contract_ids[a]}" for a in ["NQ", "ES", "RTY"]))

    # ── Parse report time ───────────────────────────────────────────────
    report_time = parse_report_datetime(report_datetime_str)
    print(f"\n  📅 Report time: {report_time:%Y-%m-%d %H:%M}")

    # ── 2. Load & normalize CSVs ─────────────────────────────────────────
    print("\n" + "─" * 80)
    tfs_to_load = [15] if quick_mode else [15, 3, 5]
    print(f"📂 STEP 2: LOADING AND NORMALIZING TIMESTAMPS  "
          f"({'15-min only' if quick_mode else '15-min, 3-min, 5-min'})")
    print("─" * 80)

    loaded: Dict[str, Dict[int, pd.DataFrame]] = {}
    for asset in ["NQ", "ES", "RTY"]:
        loaded[asset] = {}
        for tf in tfs_to_load:
            if tf not in discovered.get(asset, {}):
                continue
            info = discovered[asset][tf]
            print(f"\n  Loading {asset} {tf}-min: {info['path'].name}")
            df = load_and_normalize_csv(info["path"])
            loaded[asset][tf] = df
            print(f"    {len(df)} rows loaded.")

    # ── 3. Multi-chart construction (full mode only) ─────────────────────
    charts: Dict[str, Dict[int, pd.DataFrame]] = {}
    if not quick_mode:
        print("\n" + "─" * 80)
        print("📊 STEP 3: MULTI-CHART OHLC CONSTRUCTION")
        print("─" * 80)
        for asset in ["NQ", "ES", "RTY"]:
            print(f"\n  🔧 {asset}")
            df3 = loaded[asset].get(3, pd.DataFrame())
            df5 = loaded[asset].get(5, pd.DataFrame())
            df15 = loaded[asset][15]
            if df3.empty or df5.empty:
                print(f"  ⚠️  {asset}: missing 3-min or 5-min feed — skipping chart construction.")
                charts[asset] = {15: df15}
                continue
            open_hr = detect_session_open_hour(df3)
            print(f"    Session open detected: {open_hr}:00")
            charts[asset] = build_all_higher_timeframes(
                df3, df5, df15, asset,
                session_open_hour=open_hr,
                lookback_days=lookback_days,
            )
    else:
        print("\n  ⚡ STEP 3 SKIPPED  (quick mode — no 3-min/5-min files)")
        for asset in ["NQ", "ES", "RTY"]:
            charts[asset] = {15: loaded[asset].get(15, pd.DataFrame())}

    # ── 4. MA normalization (full mode only) ─────────────────────────────
    normalized_ma_cols: Dict[str, List[str]] = {}
    if not quick_mode:
        print("\n" + "─" * 80)
        print("〰️  STEP 4: MOVING AVERAGE NORMALIZATION")
        print("─" * 80)
        for asset in ["NQ", "ES", "RTY"]:
            if 3 not in loaded[asset]:
                print(f"  ⚠️  {asset}: no 3-min feed — skipping MA normalization.")
                normalized_ma_cols[asset] = []
                continue
            print(f"\n  {asset} — 3-min feed")
            df3_norm, name_map = normalize_ma_columns(loaded[asset][3])
            loaded[asset][3] = df3_norm
            if 12 in charts[asset] and not charts[asset][12].empty:
                df12_norm, _ = normalize_ma_columns(charts[asset][12])
                charts[asset][12] = df12_norm
            normalized_ma_cols[asset] = get_all_ma_column_names(df3_norm)
            h_series = sorted([c for c in normalized_ma_cols[asset] if c.startswith("h")])
            other_ma = sorted([c for c in normalized_ma_cols[asset] if not c.startswith("h")])
            print(f"    h-series ({len(h_series)}): {', '.join(h_series[:12])}"
                  f"{'...' if len(h_series) > 12 else ''}")
            print(f"    Other MAs ({len(other_ma)}): {', '.join(other_ma[:6])}"
                  f"{'...' if len(other_ma) > 6 else ''}")
    else:
        print("\n  ⚡ STEP 4 SKIPPED  (quick mode — no MA files)")
        for asset in ["NQ", "ES", "RTY"]:
            normalized_ma_cols[asset] = []

    # ── 5 & 6. Traveler reports + Strategic zones per asset ─────────────
    print("\n" + "─" * 80)
    print("📈 STEPS 5–6: RAW TRAVELERS + STRATEGIC ZONES (per asset)")
    print("─" * 80)
    asset_results: Dict[str, Dict] = {}

    for asset in ["NQ", "ES", "RTY"]:
        defaults = ASSET_DEFAULTS[asset].copy()
        if per_asset_overrides and asset in per_asset_overrides:
            defaults.update(per_asset_overrides[asset])

        max_spread   = defaults["max_spread"]
        radius       = defaults["radius"]
        is_composite = defaults["composite"]

        print(f"\n{'='*60}")
        print(f"  ASSET: {asset}  {'[COMPOSITE — single-feed RTY]' if is_composite else ''}")
        print(f"  max-spread={max_spread}  radius={radius}")
        print(f"{'='*60}")

        df_15 = loaded[asset][15]
        df_15_small = df_15
        df_15_big   = None if is_composite else df_15

        if asset != "RTY":
            small_candidates = [
                p for p in Path(search_dir).glob("*.csv")
                if _classify_csv(p) and
                   _classify_csv(p)["asset"] == asset and
                   _classify_csv(p)["timeframe"] == 15 and
                   _classify_csv(p)["role"] == "small"
            ]
            big_candidates = [
                p for p in Path(search_dir).glob("*.csv")
                if _classify_csv(p) and
                   _classify_csv(p)["asset"] == asset and
                   _classify_csv(p)["timeframe"] == 15 and
                   _classify_csv(p)["role"] == "big"
            ]
            if small_candidates:
                print(f"  📄 Small feed: {small_candidates[0].name}")
                df_15_small = load_and_normalize_csv(small_candidates[0])
            if big_candidates:
                print(f"  📄 Big feed:   {big_candidates[0].name}")
                df_15_big = load_and_normalize_csv(big_candidates[0])
            if not small_candidates or not big_candidates:
                print(f"  ⚠️  Could not separate small/big 15-min feeds for {asset} — using single feed.")
                df_15_small = df_15
                df_15_big = df_15

        priority_df, non_priority_df, combined_df = generate_asset_traveler_reports(
            asset_id=asset,
            df_15min_small=df_15_small,
            df_15min_big=df_15_big,
            measurement_path=measurement_path,
            report_time=report_time,
            lookback_days=lookback_days,
            window_radius=radius,
            process_non_priority=process_non_priority,
        )

        sz_results = process_strategic_zones_for_asset(
            combined_df, report_time, max_spread=max_spread, asset_id=asset
        )

        asset_results[asset] = {
            "priority_df":     priority_df,
            "non_priority_df": non_priority_df,
            "combined_df":     combined_df,
            "sz_results":      sz_results,
            "measurement_df":  pd.read_excel(measurement_path) if os.path.exists(measurement_path) else None,
        }

    # ── 7. Wick analysis (full mode only) ───────────────────────────────
    wick_results: Dict[str, pd.DataFrame] = {}

    if not quick_mode:
        print("\n" + "─" * 80)
        print("🕯️  STEP 7: WICK ANALYSIS")
        print("─" * 80)

        for asset in ["NQ", "ES", "RTY"]:
            df_12 = charts[asset].get(12, pd.DataFrame())
            if df_12.empty:
                print(f"  ⚠️  {asset}: no 12-min chart available — skipping wick analysis.")
                wick_results[asset] = pd.DataFrame()
                continue

            travelers_df = asset_results[asset].get("priority_df", pd.DataFrame())
            ma_cols_12   = get_all_ma_column_names(df_12)

            wick_df = run_wick_analysis_for_asset(
                asset_id=asset,
                df_12min=df_12,
                travelers_df=travelers_df,
                report_time=report_time,
                ma_cols=ma_cols_12,
            )
            wick_results[asset] = wick_df
    else:
        print("\n  ⚡ STEP 7 SKIPPED  (quick mode — wick analysis requires --full)")
        for asset in ["NQ", "ES", "RTY"]:
            wick_results[asset] = pd.DataFrame()

    # ── 8. Export (with contract IDs in filenames) ───────────────────────
    print("\n" + "─" * 80)
    print("📥 STEP 8: EXPORT")
    print("─" * 80)
    created_files = export_all_results(
        asset_results=asset_results,
        wick_results=wick_results,
        report_time=report_time,
        output_dir=output_dir,
        contract_ids=contract_ids,
    )

    # ── 9 & 10. Backtest / Forecast ──────────────────────────────────────
    backtest_results: Dict[str, pd.DataFrame] = {}
    forecast_results: Dict[str, pd.DataFrame] = {}

    run_bt = mode in ("backtest", "both")
    run_fc = mode in ("forecast", "both")

    if run_bt or run_fc:
        print("\n" + "─" * 80)
        print(f"{'🔁 STEP 9: BACKTEST' if run_bt else ''}  "
              f"{'🔭 STEP 10: FORECAST' if run_fc else ''}")
        print("─" * 80)

        # Load catalog
        catalog_path = os.path.join(catalog_dir, DEFAULT_CATALOG_FILENAME)
        catalog      = _load_catalog(catalog_path)
        print(f"  📂 Catalog: {catalog_path}  "
              f"({len(catalog.get('combos', {}))} combos, "
              f"{len(catalog.get('ma_usage', {}))} MAs tracked)")

        # Determine session end for backtest
        if end_datetime_str:
            session_end = parse_report_datetime(end_datetime_str)
        else:
            # Default: next RTH close (16:45 on the calendar day report_time maps to)
            trading_day = (report_time + timedelta(days=1)).date() \
                          if report_time.hour >= 18 else report_time.date()
            session_end = datetime(trading_day.year, trading_day.month,
                                   trading_day.day, 16, 45)

        for asset in ["NQ", "ES", "RTY"]:
            # Resolve the 5-min OHLC source
            ohlc_path = None
            if ohlc_paths and asset in ohlc_paths:
                ohlc_path = ohlc_paths[asset]
            elif 5 in discovered.get(asset, {}):
                ohlc_path = str(discovered[asset][5]["path"])

            if not ohlc_path or not Path(ohlc_path).exists():
                print(f"  ⚠️  {asset}: no 5-min OHLC path available — skipping BT/FC.")
                backtest_results[asset] = pd.DataFrame()
                forecast_results[asset] = pd.DataFrame()
                continue

            # Load with NY timestamps, then normalize MA column names
            # (deduplicates leading-space variants and trailing-char mismatches)
            df_5 = load_and_normalize_csv(Path(ohlc_path))
            df_5, _ = normalize_ma_columns(df_5)
            all_ma_cols = get_all_ma_column_names(df_5)

            # Build matched DF for this asset (from strategic zones results)
            matched_df = pd.DataFrame()
            sz = asset_results[asset].get("sz_results")
            if sz and sz.get("results"):
                _built = _build_matched_combined_df(
                    sz,
                    asset_results[asset].get("priority_df", pd.DataFrame()),
                    asset_results[asset].get("measurement_df"),
                )
                matched_df = _built if _built is not None else pd.DataFrame()

            if run_bt:
                defaults = ASSET_DEFAULTS[asset].copy()
                if per_asset_overrides and asset in per_asset_overrides:
                    defaults.update(per_asset_overrides[asset])
                bt_df = run_backtest_for_asset(
                    asset_id=asset,
                    df_5min=df_5,
                    matched_df=matched_df,
                    session_start=report_time,
                    session_end=session_end,
                    ma_cols=all_ma_cols,
                    catalog=catalog,
                    swing_threshold=(swing_overrides or {}).get(
                        asset, BACKTEST_SWING_THRESHOLDS.get(asset)),
                    ma_proximity=ma_proximity,
                    tv_proximity=tv_proximity,
                )
                backtest_results[asset] = bt_df

            if run_fc:
                fc_df = run_forecast_for_asset(
                    asset_id=asset,
                    df_5min=df_5,
                    matched_df=matched_df,
                    as_of=report_time,
                    ma_cols=all_ma_cols,
                    catalog=catalog,
                    ma_proximity=ma_proximity,
                    tv_proximity=tv_proximity,
                )
                forecast_results[asset] = fc_df

        # Save updated catalog
        if run_bt:
            _save_catalog(catalog, catalog_path)
            print(f"\n  💾 Catalog updated: {catalog_path}  "
                  f"({len(catalog['combos'])} combos)")

        # Export backtest
        if run_bt:
            bt_fpath = export_backtest_results(
                backtest_results, report_time, output_dir, contract_ids)
            if bt_fpath:
                created_files["BacktestLog"] = bt_fpath

        # Export forecast
        if run_fc:
            fc_fpath = export_forecast_results(
                forecast_results, report_time, output_dir, contract_ids)
            if fc_fpath:
                created_files["Forecast"] = fc_fpath

    # ── Summary ──────────────────────────────────────────────────────────
    print("\n" + "=" * 120)
    print("✅ PIPELINE v4.0.15 COMPLETE")
    print("=" * 120)
    print(f"  Finished: {datetime.now():%Y-%m-%d %H:%M:%S}")
    print("\n  📁 Created files:")
    for key, path in created_files.items():
        print(f"    {key:<30s}: {path}")
    print()

    return {
        "asset_results":    asset_results,
        "wick_results":     wick_results,
        "charts":           charts,
        "created_files":    created_files,
        "backtest_results": backtest_results,
        "forecast_results": forecast_results,
        "contract_ids":     contract_ids,
    }


# ============================================================================
# SECTION 11: BACKTEST ENGINE
# ============================================================================
# Backtest mode processes a complete trading day (18:00 → next 16:45 RTH close)
# using the MA-enriched 5-min OHLC.  For each detected swing it:
#   • Classifies the extreme as SWING ZONE or DAILY TARGET
#   • Tags which MAs (h-series + others) sit within DEFAULT_MA_PROXIMITY
#   • Tags which Traveler Output1 values sit within DEFAULT_TV_PROXIMITY
# Results are written to BacktestLog_{CONTRACT}_{DATE}.xlsx and merged into
# the persistent JSON catalog for use by the Forecast engine.
# ============================================================================



def _load_catalog(catalog_path: str) -> Dict:
    """Load the JSON catalog, returning an empty structure if it doesn't exist."""
    if os.path.exists(catalog_path):
        try:
            with open(catalog_path, "r") as f:
                return json.load(f)
        except Exception:
            pass
    return {"combos": {}, "ma_usage": {}, "version": 1}


def _save_catalog(catalog: Dict, catalog_path: str) -> None:
    """Write catalog to disk (atomic via temp-rename)."""
    tmp = catalog_path + ".tmp"
    try:
        with open(tmp, "w") as f:
            json.dump(catalog, f, indent=2, default=str)
        os.replace(tmp, catalog_path)
    except Exception as e:
        print(f"  ⚠️  Catalog save error: {e}")


def _detect_swings_zigzag(df: pd.DataFrame,
                           min_swing: float,
                           tcol: str = "datetime_ny") -> List[Dict]:
    """
    Simple ZigZag swing detector on a 5-min OHLC DataFrame.

    Walks bars sequentially maintaining a current swing direction.  A new
    swing point is confirmed when price reverses by at least min_swing from
    the current extreme.

    Returns list of dicts with keys:
      time, price, direction ('High'|'Low'), bar_idx
    """
    if df.empty or len(df) < 3:
        return []

    highs  = df["high"].values
    lows   = df["low"].values
    times  = df[tcol].values

    points: List[Dict] = []
    # Seed with first bar
    last_high_idx  = 0
    last_low_idx   = 0
    last_high_val  = highs[0]
    last_low_val   = lows[0]
    direction      = None  # 'up' or 'down' — unknown until first reversal confirmed

    for i in range(1, len(df)):
        h, lo = highs[i], lows[i]

        if direction is None:
            # Determine initial direction
            if h > last_high_val:
                last_high_val, last_high_idx = h, i
            if lo < last_low_val:
                last_low_val, last_low_idx = lo, i
            move_up   = last_high_val - last_low_val
            move_down = last_high_val - last_low_val
            if move_up >= min_swing:
                direction = "up"
                points.append({"time": times[last_low_idx],
                                "price": last_low_val,
                                "direction": "Low",
                                "bar_idx": int(last_low_idx)})
            elif move_down >= min_swing:
                direction = "down"
                points.append({"time": times[last_high_idx],
                                "price": last_high_val,
                                "direction": "High",
                                "bar_idx": int(last_high_idx)})
        elif direction == "up":
            if h > last_high_val:
                last_high_val, last_high_idx = h, i
            elif last_high_val - lo >= min_swing:
                # Reversal confirmed downward
                points.append({"time": times[last_high_idx],
                                "price": last_high_val,
                                "direction": "High",
                                "bar_idx": int(last_high_idx)})
                direction  = "down"
                last_low_val, last_low_idx = lo, i
        else:  # direction == "down"
            if lo < last_low_val:
                last_low_val, last_low_idx = lo, i
            elif h - last_low_val >= min_swing:
                # Reversal confirmed upward
                points.append({"time": times[last_low_idx],
                                "price": last_low_val,
                                "direction": "Low",
                                "bar_idx": int(last_low_idx)})
                direction   = "up"
                last_high_val, last_high_idx = h, i

    # Append the final open swing point
    if direction == "up" and last_high_val not in [p["price"] for p in points[-1:]]:
        points.append({"time": times[last_high_idx],
                        "price": last_high_val,
                        "direction": "High",
                        "bar_idx": int(last_high_idx)})
    elif direction == "down" and last_low_val not in [p["price"] for p in points[-1:]]:
        points.append({"time": times[last_low_idx],
                        "price": last_low_val,
                        "direction": "Low",
                        "bar_idx": int(last_low_idx)})

    return points


def _classify_swing_point(swing: Dict,
                           df_5min: pd.DataFrame,
                           session_end: datetime,
                           tcol: str = "datetime_ny",
                           linger_window_min: int = 60,
                           linger_bars: int = 3,
                           linger_tolerance: float = 5.0) -> str:
    """
    Classify a swing extreme as 'SWING' or 'TARGET'.

    TARGET: ≥ linger_bars bars within linger_tolerance points of the extreme
            occur in the final linger_window_min minutes of the session.
    SWING:  everything else.
    """
    extreme  = swing["price"]
    cutoff   = pd.Timestamp(session_end) - pd.Timedelta(minutes=linger_window_min)
    end_ts   = pd.Timestamp(session_end)

    mask  = (df_5min[tcol] >= cutoff) & (df_5min[tcol] <= end_ts)
    final = df_5min[mask]

    count = ((final["high"] >= extreme - linger_tolerance) &
             (final["low"]  <= extreme + linger_tolerance)).sum()

    return "TARGET" if count >= linger_bars else "SWING"


def _tag_mas_at_extreme(swing: Dict,
                         df_5min: pd.DataFrame,
                         ma_cols: List[str],
                         proximity: float = DEFAULT_MA_PROXIMITY,
                         tcol: str = "datetime_ny") -> List[Dict]:
    """
    Return list of MAs whose value at the swing bar falls within proximity of
    the swing extreme price.  Each entry: {ma_name, ma_value, distance}.
    """
    extreme  = swing["price"]
    bar_time = pd.Timestamp(swing["time"])

    # Find the bar at or just before the swing time
    mask = df_5min[tcol] <= bar_time
    if not mask.any():
        return []
    row = df_5min[mask].iloc[-1]

    hits = []
    for col in ma_cols:
        val = row.get(col, np.nan)
        if pd.isna(val):
            continue
        dist = abs(float(val) - extreme)
        if dist <= proximity:
            hits.append({"ma_name": col, "ma_value": round(float(val), 4),
                         "distance": round(dist, 4)})

    hits.sort(key=lambda x: x["distance"])
    return hits


def _tag_travelers_at_extreme(swing: Dict,
                               matched_df: pd.DataFrame,
                               proximity: float = DEFAULT_TV_PROXIMITY,
                               top_groups: int = 3) -> List[Dict]:
    """
    Return MatchedTraveler rows whose Output1 is within proximity of the
    swing extreme, using a tiered group-rank cap instead of a flat 20-row limit.

    Tier caps (rows returned per group rank level):
      Group 1.SAA          → all matches (no cap; rarest, most meaningful)
      Groups 2.STT / 2.TT  → up to 10 per group
      Groups 3.TA / 3.AT   → up to 5 per group
      Groups 4+ (AA, oA…)  → up to 3 per group
    Only the top `top_groups` group-rank tiers are considered at all.
    """
    if matched_df is None or matched_df.empty:
        return []

    extreme = swing["price"]
    out_col = "Output1" if "Output1" in matched_df.columns else None
    if out_col is None:
        return []

    mask = matched_df[out_col].apply(
        lambda v: not pd.isna(v) and abs(float(v) - extreme) <= proximity)
    hits = matched_df[mask].copy()
    if hits.empty:
        return []

    # Group rank scoring
    group_rank = {g: i for i, g in enumerate(_GROUP_RANK_ORDER)}
    hits["_grp_rank"] = hits["Group"].map(lambda g: group_rank.get(g, 99)) \
        if "Group" in hits.columns else 99
    hits["_dist"] = (hits[out_col].astype(float) - extreme).abs()

    # Keep only top_groups tiers
    present_ranks = sorted(hits["_grp_rank"].unique())[:top_groups]
    hits = hits[hits["_grp_rank"].isin(present_ranks)]

    # Tiered cap per group rank
    tier_caps = {0: None, 1: 10, 2: 10, 3: 5, 4: 5, 5: 3, 6: 3, 7: 3, 8: 3}
    parts = []
    for rank in present_ranks:
        grp_rows = hits[hits["_grp_rank"] == rank].sort_values("_dist")
        cap = tier_caps.get(rank, 3)
        parts.append(grp_rows if cap is None else grp_rows.head(cap))
    hits = pd.concat(parts).drop(columns=["_grp_rank", "_dist"], errors="ignore")

    result = []
    for _, row in hits.iterrows():
        result.append({
            "output1":  round(float(row.get(out_col, np.nan)), 4),
            "origin1":  str(row.get("Origin1", "")),
            "origin2":  str(row.get("Origin2", "")),
            "m1":       row.get("M1", ""),
            "m2":       row.get("M2", ""),
            "group":    str(row.get("Group", "")),
            "match":    str(row.get("Match", "")),
            "model_no": row.get("Model_Number", ""),
            "recent":   str(row.get("Recent", "")),
            "feed1":    str(row.get("Feed1", "")),
            "feed2":    str(row.get("Feed2", "")),
            "distance": round(abs(float(row.get(out_col, 0)) - extreme), 4),
        })
    return result


def run_backtest_for_asset(asset_id: str,
                            df_5min: pd.DataFrame,
                            matched_df: pd.DataFrame,
                            session_start: datetime,
                            session_end: datetime,
                            ma_cols: List[str],
                            catalog: Dict,
                            swing_threshold: Optional[float] = None,
                            ma_proximity: float = DEFAULT_MA_PROXIMITY,
                            tv_proximity: float = DEFAULT_TV_PROXIMITY) -> pd.DataFrame:
    """
    Run backtest for one asset over a complete session window.

    Returns a DataFrame with one row per swing point × MA/TV tag combination.
    Also updates the catalog in-place.
    """
    threshold = swing_threshold or BACKTEST_SWING_THRESHOLDS.get(asset_id, 27.0)
    tcol = "datetime_ny" if "datetime_ny" in df_5min.columns else "time"

    print(f"\n  📉 Backtest {asset_id}  "
          f"({session_start:%Y-%m-%d %H:%M} → {session_end:%Y-%m-%d %H:%M})  "
          f"swing_threshold={threshold} pts")

    # Slice to session
    mask = (df_5min[tcol] >= pd.Timestamp(session_start)) & \
           (df_5min[tcol] <= pd.Timestamp(session_end))
    df_session = df_5min[mask].copy().reset_index(drop=True)

    if df_session.empty:
        print(f"    ⚠️  No data in session window — skipping.")
        return pd.DataFrame()

    # Detect swings
    swings = _detect_swings_zigzag(df_session, threshold, tcol)
    print(f"    Found {len(swings)} swing points.")

    if not swings:
        return pd.DataFrame()

    # Identify session HOD / LOD
    session_hod = df_session["high"].max()
    session_lod = df_session["low"].min()

    # Session open price — used to seed swing_size for the first swing point
    session_open_price = float(df_session.iloc[0]["open"])

    # All MA cols present in df_5min
    all_ma = [c for c in ma_cols if c in df_5min.columns]

    rows = []
    for swing in swings:
        extreme     = swing["price"]
        swing_dir   = swing["direction"]  # 'High' or 'Low'
        swing_time  = pd.Timestamp(swing["time"])

        is_hod = (swing_dir == "High" and abs(extreme - session_hod) < 0.01)
        is_lod = (swing_dir == "Low"  and abs(extreme - session_lod) < 0.01)
        zone_role = ("HOD" if is_hod else "LOD" if is_lod else
                     "Swing High" if swing_dir == "High" else "Swing Low")

        # Classify HOD/LOD; intermediate swings default to SWING
        if is_hod or is_lod:
            classification = _classify_swing_point(
                swing, df_session, session_end, tcol)
        else:
            classification = "SWING"

        # Tag MAs
        ma_tags = _tag_mas_at_extreme(swing, df_session, all_ma, ma_proximity, tcol)

        # Tag travelers
        tv_tags = _tag_travelers_at_extreme(swing, matched_df, tv_proximity)

        # Swing size: distance from prev swing to this one.
        # For the seed (first) swing use distance from session open price.
        swing_idx  = swings.index(swing)
        prev_swing = swings[swing_idx - 1] if swing_idx > 0 else None
        swing_size = (abs(extreme - prev_swing["price"]) if prev_swing
                      else abs(extreme - session_open_price))

        # HOD profile: breakout (no MA confluence) vs resistance-touch (MAs present)
        hod_profile = None
        if is_hod or is_lod:
            hod_profile = "BREAKOUT" if len(ma_tags) == 0 else "RESISTANCE-TOUCH"

        # Emit rows — one per MA tag and one per TV tag (if any)
        base = {
            "Asset":          asset_id,
            "Session_Start":  session_start,
            "Swing_Time":     swing_time,
            "Swing_Dir":      swing_dir,
            "Extreme_Price":  extreme,
            "Zone_Role":      zone_role,
            "Classification": classification,
            "HOD_Profile":    hod_profile,
            "Swing_Size":     round(swing_size, 2),
            "MA_Count":       len(ma_tags),
            "TV_Count":       len(tv_tags),
        }

        if ma_tags:
            for ma in ma_tags:
                row = base.copy()
                row.update({
                    "Tag_Type":   "MA",
                    "Tag_Name":   ma["ma_name"],
                    "Tag_Value":  ma["ma_value"],
                    "Tag_Dist":   ma["distance"],
                    "TV_M1":      None,
                    "TV_M2":      None,
                    "TV_Group":   None,
                    "TV_Match":   None,
                    "TV_Model":   None,
                    "TV_Recent":  None,
                    "TV_Feeds":   None,
                })
                rows.append(row)

        if tv_tags:
            for tv in tv_tags:
                row = base.copy()
                row.update({
                    "Tag_Type":   "Traveler",
                    "Tag_Name":   f"{tv['origin1']} / {tv['origin2']}",
                    "Tag_Value":  tv["output1"],
                    "Tag_Dist":   tv["distance"],
                    "TV_M1":      tv["m1"],
                    "TV_M2":      tv["m2"],
                    "TV_Group":   tv["group"],
                    "TV_Match":   tv["match"],
                    "TV_Model":   tv["model_no"],
                    "TV_Recent":  tv["recent"],
                    "TV_Feeds":   f"{tv['feed1']}/{tv['feed2']}",
                })
                rows.append(row)

        if not ma_tags and not tv_tags:
            # Emit bare swing point with no tags so it still appears in log
            row = base.copy()
            row.update({
                "Tag_Type": "None", "Tag_Name": None, "Tag_Value": None,
                "Tag_Dist": None,   "TV_M1": None,    "TV_M2": None,
                "TV_Group": None,   "TV_Match": None,  "TV_Model": None,
                "TV_Recent": None,  "TV_Feeds": None,
            })
            rows.append(row)

        # Update catalog for MA+TV combo rows
        for ma in ma_tags:
            for tv in tv_tags:
                combo_key = f"{ma['ma_name']}|{tv['m1']},{tv['m2']}|{tv['group']}"
                entry = catalog["combos"].setdefault(combo_key, {
                    "ma_name": ma["ma_name"], "m1": tv["m1"], "m2": tv["m2"],
                    "group": tv["group"], "match": tv["match"],
                    "asset": asset_id, "instances": 0,
                    "swing_swings": 0, "swing_targets": 0,
                    "swing_sizes": [], "mae_values": [],
                })
                entry["instances"] += 1
                if classification == "SWING":
                    entry["swing_swings"] += 1
                else:
                    entry["swing_targets"] += 1
                if swing_size > 0:
                    entry["swing_sizes"].append(round(swing_size, 2))

        # Update MA usage log (tracks how often each MA is used at swing zones)
        for ma in ma_tags:
            usage = catalog["ma_usage"].setdefault(ma["ma_name"], {
                "total": 0, "swing": 0, "target": 0,
                "hod": 0, "lod": 0, "other": 0,
            })
            usage["total"] += 1
            usage[classification.lower()] = usage.get(classification.lower(), 0) + 1
            if is_hod:   usage["hod"] += 1
            elif is_lod: usage["lod"] += 1
            else:        usage["other"] += 1

    result_df = pd.DataFrame(rows)
    n_swing = sum(1 for s in swings
                  if _classify_swing_point(s, df_session, session_end, tcol) == "SWING"
                  or not (abs(s["price"] - session_hod) < 0.01 or
                          abs(s["price"] - session_lod) < 0.01))
    print(f"    ✅ {len(result_df)} tagged rows  "
          f"({len([s for s in swings if s['direction']=='High'])} highs, "
          f"{len([s for s in swings if s['direction']=='Low'])} lows)")
    return result_df


def export_backtest_results(backtest_results: Dict[str, pd.DataFrame],
                             session_start: datetime,
                             output_dir: str,
                             contract_ids: Optional[Dict] = None) -> str:
    """Write BacktestLog_{CONTRACT}_{DATE}.xlsx with one tab per asset."""
    if contract_ids is None:
        contract_ids = {}

    date_str = session_start.strftime("%Y%m%d")
    cid_list = [contract_ids.get(a, a) for a in ["NQ", "ES", "RTY"]
                if a in backtest_results and not backtest_results[a].empty]
    cid_prefix = "_".join(cid_list) + "_" if cid_list else ""
    fname = f"{cid_prefix}BacktestLog_{date_str}.xlsx"
    fpath = os.path.join(output_dir, fname)

    has_data = any(not df.empty for df in backtest_results.values())
    if not has_data:
        print("  ℹ️  No backtest data to export.")
        return ""

    try:
        with pd.ExcelWriter(fpath, engine="openpyxl") as writer:
            for asset_id, df in backtest_results.items():
                if df.empty:
                    continue
                df.to_excel(writer, sheet_name=f"{asset_id}_BT", index=False)

        wb = load_workbook(fpath)
        for asset_id in backtest_results:
            if backtest_results[asset_id].empty:
                continue
            ws = wb[f"{asset_id}_BT"]
            color = HEADER_FILLS.get(asset_id, "4472C4")
            _fmt_ws_header(ws, color)
            _auto_col_width(ws)

            # Color-code Classification column: SWING=green, TARGET=amber
            cls_col_idx = None
            hod_col_idx = None
            for i, cell in enumerate(ws[1], start=1):
                if str(cell.value) == "Classification":
                    cls_col_idx = i
                if str(cell.value) == "HOD_Profile":
                    hod_col_idx = i
            if cls_col_idx:
                swing_fill  = PatternFill(start_color="C8E6C9", end_color="C8E6C9",
                                          fill_type="solid")
                target_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD",
                                          fill_type="solid")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    val = row[cls_col_idx - 1].value
                    f   = swing_fill if val == "SWING" else \
                          target_fill if val == "TARGET" else None
                    if f:
                        for cell in row:
                            cell.fill = f

            # Color-code HOD_Profile: RESISTANCE-TOUCH=blue, BREAKOUT=orange
            if hod_col_idx:
                rt_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE",
                                      fill_type="solid")   # light blue
                bo_fill = PatternFill(start_color="F4B942", end_color="F4B942",
                                      fill_type="solid")   # orange
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    val = row[hod_col_idx - 1].value
                    f   = rt_fill if val == "RESISTANCE-TOUCH" else \
                          bo_fill if val == "BREAKOUT" else None
                    if f:
                        row[hod_col_idx - 1].fill = f   # color only that cell

        wb.save(fpath)
        total_rows = sum(len(d) for d in backtest_results.values())
        print(f"  ✅ {fname}  ({total_rows} rows)")
        return fpath
    except Exception as e:
        print(f"  ❌ Error writing {fname}: {e}")
        import traceback; traceback.print_exc()
        return ""


# ============================================================================
# SECTION 12: FORECAST ENGINE
# ============================================================================
# Forecast mode uses the MA-enriched 5-min OHLC up to report_time to:
#   • Calculate velocity (pts/bar) for each MA over the last N bars
#   • For each MatchedTraveler zone (Output1), project how many bars/minutes
#     until each MA reaches within DEFAULT_MA_PROXIMITY of that zone
#   • Look up historical hit rate in the catalog for each (MA, M1, M2) combo
#   • Rank all (MA × Traveler) alert pairs by urgency and conviction
# Output: Forecast_{CONTRACT}_{TS}.xlsx
# ============================================================================

def _compute_ma_velocities(df_5min: pd.DataFrame,
                             ma_cols: List[str],
                             as_of: datetime,
                             lookback_bars: int = DEFAULT_VELOCITY_LOOKBACK,
                             tcol: str = "datetime_ny") -> Dict[str, Dict]:
    """
    For each MA column, compute current value + velocity (pts / 5-min bar)
    using a linear regression over the last lookback_bars bars ending at as_of.

    Returns dict: {ma_name: {value, velocity, direction ('Rising'|'Falling'|'Flat')}}
    """
    mask = df_5min[tcol] <= pd.Timestamp(as_of)
    df_slice = df_5min[mask].tail(lookback_bars + 1)
    if len(df_slice) < 2:
        return {}

    n = len(df_slice)
    x = np.arange(n, dtype=float)

    result = {}
    for col in ma_cols:
        if col not in df_slice.columns:
            continue
        y = df_slice[col].values.astype(float)
        valid = ~np.isnan(y)
        if valid.sum() < 2:
            continue
        # Simple linear regression slope
        xv, yv = x[valid], y[valid]
        slope = np.polyfit(xv, yv, 1)[0]  # pts per bar
        current_val = float(y[-1]) if not np.isnan(y[-1]) else float(yv[-1])
        direction = "Rising" if slope > 0.05 else "Falling" if slope < -0.05 else "Flat"
        result[col] = {
            "value":     round(current_val, 4),
            "velocity":  round(float(slope), 4),
            "direction": direction,
        }
    return result


def _project_ma_arrival(ma_info: Dict,
                         target_price: float,
                         proximity: float = DEFAULT_MA_PROXIMITY,
                         bar_minutes: int = 5,
                         max_bars_ahead: int = 96) -> Optional[Dict]:
    """
    Given MA current value and velocity, project how many bars until it
    reaches within proximity of target_price.

    Returns dict with bars_away, minutes_away, projected_price, confidence
    — or None if the MA is moving away from the target or already there.
    """
    current   = ma_info["value"]
    velocity  = ma_info["velocity"]
    dist      = target_price - current   # positive = target above MA

    # Already within proximity?
    if abs(dist) <= proximity:
        return {"bars_away": 0, "minutes_away": 0,
                "projected_price": current, "confidence": "At Zone"}

    # MA must be moving toward the target
    if velocity == 0:
        return None
    if (dist > 0 and velocity <= 0) or (dist < 0 and velocity >= 0):
        return None   # Moving away

    # Bars to reach proximity band edge
    edge_dist  = abs(dist) - proximity
    bars_to_edge = edge_dist / abs(velocity)

    if bars_to_edge > max_bars_ahead:
        return None

    projected = current + velocity * bars_to_edge
    confidence = ("High" if bars_to_edge <= 6 else
                  "Medium" if bars_to_edge <= 24 else "Low")

    return {
        "bars_away":       round(bars_to_edge, 1),
        "minutes_away":    round(bars_to_edge * bar_minutes, 0),
        "projected_price": round(projected, 4),
        "confidence":      confidence,
    }


def run_forecast_for_asset(asset_id: str,
                            df_5min: pd.DataFrame,
                            matched_df: pd.DataFrame,
                            as_of: datetime,
                            ma_cols: List[str],
                            catalog: Dict,
                            ma_proximity: float = DEFAULT_MA_PROXIMITY,
                            tv_proximity: float = DEFAULT_TV_PROXIMITY,
                            velocity_lookback: int = DEFAULT_VELOCITY_LOOKBACK,
                            top_groups: int = 3) -> pd.DataFrame:
    """
    Generate ranked MA×Traveler convergence alerts for one asset.

    Returns a DataFrame sorted by urgency (bars_away asc) then conviction
    (catalog_hit_rate desc, group_rank asc).
    """
    tcol = "datetime_ny" if "datetime_ny" in df_5min.columns else "time"
    print(f"\n  🔭 Forecast {asset_id}  as-of {as_of:%Y-%m-%d %H:%M}")

    # Slice data up to as_of
    mask     = df_5min[tcol] <= pd.Timestamp(as_of)
    df_slice = df_5min[mask].copy()
    if df_slice.empty:
        print(f"    ⚠️  No data up to {as_of} — skipping.")
        return pd.DataFrame()

    # 1. Compute MA velocities
    all_ma_cols = [c for c in ma_cols if c in df_5min.columns]
    velocities  = _compute_ma_velocities(df_slice, all_ma_cols, as_of,
                                          velocity_lookback, tcol)
    print(f"    {len(velocities)} MAs with valid velocity  "
          f"({sum(1 for v in velocities.values() if v['direction']=='Rising')} rising, "
          f"{sum(1 for v in velocities.values() if v['direction']=='Falling')} falling)")

    if not velocities:
        return pd.DataFrame()

    # 2. Build traveler zone list (top-group filtered, deduplicated by Output1 cluster)
    if matched_df is None or matched_df.empty:
        print("    ⚠️  No matched traveler data — skipping TV layer.")
        tv_zones: List[Dict] = []
    else:
        out_col = "Output1" if "Output1" in matched_df.columns else None
        if out_col is None:
            tv_zones = []
        else:
            group_rank = {g: i for i, g in enumerate(_GROUP_RANK_ORDER)}
            df_tv = matched_df.copy()
            if "Group" in df_tv.columns:
                df_tv["_grp_rank"] = df_tv["Group"].map(lambda g: group_rank.get(g, 99))
                max_rank = sorted(df_tv["_grp_rank"].unique())[:top_groups][-1] \
                           if not df_tv.empty else 99
                df_tv = df_tv[df_tv["_grp_rank"] <= max_rank]

            # Deduplicate by price cluster (merge rows within tv_proximity of each other)
            tv_zones = []
            df_tv_sorted = df_tv.sort_values(out_col).reset_index(drop=True)
            cluster_start = None
            cluster_rows: List = []

            def _flush_cluster(rows_: List) -> None:
                if not rows_:
                    return
                best = min(rows_, key=lambda r: group_rank.get(str(r.get("Group", "")), 99))
                tv_zones.append({
                    "output1":  round(float(best.get(out_col, 0)), 4),
                    "origin1":  str(best.get("Origin1", "")),
                    "origin2":  str(best.get("Origin2", "")),
                    "m1":       best.get("M1", ""),
                    "m2":       best.get("M2", ""),
                    "group":    str(best.get("Group", "")),
                    "match":    str(best.get("Match", "")),
                    "model_no": best.get("Model_Number", ""),
                    "recent":   str(best.get("Recent", "")),
                    "feed1":    str(best.get("Feed1", "")),
                    "feed2":    str(best.get("Feed2", "")),
                    "n_in_cluster": len(rows_),
                })

            for _, row in df_tv_sorted.iterrows():
                val = float(row.get(out_col, np.nan))
                if np.isnan(val):
                    continue
                if cluster_start is None or val - cluster_start > tv_proximity:
                    _flush_cluster(cluster_rows)
                    cluster_start = val
                    cluster_rows  = [row]
                else:
                    cluster_rows.append(row)
            _flush_cluster(cluster_rows)

    print(f"    {len(tv_zones)} traveler zones (top {top_groups} groups, deduplicated)")

    # 3. Cross MA velocities × traveler zones → alert rows
    rows = []
    for zone in tv_zones:
        target = zone["output1"]
        combo_key_prefix = f"|{zone['m1']},{zone['m2']}|{zone['group']}"

        for ma_name, ma_info in velocities.items():
            projection = _project_ma_arrival(ma_info, target, ma_proximity)
            if projection is None:
                continue

            # Catalog lookup
            combo_key    = f"{ma_name}{combo_key_prefix}"
            cat_entry    = catalog["combos"].get(combo_key, {})
            instances    = cat_entry.get("instances", 0)
            n_swing      = cat_entry.get("swing_swings", 0)
            hit_rate     = round(n_swing / instances, 3) if instances > 0 else None

            # Urgency tier
            bars = projection["bars_away"]
            urgency = ("🔴 IMMEDIATE" if bars == 0 else
                       "🟠 SOON (<30m)" if bars <= 6 else
                       "🟡 WATCH (30–120m)" if bars <= 24 else
                       "🟢 LATER")

            rows.append({
                "Asset":          asset_id,
                "As_Of":          as_of,
                "Urgency":        urgency,
                "Bars_Away":      projection["bars_away"],
                "Min_Away":       int(projection["minutes_away"]),
                "MA_Name":        ma_name,
                "MA_Current":     ma_info["value"],
                "MA_Direction":   ma_info["direction"],
                "MA_Velocity":    ma_info["velocity"],
                "MA_Projected":   projection["projected_price"],
                "TV_Zone":        target,
                "TV_Origin1":     zone["origin1"],
                "TV_Origin2":     zone["origin2"],
                "TV_M1":          zone["m1"],
                "TV_M2":          zone["m2"],
                "TV_Group":       zone["group"],
                "TV_Match":       zone["match"],
                "TV_Model":       zone["model_no"],
                "TV_Recent":      zone["recent"],
                "TV_Feeds":       f"{zone['feed1']}/{zone['feed2']}",
                "Zone_Cluster_N": zone["n_in_cluster"],
                "Catalog_Hits":   instances,
                "Catalog_Rate":   hit_rate,
                "Confidence":     projection["confidence"],
            })

    if not rows:
        print(f"    ℹ️  No convergence alerts generated.")
        return pd.DataFrame()

    df_out = pd.DataFrame(rows)

    # Sort: urgency (bars_away asc), then group rank, then catalog rate desc
    group_rank_map = {g: i for i, g in enumerate(_GROUP_RANK_ORDER)}
    df_out["_grp_rank"] = df_out["TV_Group"].map(lambda g: group_rank_map.get(g, 99))
    df_out["_cat_rate"] = df_out["Catalog_Rate"].fillna(-1)
    df_out = df_out.sort_values(
        ["Bars_Away", "_grp_rank", "_cat_rate"],
        ascending=[True, True, False]
    ).drop(columns=["_grp_rank", "_cat_rate"]).reset_index(drop=True)

    print(f"    ✅ {len(df_out)} convergence alerts  "
          f"({sum(df_out['Bars_Away'] == 0)} immediate, "
          f"{sum(df_out['Bars_Away'] <= 6)} within 30 min)")
    return df_out


def export_forecast_results(forecast_results: Dict[str, pd.DataFrame],
                              report_time: datetime,
                              output_dir: str,
                              contract_ids: Optional[Dict] = None) -> str:
    """Write Forecast_{CONTRACT}_{TS}.xlsx with one tab per asset."""
    if contract_ids is None:
        contract_ids = {}

    ts = report_time.strftime("%Y%m%d_%H%M")
    cid_list = [contract_ids.get(a, a) for a in ["NQ", "ES", "RTY"]
                if a in forecast_results and not forecast_results[a].empty]
    cid_prefix = "_".join(cid_list) + "_" if cid_list else ""
    fname = f"{cid_prefix}Forecast_{ts}.xlsx"
    fpath = os.path.join(output_dir, fname)

    has_data = any(not df.empty for df in forecast_results.values())
    if not has_data:
        print("  ℹ️  No forecast data to export.")
        return ""

    try:
        with pd.ExcelWriter(fpath, engine="openpyxl") as writer:
            for asset_id, df in forecast_results.items():
                if df.empty:
                    continue
                df.to_excel(writer, sheet_name=f"{asset_id}_FC", index=False)

        # Urgency color coding
        urgency_fills = {
            "🔴 IMMEDIATE":     PatternFill(start_color="FFCCCC", end_color="FFCCCC",
                                             fill_type="solid"),
            "🟠 SOON (<30m)":   PatternFill(start_color="FFE5CC", end_color="FFE5CC",
                                             fill_type="solid"),
            "🟡 WATCH (30–120m)": PatternFill(start_color="FFFACC", end_color="FFFACC",
                                               fill_type="solid"),
            "🟢 LATER":         PatternFill(start_color="CCFFCC", end_color="CCFFCC",
                                             fill_type="solid"),
        }

        wb = load_workbook(fpath)
        for asset_id, df in forecast_results.items():
            if df.empty:
                continue
            ws = wb[f"{asset_id}_FC"]
            color = HEADER_FILLS.get(asset_id, "4472C4")
            _fmt_ws_header(ws, color)
            _auto_col_width(ws)

            urg_col_idx = None
            for i, cell in enumerate(ws[1], start=1):
                if str(cell.value) == "Urgency":
                    urg_col_idx = i
                    break
            if urg_col_idx:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    val = row[urg_col_idx - 1].value
                    fill = urgency_fills.get(str(val) if val else "")
                    if fill:
                        for cell in row:
                            cell.fill = fill

        wb.save(fpath)
        total_alerts = sum(len(d) for d in forecast_results.values())
        print(f"  ✅ {fname}  ({total_alerts} alerts)")
        return fpath
    except Exception as e:
        print(f"  ❌ Error writing {fname}: {e}")
        return ""


# ============================================================================
# SECTION 13: COMMAND-LINE INTERFACE
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Unified Traveler Pipeline v4.0.15 — NQ + ES + RTY multi-timeframe",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Run Styles:
  quick (DEFAULT)  Requires only 5×15-min CSVs + measurement file.
                   Produces Raw and Matched traveler reports.
                   No wick analysis, no OHLC construction, no MA normalization.

  --full           Requires all 16 files (5×15-min, 5×3-min, 5×5-min + measurement).
                   Full pipeline: OHLC construction, MA normalization, wick analysis.

Pipeline Modes (apply to both quick and full):
  standard   Run traveler reports [+ wick analysis if --full] (default)
  backtest   standard + full-day swing detection, MA/TV tagging, catalog update
  forecast   standard + MA velocity projections, convergence alerts
  both       standard + backtest + forecast

Examples:
  # Quick run (default) — just drop the 5×15-min files + measurement in the folder
  python unified_traveler_pipeline_4_0_15.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-18 18:00"

  # Full run with wick analysis
  python unified_traveler_pipeline_4_0_15.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-18 18:00" \\
      --full

  # Backtest (quick mode — supply explicit 5-min paths or place them in --dir)
  python unified_traveler_pipeline_4_0_15.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-18 18:00" \\
      --mode backtest \\
      --ohlc-nq "CME_MINI_NQM2026__5.csv" \\
      --ohlc-es "CME_MINI_ESM2026__5.csv"

  # Backtest with explicit session end
  python unified_traveler_pipeline_4_0_15.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-18 18:00" --end-dt "2026-03-19 16:45" \\
      --mode backtest

  # Forecast at an intraday grab time
  python unified_traveler_pipeline_4_0_15.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-19 09:00" \\
      --mode forecast

  # Both backtest + forecast, full pipeline
  python unified_traveler_pipeline_4_0_15.py \\
      --measurement "Meas 4_8 all.xlsx" \\
      --dt "2026-03-18 18:00" \\
      --mode both --full --catalog-dir "C:/TradingData/catalog"

Required Files — QUICK mode (default, 6 total, auto-discovered in --dir):
  15-min (5):  CME_MINI_MNQ*, CME_MINI_NQ*, CME_MINI_MES*, CME_MINI_ES*, CME_MINI_RTY*  with ", 15"
  + 1 measurement Excel file (--measurement)

Required Files — FULL mode (--full, 16 total):
  15-min (5):  same as quick
   3-min (5):  same assets with ", 3"
   5-min (5):  same assets with ", 5"
  + 1 measurement Excel file (--measurement)

Per-Asset Defaults:
  NQ:  max-spread=3.0  radius=1000
  ES:  max-spread=1.0  radius=215
  RTY: max-spread=0.5  radius=100  (composite feed — no MRTY)
""",
    )

    # Run style
    parser.add_argument("--full", action="store_true",
                        help="Run full pipeline (16 files: +3-min, +5-min, wick analysis, MA normalization). "
                             "Default is quick mode (15-min + measurement only).")

    # Required
    parser.add_argument("--measurement", "-m", required=True,
                        help="Path to measurement Excel file")
    parser.add_argument("--dt", "-dt", required=True,
                        help="Report/grab datetime  (YYYY-MM-DD HH:MM)")

    # Optional - directory
    parser.add_argument("--dir", "-d", default=".",
                        help="Directory to scan for feed CSVs (default: current dir)")
    parser.add_argument("--output-dir", "-o", default=".",
                        help="Output directory for Excel files (default: current dir)")

    # Mode
    parser.add_argument("--mode", default="standard",
                        choices=["standard", "backtest", "forecast", "both"],
                        help="Pipeline mode (default: standard)")

    # Backtest / forecast options
    parser.add_argument("--end-dt", default=None,
                        help="Session end for backtest (YYYY-MM-DD HH:MM); "
                             "defaults to next 16:45 RTH close")
    parser.add_argument("--catalog-dir", default=".",
                        help="Directory for backtest_catalog.json (default: current dir)")
    parser.add_argument("--ohlc-nq",  default=None,
                        help="Explicit path to NQ 5-min OHLC CSV for BT/FC")
    parser.add_argument("--ohlc-es",  default=None,
                        help="Explicit path to ES 5-min OHLC CSV for BT/FC")
    parser.add_argument("--ohlc-rty", default=None,
                        help="Explicit path to RTY 5-min OHLC CSV for BT/FC")
    parser.add_argument("--ma-prox", type=float, default=DEFAULT_MA_PROXIMITY,
                        help=f"MA proximity threshold in pts (default: {DEFAULT_MA_PROXIMITY})")
    parser.add_argument("--tv-prox", type=float, default=DEFAULT_TV_PROXIMITY,
                        help=f"Traveler proximity threshold in pts (default: {DEFAULT_TV_PROXIMITY})")
    parser.add_argument("--swing-nq",  type=float, default=None,
                        help="NQ swing threshold in pts (default: 60.0)")
    parser.add_argument("--swing-es",  type=float, default=None,
                        help="ES swing threshold in pts (default: 20.0)")
    parser.add_argument("--swing-rty", type=float, default=None,
                        help="RTY swing threshold in pts (default: 16.0)")

    # Processing options
    parser.add_argument("--lookback", "-l", type=int, default=DEFAULT_LOOKBACK_DAYS,
                        help=f"Lookback days (default: {DEFAULT_LOOKBACK_DAYS})")
    parser.add_argument("--process-non-priority", action="store_true",
                        help="Also process non-priority origins (slow)")

    # Per-asset overrides
    parser.add_argument("--max-spread-nq",  type=float, default=None)
    parser.add_argument("--max-spread-es",  type=float, default=None)
    parser.add_argument("--max-spread-rty", type=float, default=None)
    parser.add_argument("--radius-nq",      type=int,   default=None)
    parser.add_argument("--radius-es",      type=int,   default=None)
    parser.add_argument("--radius-rty",     type=int,   default=None)

    args = parser.parse_args()

    # Build override dict
    overrides = {}
    if args.max_spread_nq  is not None: overrides.setdefault("NQ",  {})["max_spread"] = args.max_spread_nq
    if args.max_spread_es  is not None: overrides.setdefault("ES",  {})["max_spread"] = args.max_spread_es
    if args.max_spread_rty is not None: overrides.setdefault("RTY", {})["max_spread"] = args.max_spread_rty
    if args.radius_nq      is not None: overrides.setdefault("NQ",  {})["radius"]     = args.radius_nq
    if args.radius_es      is not None: overrides.setdefault("ES",  {})["radius"]     = args.radius_es
    if args.radius_rty     is not None: overrides.setdefault("RTY", {})["radius"]     = args.radius_rty

    # Build explicit OHLC path dict
    ohlc_paths = {}
    if args.ohlc_nq:  ohlc_paths["NQ"]  = args.ohlc_nq
    if args.ohlc_es:  ohlc_paths["ES"]  = args.ohlc_es
    if args.ohlc_rty: ohlc_paths["RTY"] = args.ohlc_rty

    # Build swing threshold overrides
    swing_overrides = {}
    if args.swing_nq  is not None: swing_overrides["NQ"]  = args.swing_nq
    if args.swing_es  is not None: swing_overrides["ES"]  = args.swing_es
    if args.swing_rty is not None: swing_overrides["RTY"] = args.swing_rty

    # Run
    result = run_pipeline_v4(
        search_dir=args.dir,
        measurement_path=args.measurement,
        report_datetime_str=args.dt,
        lookback_days=args.lookback,
        output_dir=args.output_dir,
        process_non_priority=args.process_non_priority,
        per_asset_overrides=overrides if overrides else None,
        mode=args.mode,
        end_datetime_str=args.end_dt,
        ohlc_paths=ohlc_paths if ohlc_paths else None,
        catalog_dir=args.catalog_dir,
        ma_proximity=args.ma_prox,
        tv_proximity=args.tv_prox,
        swing_overrides=swing_overrides if swing_overrides else None,
        quick_mode=not args.full,
    )

    return 0 if result else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
