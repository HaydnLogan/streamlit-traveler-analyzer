"""
Excel Exporter v23 enhanced (v6) - Unified Export for Both Bypass and Normal Modes

This module handles Excel export for all model results with standardized 27-column format:
- Open, Output1, Output2, Prox, Origin1, Origin2, Group, M1, M2, R1, R2, M_#s, 
  Arrival_Order, Match, Tag1, Tag2, Family1, Family2, Families, Arrival1, Arrival2, 
  Day1, Day2, Arrival_Brackets, Category, Recent, Feed1, Feed2

Features:
- Report Time in sheet headers
- Report datetime in filename
- Highlighting (yellow for [0], blue for [-1][-2][-3])
- Freeze panes at row 3
- Autofilter on row 3
- Works with both bypass mode (split columns) and normal mode (combined columns)
- Recent column shows "Today" for [0] or "Recent" for [-1] or [-3]
- Excel export timing (prints completion time)
"""

import pandas as pd
import io
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def derive_missing_columns(df):
    """
    Derive missing columns from available data (for bypass mode).
    
    Calculates:
    - Group (from Origin1, Origin2)
    - Match (from M1, M2)
    - Arrival_Order (from M1, M2)
    """
    result_df = df.copy()
    
    # Define origin classifications
    ANCHOR_ORIGINS = {"spain", "saturn", "jupiter", "kepler-62", "kepler-44"}
    TT_ORIGINS = {"trinidad", "tobago"}
    
    # Calculate Group if not present
    if 'Group' not in result_df.columns or result_df['Group'].isna().all():
        def calculate_group(row):
            if pd.isna(row.get('Origin1')) or pd.isna(row.get('Origin2')):
                return ''
            
            origin1_lower = str(row['Origin1']).lower()
            origin2_lower = str(row['Origin2']).lower()
            
            earlier_is_anchor = origin1_lower in ANCHOR_ORIGINS
            later_is_anchor = origin2_lower in ANCHOR_ORIGINS
            earlier_is_tt = origin1_lower in TT_ORIGINS
            later_is_tt = origin2_lower in TT_ORIGINS
            
            # Group 1: Both Anchors, same Anchor
            if earlier_is_anchor and later_is_anchor and origin1_lower == origin2_lower:
                return "1. SAA"
            # Group 2: Both Trinidad/Tobago
            elif earlier_is_tt and later_is_tt:
                return "2. STT"
            # Group 3: One Trinidad/Tobago, one Anchor
            elif earlier_is_tt and later_is_anchor:
                return "3. TA"
            elif earlier_is_anchor and later_is_tt:
                return "3. AT"
            # Group 4: Both Anchors, different Anchors
            elif earlier_is_anchor and later_is_anchor and origin1_lower != origin2_lower:
                return "4. AA"
            # Group 5: Later is Anchor, earlier is not
            elif later_is_anchor and not earlier_is_anchor:
                return "5. oA"
            # Group 6: Earlier is Anchor, later is not
            elif earlier_is_anchor and not later_is_anchor:
                return "6. Ao"
            else:
                return "7. oo"  # Neither is Anchor or TT
        
        result_df['Group'] = result_df.apply(calculate_group, axis=1)
    
    # Calculate Match if not present
    if 'Match' not in result_df.columns or result_df['Match'].isna().all():
        def calculate_match(row):
            if pd.isna(row.get('M1')) or pd.isna(row.get('M2')):
                return ''
            
            try:
                m1_val = int(float(row['M1']))
                m2_val = int(float(row['M2']))
                m1_abs = abs(m1_val)
                m2_abs = abs(m2_val)
                
                # Determine signage
                if (m1_val > 0 and m2_val > 0) or (m1_val < 0 and m2_val < 0):
                    signage = "same"
                else:
                    signage = "flip"
                
                # Determine direction
                if m1_abs > m2_abs:
                    direction = "up"
                elif m1_abs < m2_abs:
                    direction = "down"
                else:  # m1_abs == m2_abs
                    if signage == "flip":
                        direction = "OPP"
                    else:
                        direction = "same"
                
                return f"{signage} {direction}"
            except:
                return ''
        
        result_df['Match'] = result_df.apply(calculate_match, axis=1)
    
    # Calculate Arrival_Order if not present
    if 'Arrival_Order' not in result_df.columns or result_df['Arrival_Order'].isna().all():
        def calculate_arrival_order(row):
            if pd.isna(row.get('M1')) or pd.isna(row.get('M2')):
                return ''
            
            try:
                m1_abs = abs(int(float(row['M1'])))
                m2_abs = abs(int(float(row['M2'])))
                
                if m1_abs < 40 and m2_abs >= 40:
                    return "PD"  # Premium first (M2), then Discount (M1)
                elif m1_abs >= 40 and m2_abs < 40:
                    return "DP"  # Discount first (M1), then Premium (M2)
                elif m1_abs < 40 and m2_abs < 40:
                    return "DD"  # Both Discounts
                else:
                    return "PP"  # Both Premiums
            except:
                return ''
        
        result_df['Arrival_Order'] = result_df.apply(calculate_arrival_order, axis=1)
    
    return result_df


def standardize_dataframe(df, measurement_df=None, model_name=None):
    """
    Transform any dataframe (bypass or normal mode) into standardized 26-column format.
    
    Parameters:
    -----------
    df : DataFrame
        Results dataframe from either bypass or normal mode
    measurement_df : DataFrame, optional
        Measurement data with M# and R# columns (for extracting R values if needed)
    model_name : str, optional
        Model name to populate Category column (e.g., "Fogz PD")
    
    Returns:
    --------
    DataFrame with standardized columns
    """
    if df.empty:
        return df
    
    result_df = df.copy()
    
    # ========================================================================
    # STEP 1: Handle Output1, Output2 (rename from Input1, Input2 or split from Outputs)
    # ========================================================================
    if 'Input1' in result_df.columns and 'Input2' in result_df.columns:
        # Bypass mode - rename Input1/Input2 to Output1/Output2
        result_df = result_df.rename(columns={'Input1': 'Output1', 'Input2': 'Output2'})
    elif 'Outputs' in result_df.columns:
        # Normal mode - split "25818.50, 25818.67" into Output1, Output2
        def split_outputs(outputs_str):
            try:
                if pd.isna(outputs_str):
                    return pd.Series([None, None])
                parts = str(outputs_str).split(',')
                return pd.Series([float(parts[0].strip()), float(parts[1].strip())])
            except:
                return pd.Series([None, None])
        
        result_df[['Output1', 'Output2']] = result_df['Outputs'].apply(split_outputs)
    
    # ========================================================================
    # STEP 2: Handle Origin1, Origin2 (already split or split from Origins)
    # ========================================================================
    if 'Origin1' not in result_df.columns and 'Origins' in result_df.columns:
        # Normal mode - split "Kepler-44, Kepler-62" into Origin1, Origin2
        def split_origins(origins_str):
            try:
                if pd.isna(origins_str):
                    return pd.Series(['', ''])
                parts = str(origins_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Origin1', 'Origin2']] = result_df['Origins'].apply(split_origins)
    
    # ========================================================================
    # STEP 3: Handle M1, M2 (already split or split from M_#s)
    # ========================================================================
    if 'M1' not in result_df.columns and 'M_#s' in result_df.columns:
        # Normal mode - split "6, 43" into M1, M2
        def split_m_numbers(m_str):
            try:
                if pd.isna(m_str):
                    return pd.Series([None, None])
                parts = str(m_str).split(',')
                return pd.Series([int(float(parts[0].strip())), int(float(parts[1].strip()))])
            except:
                return pd.Series([None, None])
        
        result_df[['M1', 'M2']] = result_df['M_#s'].apply(split_m_numbers)
    
    # ========================================================================
    # STEP 4: Create M_#s combined format if not present
    # ========================================================================
    if 'M_#s' not in result_df.columns and 'M1' in result_df.columns and 'M2' in result_df.columns:
        # Bypass mode - combine M1, M2 into M_#s
        result_df['M_#s'] = result_df.apply(
            lambda row: f"{int(row['M1'])}, {int(row['M2'])}" if pd.notna(row['M1']) and pd.notna(row['M2']) else '',
            axis=1
        )
    
    # ========================================================================
    # STEP 5: Handle R1, R2 (extract from measurement_df if needed)
    # ========================================================================
    if 'R1' not in result_df.columns or 'R2' not in result_df.columns:
        if measurement_df is not None and 'M1' in result_df.columns and 'M2' in result_df.columns:
            # Try to extract R values from measurement_df
            # Find M# and R# columns
            m_col = next((c for c in measurement_df.columns if c.lower().replace(' ', '') in ['m#', 'm', 'mnumber']), None)
            r_col = next((c for c in measurement_df.columns if c.lower().replace(' ', '') in ['r#', 'r', 'recipr#', 'recipr', 'reciprocal']), None)
            
            if m_col and r_col:
                # Create lookup dict
                m_to_r = {}
                for _, row in measurement_df.iterrows():
                    try:
                        m_val = int(float(row[m_col]))
                        r_val = int(float(row[r_col]))
                        m_to_r[m_val] = r_val
                    except:
                        pass
                
                # Map M1 -> R1, M2 -> R2
                result_df['R1'] = result_df['M1'].map(m_to_r)
                result_df['R2'] = result_df['M2'].map(m_to_r)
        
        # If still not present, set to None
        if 'R1' not in result_df.columns:
            result_df['R1'] = None
        if 'R2' not in result_df.columns:
            result_df['R2'] = None
    
    # ========================================================================
    # STEP 6: Handle Tag1, Tag2 (already split or split from Tags)
    # ========================================================================
    if 'Tag1' not in result_df.columns and 'Tags' in result_df.columns:
        # Normal mode - split "X05d, X1" into Tag1, Tag2
        def split_tags(tags_str):
            try:
                if pd.isna(tags_str):
                    return pd.Series(['', ''])
                parts = str(tags_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Tag1', 'Tag2']] = result_df['Tags'].apply(split_tags)
    
    # ========================================================================
    # STEP 7: Handle Family1, Family2 (already split or split from Families)
    # ========================================================================
    if 'Family1' not in result_df.columns and 'Families' in result_df.columns:
        # Normal mode - split "Charlie, Alpha" into Family1, Family2
        def split_families(families_str):
            try:
                if pd.isna(families_str):
                    return pd.Series(['', ''])
                parts = str(families_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Family1', 'Family2']] = result_df['Families'].apply(split_families)
    
    # ========================================================================
    # STEP 8: Create Families combined format if not present
    # ========================================================================
    if 'Families' not in result_df.columns and 'Family1' in result_df.columns and 'Family2' in result_df.columns:
        # Bypass mode - combine Family1, Family2 into Families
        result_df['Families'] = result_df.apply(
            lambda row: f"{row['Family1']}, {row['Family2']}" if pd.notna(row['Family1']) and pd.notna(row['Family2']) else '',
            axis=1
        )
    
    # ========================================================================
    # STEP 9: Handle Arrival1, Arrival2 (already split or from Arrival_DateTime)
    # ========================================================================
    if 'Arrival1' not in result_df.columns:
        if 'Arrival_DateTime' in result_df.columns:
            # Normal mode - Arrival_DateTime represents the first arrival
            result_df['Arrival1'] = result_df['Arrival_DateTime']
            # For Arrival2, we don't have a separate value in normal mode
            # Leave it as None for now - may need to extract from other data
            result_df['Arrival2'] = None
        else:
            result_df['Arrival1'] = None
            result_df['Arrival2'] = None
    
    # If Arrival1 exists but Arrival2 doesn't, set Arrival2 to None
    if 'Arrival1' in result_df.columns and 'Arrival2' not in result_df.columns:
        result_df['Arrival2'] = None
    
    # ========================================================================
    # STEP 10: Handle Day1, Day2 (already split or split from Arrival_Brackets)
    # ========================================================================
    if 'Day1' not in result_df.columns and 'Arrival_Brackets' in result_df.columns:
        # Normal mode - split "[0], [-1]" into Day1, Day2
        def split_days(brackets_str):
            try:
                if pd.isna(brackets_str):
                    return pd.Series(['', ''])
                parts = str(brackets_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Day1', 'Day2']] = result_df['Arrival_Brackets'].apply(split_days)
    
    # ========================================================================
    # STEP 11: Create Arrival_Brackets combined format if not present
    # ========================================================================
    if 'Arrival_Brackets' not in result_df.columns and 'Day1' in result_df.columns and 'Day2' in result_df.columns:
        # Bypass mode - combine Day1, Day2 into Arrival_Brackets
        result_df['Arrival_Brackets'] = result_df.apply(
            lambda row: f"{row['Day1']}, {row['Day2']}" if pd.notna(row['Day1']) and pd.notna(row['Day2']) else '',
            axis=1
        )
    
    # ========================================================================
    # STEP 13: Derive missing columns from available data (for bypass mode)
    # ========================================================================
    result_df = derive_missing_columns(result_df)
    
    # ========================================================================
    # STEP 13a: Set Category from model_name if provided
    # ========================================================================
    if model_name:
        if 'Category' not in result_df.columns or result_df['Category'].isna().all():
            result_df['Category'] = model_name
    
    # ========================================================================
    # STEP 13b: Derive Recent column from Day1 if not present
    # ========================================================================
    if 'Recent' not in result_df.columns or result_df['Recent'].isna().all():
        def derive_recent(row):
            """Derive Recent value from Day1"""
            if pd.isna(row.get('Day1')):
                return ''
            day1 = str(row['Day1'])
            if '[0]' in day1:
                return 'Today'
            elif '[-1]' in day1 or '[-3]' in day1:
                return 'Recent'
            else:
                return ''  # Other day values
        
        result_df['Recent'] = result_df.apply(derive_recent, axis=1)
    
    # ========================================================================
    # STEP 14: Ensure all required columns exist with defaults
    # ========================================================================
    required_columns = {
        'Open': '',
        'Output1': None,
        'Output2': None,
        'Prox': None,
        'Origin1': '',
        'Origin2': '',
        'Group': '',
        'M1': None,
        'M2': None,
        'R1': None,
        'R2': None,
        'M_#s': '',
        'Arrival_Order': '',
        'Match': '',
        'Tag1': '',
        'Tag2': '',
        'Family1': '',
        'Family2': '',
        'Families': '',
        'Arrival1': None,
        'Arrival2': None,
        'Day1': '',
        'Day2': '',
        'Arrival_Brackets': '',
        'Category': '',
        'Feed1': '',
        'Feed2': '',
        'Recent': ''  # Today or Recent based on Day1
    }
    
    for col, default in required_columns.items():
        if col not in result_df.columns:
            result_df[col] = default
    
    # ========================================================================
    # STEP 15: Select and order columns (27 standardized columns)
    # ========================================================================
    final_columns = [
        'Open', 'Output1', 'Output2', 'Prox', 'Origin1', 'Origin2', 'Group',
        'M1', 'M2', 'R1', 'R2', 'M_#s', 'Arrival_Order', 'Match',
        'Tag1', 'Tag2', 'Family1', 'Family2', 'Families',
        'Arrival1', 'Arrival2', 'Day1', 'Day2', 'Arrival_Brackets',
        'Category', 'Recent', 'Feed1', 'Feed2'
    ]
    
    # Ensure columns exist before selecting
    available_columns = [col for col in final_columns if col in result_df.columns]
    result_df = result_df[available_columns]
    
    return result_df


def export_all_models_to_excel(all_results, report_time, measurement_df=None, 
                               prox_threshold_yellow=0.25, prox_threshold_blue=1.0,
                               highlight_origins=False, highlight_group=False):
    """
    Export all model results to Excel with standardized 27-column format.
    
    Includes 'Recent' column showing "Today" for [0] arrivals, "Recent" for [-1] or [-3].
    
    Parameters:
    -----------
    all_results : dict
        Dictionary from model_processor with 'results', 'timings', etc.
    report_time : datetime
        Report datetime to include in headers and filename
    measurement_df : DataFrame, optional
        Measurement data for extracting R# values if needed
    prox_threshold_yellow : float, default 0.25
        Prox values below this threshold get bright yellow highlighting
    prox_threshold_blue : float, default 1.0
        Prox values below this threshold get light blue highlighting
    highlight_origins : bool, default False
        If True, highlight anchor origins (Spain, Saturn, Jupiter, Kepler-62, Kepler-44)
    highlight_group : bool, default False
        If True, highlight SAA groups differently
    
    Returns:
    --------
    tuple : (output_bytes, filename)
    """
    # Start timing Excel export
    export_start_time = time.time()
    
    output = io.BytesIO()
    results = all_results['results']
    timings = all_results.get('timings', {})
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write each model to its own sheet
        for model_name, df in results.items():
            if len(df) > 0:
                # Standardize dataframe to 26-column format with model_name for Category
                standardized_df = standardize_dataframe(df, measurement_df, model_name)
                
                # Create sheet name (Excel limit: 31 chars)
                sheet_name = model_name[:31]
                
                # Write dataframe
                standardized_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                
                # Get worksheet to add header and formatting
                worksheet = writer.sheets[sheet_name]
                
                # Add Report Time header (row 1)
                report_str = report_time.strftime('%Y-%m-%d %H:%M')
                worksheet['A1'] = f"Report Time: {report_str}"
                worksheet['A1'].font = worksheet['A1'].font.copy(bold=True)
                
                # Add timing info if available (row 2)
                if model_name in timings:
                    timing = timings[model_name]
                    worksheet['A2'] = f"Generated in {timing:.2f}s | {len(standardized_df)} matches"
                    worksheet['A2'].font = worksheet['A2'].font.copy(italic=True)
        
        # Create Combined sheet with all results
        all_tables = []
        for model_name, df in results.items():
            if len(df) > 0:
                standardized_df = standardize_dataframe(df, measurement_df, model_name)
                all_tables.append(standardized_df)
        
        if all_tables:
            combined_df = pd.concat(all_tables, ignore_index=True)
            
            # Sort by Arrival_Output if present, otherwise by Output1
            if 'Arrival_Output' in combined_df.columns:
                combined_df = combined_df.sort_values('Arrival_Output', ascending=False)
            elif 'Output1' in combined_df.columns:
                combined_df = combined_df.sort_values('Output1', ascending=False)
            
            # Write combined sheet
            combined_df.to_excel(writer, sheet_name='Combined', index=False, startrow=2)
            
            # Add headers to combined sheet
            combined_ws = writer.sheets['Combined']
            combined_ws['A1'] = f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')}"
            combined_ws['A1'].font = combined_ws['A1'].font.copy(bold=True)
            combined_ws['A2'] = f"All Models Combined | {len(combined_df)} total matches"
            combined_ws['A2'].font = combined_ws['A2'].font.copy(italic=True)
    
    # Reopen workbook to apply highlighting, freeze panes, and filters
    output.seek(0)
    workbook = load_workbook(output)
    
    # Define fill colors
    yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # For [0] brackets
    blue_fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')  # For [-1][-2][-3] brackets
    pink_fill = PatternFill(start_color='FF99CC', end_color='FF99CC', fill_type='solid')  # Pink for Prox < 0.25
    bright_yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Bright yellow for Prox < 1.0
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Light green for anchor origins
    orange_fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')  # Light orange for SAA groups
    
    # Define font colors for M1, M2, R1, R2
    from openpyxl.styles import Font
    red_font = Font(color='FF0000')  # Red for positive numbers
    blue_font = Font(color='0070C0')  # Blue for negative numbers
    
    # Apply highlighting, freeze panes, and filters to each sheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Freeze row 3 (header row with column names)
        # Row 1 = Report Time, Row 2 = Timing info, Row 3 = Column headers
        worksheet.freeze_panes = 'A4'  # Freeze everything above row 4
        
        # Set column widths (3.3677 = 40 pixels). Actual results: 2.56, 30 pixels. 4.44 gets the desired result with a 7 pixel offset.
        # Find column indices for specific columns
        header_row = 3
        col_widths = {
            'Prox': 5.56,
            'M1': 4.44,
            'M2': 4.44,
            'R1': 4.44,
            'R2': 4.44,
            'Arrival_Order': 3.67,
            'Day1': 4.44,
            'Day2': 4.44,
            'Feed1': 5.44,
            'Feed2': 5.44
        }
        
        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            col_name = cell.value
            if col_name in col_widths:
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = col_widths[col_name]
        
        # Add autofilter to row 3 (column headers)
        max_col = worksheet.max_column
        if max_col > 0:
            last_col_letter = get_column_letter(max_col)
            worksheet.auto_filter.ref = f'A3:{last_col_letter}3'
        
        # Find columns for highlighting (data starts at row 4)
        
        arrival_brackets_col = None
        prox_col = None
        origin1_col = None
        origin2_col = None
        group_col = None
        m1_col = None
        m2_col = None
        r1_col = None
        r2_col = None
        
        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            col_name = cell.value
            if col_name == 'Arrival_Brackets':
                arrival_brackets_col = col_idx
            elif col_name == 'Prox':
                prox_col = col_idx
            elif col_name == 'Origin1':
                origin1_col = col_idx
            elif col_name == 'Origin2':
                origin2_col = col_idx
            elif col_name == 'Group':
                group_col = col_idx
            elif col_name == 'M1':
                m1_col = col_idx
            elif col_name == 'M2':
                m2_col = col_idx
            elif col_name == 'R1':
                r1_col = col_idx
            elif col_name == 'R2':
                r2_col = col_idx
        
        # Define anchor origins for highlighting
        ANCHOR_ORIGINS = {'spain', 'saturn', 'jupiter', 'kepler-62', 'kepler-44'}
        
        # Apply highlighting to relevant columns
        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row):
            # Highlight Arrival_Brackets
            if arrival_brackets_col and len(row) >= arrival_brackets_col:
                cell = row[arrival_brackets_col - 1]
                value = str(cell.value) if cell.value else ''
                
                # Check for [0] - yellow
                if '[0]' in value:
                    cell.fill = yellow_fill
                # Check for [-1], [-2], [-3] - blue
                elif any(x in value for x in ['[-1]', '[-2]', '[-3]']):
                    cell.fill = blue_fill
            
            # Highlight Prox with configurable thresholds and format to 2 decimals
            if prox_col and len(row) >= prox_col:
                cell = row[prox_col - 1]
                try:
                    prox_value = float(cell.value) if cell.value is not None else None
                    if prox_value is not None:
                        # Format to 2 decimal places
                        cell.number_format = '0.00'
                        
                        # Prox < prox_threshold_yellow → pink
                        if prox_value < prox_threshold_yellow:
                            cell.fill = pink_fill
                        # Prox < prox_threshold_blue → bright yellow
                        elif prox_value < prox_threshold_blue:
                            cell.fill = bright_yellow_fill
                except (ValueError, TypeError):
                    pass  # Skip non-numeric values
            
            # Color M1, M2, R1, R2 fonts (red for positive, blue for negative)
            for col_idx, col_name in [(m1_col, 'M1'), (m2_col, 'M2'), (r1_col, 'R1'), (r2_col, 'R2')]:
                if col_idx and len(row) >= col_idx:
                    cell = row[col_idx - 1]
                    try:
                        value = float(cell.value) if cell.value is not None else None
                        if value is not None:
                            if value > 0:
                                cell.font = red_font
                            elif value < 0:
                                cell.font = blue_font
                    except (ValueError, TypeError):
                        pass  # Skip non-numeric values
            
            # Highlight Origins if enabled
            if highlight_origins:
                if origin1_col and len(row) >= origin1_col:
                    cell = row[origin1_col - 1]
                    value = str(cell.value).lower() if cell.value else ''
                    if value in ANCHOR_ORIGINS:
                        cell.fill = green_fill
                
                if origin2_col and len(row) >= origin2_col:
                    cell = row[origin2_col - 1]
                    value = str(cell.value).lower() if cell.value else ''
                    if value in ANCHOR_ORIGINS:
                        cell.fill = green_fill
            
            # Highlight Group if enabled
            if highlight_group and group_col and len(row) >= group_col:
                cell = row[group_col - 1]
                value = str(cell.value) if cell.value else ''
                if 'SAA' in value:  # Highlight SAA groups
                    cell.fill = orange_fill
    
    # Save modified workbook
    output_final = io.BytesIO()
    workbook.save(output_final)
    output_final.seek(0)
    
    # Calculate export time
    export_time = time.time() - export_start_time
    print(f"✓ Excel export completed in {export_time:.2f}s")
    
    # Create filename with report datetime
    report_dt_str = report_time.strftime('%Y%m%d_%H%M')
    filename = f"swing_analysis_{report_dt_str}_v23.xlsx"
    
    return output_final.getvalue(), filename


def create_download_button(all_results, report_time, measurement_df=None,
                          prox_threshold_yellow=0.25, prox_threshold_blue=1.0,
                          highlight_origins=False, highlight_group=False):
    """
    Create Streamlit download button for Excel export.
    
    Parameters:
    -----------
    all_results : dict
        Dictionary from model_processor
    report_time : datetime
        Report datetime
    measurement_df : DataFrame, optional
        Measurement data for R# extraction
    prox_threshold_yellow : float, default 0.25
        Prox values below this get bright yellow highlighting
    prox_threshold_blue : float, default 1.0
        Prox values below this get light blue highlighting
    highlight_origins : bool, default False
        If True, highlight anchor origins
    highlight_group : bool, default False
        If True, highlight SAA groups
    
    Returns:
    --------
    None (creates Streamlit widget)
    """
    import streamlit as st
    
    try:
        excel_bytes, filename = export_all_models_to_excel(
            all_results, report_time, measurement_df,
            prox_threshold_yellow, prox_threshold_blue,
            highlight_origins, highlight_group
        )
        
        # Count how many models have results
        num_models = sum(1 for df in all_results['results'].values() if len(df) > 0)
        total_matches = sum(len(df) for df in all_results['results'].values())
        
        st.download_button(
            label=f"📥 Download All Model Results (Excel)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_v23"
        )
        
        
        # Build highlighting description
        highlight_desc = []
        highlight_desc.append("🟡 Yellow = [0] (Today)")
        highlight_desc.append("🔵 Blue = [-1][-2][-3] (Recent)")
        highlight_desc.append(f"💛 Prox < {prox_threshold_yellow}")
        highlight_desc.append(f"🔷 Prox < {prox_threshold_blue}")
        if highlight_origins:
            highlight_desc.append("🟢 Anchor Origins")
        if highlight_group:
            highlight_desc.append("🟠 SAA Groups")
        
        st.caption(
            f"Excel file contains {num_models} model sheets + Combined sheet "
            f"({total_matches} total matches) | "
            f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')} | "
            f"{' | '.join(highlight_desc)} | "
            f"Format: 26 standardized columns"
        )
        
    except Exception as e:
        st.error(f"Error creating Excel export: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
