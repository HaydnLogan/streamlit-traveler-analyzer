"""
Excel Exporter v23 - Unified Export for Both Bypass and Normal Modes

This module handles Excel export for all model results with standardized 26-column format:
- Open, Output1, Output2, Prox, Origin1, Origin2, Group, M1, M2, R1, R2, M_#s, 
  Arrival_Order, Match, Tag1, Tag2, Family1, Family2, Families, Arrival1, Arrival2, 
  Day1, Day2, Arrival_Brackets, Category, Feed1, Feed2

Features:
- Report Time in sheet headers
- Report datetime in filename
- Highlighting (yellow for [0], blue for [-1][-2][-3])
- Freeze panes at row 3
- Autofilter on row 3
- Works with both bypass mode (split columns) and normal mode (combined columns)
"""

import pandas as pd
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def standardize_dataframe(df, measurement_df=None):
    """
    Transform any dataframe (bypass or normal mode) into standardized 26-column format.
    
    Parameters:
    -----------
    df : DataFrame
        Results dataframe from either bypass or normal mode
    measurement_df : DataFrame, optional
        Measurement data with M# and R# columns (for extracting R values if needed)
    
    Returns:
    --------
    DataFrame with standardized columns
    """
    if df.empty:
        return df
    
    result_df = df.copy()
    
    # ========================================================================
    # STEP 1: Handle Output1, Output2 (rename from Input1, Input2 or split from Outputs)
    # ========================================================================
    if 'Input1' in result_df.columns and 'Input2' in result_df.columns:
        # Bypass mode - rename Input1/Input2 to Output1/Output2
        result_df = result_df.rename(columns={'Input1': 'Output1', 'Input2': 'Output2'})
    elif 'Outputs' in result_df.columns:
        # Normal mode - split "25818.50, 25818.67" into Output1, Output2
        def split_outputs(outputs_str):
            try:
                if pd.isna(outputs_str):
                    return pd.Series([None, None])
                parts = str(outputs_str).split(',')
                return pd.Series([float(parts[0].strip()), float(parts[1].strip())])
            except:
                return pd.Series([None, None])
        
        result_df[['Output1', 'Output2']] = result_df['Outputs'].apply(split_outputs)
    
    # ========================================================================
    # STEP 2: Handle Origin1, Origin2 (already split or split from Origins)
    # ========================================================================
    if 'Origin1' not in result_df.columns and 'Origins' in result_df.columns:
        # Normal mode - split "Kepler-44, Kepler-62" into Origin1, Origin2
        def split_origins(origins_str):
            try:
                if pd.isna(origins_str):
                    return pd.Series(['', ''])
                parts = str(origins_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Origin1', 'Origin2']] = result_df['Origins'].apply(split_origins)
    
    # ========================================================================
    # STEP 3: Handle M1, M2 (already split or split from M_#s)
    # ========================================================================
    if 'M1' not in result_df.columns and 'M_#s' in result_df.columns:
        # Normal mode - split "6, 43" into M1, M2
        def split_m_numbers(m_str):
            try:
                if pd.isna(m_str):
                    return pd.Series([None, None])
                parts = str(m_str).split(',')
                return pd.Series([int(float(parts[0].strip())), int(float(parts[1].strip()))])
            except:
                return pd.Series([None, None])
        
        result_df[['M1', 'M2']] = result_df['M_#s'].apply(split_m_numbers)
    
    # ========================================================================
    # STEP 4: Create M_#s combined format if not present
    # ========================================================================
    if 'M_#s' not in result_df.columns and 'M1' in result_df.columns and 'M2' in result_df.columns:
        # Bypass mode - combine M1, M2 into M_#s
        result_df['M_#s'] = result_df.apply(
            lambda row: f"{int(row['M1'])}, {int(row['M2'])}" if pd.notna(row['M1']) and pd.notna(row['M2']) else '',
            axis=1
        )
    
    # ========================================================================
    # STEP 5: Handle R1, R2 (extract from measurement_df if needed)
    # ========================================================================
    if 'R1' not in result_df.columns or 'R2' not in result_df.columns:
        if measurement_df is not None and 'M1' in result_df.columns and 'M2' in result_df.columns:
            # Try to extract R values from measurement_df
            # Find M# and R# columns
            m_col = next((c for c in measurement_df.columns if c.lower().replace(' ', '') in ['m#', 'm', 'mnumber']), None)
            r_col = next((c for c in measurement_df.columns if c.lower().replace(' ', '') in ['r#', 'r', 'recipr#', 'recipr', 'reciprocal']), None)
            
            if m_col and r_col:
                # Create lookup dict
                m_to_r = {}
                for _, row in measurement_df.iterrows():
                    try:
                        m_val = int(float(row[m_col]))
                        r_val = int(float(row[r_col]))
                        m_to_r[m_val] = r_val
                    except:
                        pass
                
                # Map M1 -> R1, M2 -> R2
                result_df['R1'] = result_df['M1'].map(m_to_r)
                result_df['R2'] = result_df['M2'].map(m_to_r)
        
        # If still not present, set to None
        if 'R1' not in result_df.columns:
            result_df['R1'] = None
        if 'R2' not in result_df.columns:
            result_df['R2'] = None
    
    # ========================================================================
    # STEP 6: Handle Tag1, Tag2 (already split or split from Tags)
    # ========================================================================
    if 'Tag1' not in result_df.columns and 'Tags' in result_df.columns:
        # Normal mode - split "X05d, X1" into Tag1, Tag2
        def split_tags(tags_str):
            try:
                if pd.isna(tags_str):
                    return pd.Series(['', ''])
                parts = str(tags_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Tag1', 'Tag2']] = result_df['Tags'].apply(split_tags)
    
    # ========================================================================
    # STEP 7: Handle Family1, Family2 (already split or split from Families)
    # ========================================================================
    if 'Family1' not in result_df.columns and 'Families' in result_df.columns:
        # Normal mode - split "Charlie, Alpha" into Family1, Family2
        def split_families(families_str):
            try:
                if pd.isna(families_str):
                    return pd.Series(['', ''])
                parts = str(families_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Family1', 'Family2']] = result_df['Families'].apply(split_families)
    
    # ========================================================================
    # STEP 8: Create Families combined format if not present
    # ========================================================================
    if 'Families' not in result_df.columns and 'Family1' in result_df.columns and 'Family2' in result_df.columns:
        # Bypass mode - combine Family1, Family2 into Families
        result_df['Families'] = result_df.apply(
            lambda row: f"{row['Family1']}, {row['Family2']}" if pd.notna(row['Family1']) and pd.notna(row['Family2']) else '',
            axis=1
        )
    
    # ========================================================================
    # STEP 9: Handle Arrival1, Arrival2 (already split or split from Arrival_DateTime)
    # ========================================================================
    if 'Arrival1' not in result_df.columns and 'Arrival_DateTime' in result_df.columns:
        # Normal mode - use Arrival_DateTime for Arrival1, need to extract Arrival2 somehow
        # For now, duplicate Arrival_DateTime (may need refinement based on actual data structure)
        result_df['Arrival1'] = result_df['Arrival_DateTime']
        result_df['Arrival2'] = result_df['Arrival_DateTime']  # Placeholder
    
    # ========================================================================
    # STEP 10: Handle Day1, Day2 (already split or split from Arrival_Brackets)
    # ========================================================================
    if 'Day1' not in result_df.columns and 'Arrival_Brackets' in result_df.columns:
        # Normal mode - split "[0], [-1]" into Day1, Day2
        def split_days(brackets_str):
            try:
                if pd.isna(brackets_str):
                    return pd.Series(['', ''])
                parts = str(brackets_str).split(',')
                return pd.Series([parts[0].strip(), parts[1].strip()])
            except:
                return pd.Series(['', ''])
        
        result_df[['Day1', 'Day2']] = result_df['Arrival_Brackets'].apply(split_days)
    
    # ========================================================================
    # STEP 11: Create Arrival_Brackets combined format if not present
    # ========================================================================
    if 'Arrival_Brackets' not in result_df.columns and 'Day1' in result_df.columns and 'Day2' in result_df.columns:
        # Bypass mode - combine Day1, Day2 into Arrival_Brackets
        result_df['Arrival_Brackets'] = result_df.apply(
            lambda row: f"{row['Day1']}, {row['Day2']}" if pd.notna(row['Day1']) and pd.notna(row['Day2']) else '',
            axis=1
        )
    
    # ========================================================================
    # STEP 12: Ensure all required columns exist with defaults
    # ========================================================================
    required_columns = {
        'Open': '',
        'Output1': None,
        'Output2': None,
        'Prox': None,
        'Origin1': '',
        'Origin2': '',
        'Group': '',
        'M1': None,
        'M2': None,
        'R1': None,
        'R2': None,
        'M_#s': '',
        'Arrival_Order': '',
        'Match': '',
        'Tag1': '',
        'Tag2': '',
        'Family1': '',
        'Family2': '',
        'Families': '',
        'Arrival1': None,
        'Arrival2': None,
        'Day1': '',
        'Day2': '',
        'Arrival_Brackets': '',
        'Category': '',
        'Feed1': '',
        'Feed2': ''
    }
    
    for col, default in required_columns.items():
        if col not in result_df.columns:
            result_df[col] = default
    
    # ========================================================================
    # STEP 13: Select and order columns
    # ========================================================================
    final_columns = [
        'Open', 'Output1', 'Output2', 'Prox', 'Origin1', 'Origin2', 'Group',
        'M1', 'M2', 'R1', 'R2', 'M_#s', 'Arrival_Order', 'Match',
        'Tag1', 'Tag2', 'Family1', 'Family2', 'Families',
        'Arrival1', 'Arrival2', 'Day1', 'Day2', 'Arrival_Brackets',
        'Category', 'Feed1', 'Feed2'
    ]
    
    # Ensure columns exist before selecting
    available_columns = [col for col in final_columns if col in result_df.columns]
    result_df = result_df[available_columns]
    
    return result_df


def export_all_models_to_excel(all_results, report_time, measurement_df=None):
    """
    Export all model results to Excel with standardized 26-column format.
    
    Parameters:
    -----------
    all_results : dict
        Dictionary from model_processor with 'results', 'timings', etc.
    report_time : datetime
        Report datetime to include in headers and filename
    measurement_df : DataFrame, optional
        Measurement data for extracting R# values if needed
    
    Returns:
    --------
    tuple : (output_bytes, filename)
    """
    output = io.BytesIO()
    results = all_results['results']
    timings = all_results.get('timings', {})
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write each model to its own sheet
        for model_name, df in results.items():
            if len(df) > 0:
                # Standardize dataframe to 26-column format
                standardized_df = standardize_dataframe(df, measurement_df)
                
                # Create sheet name (Excel limit: 31 chars)
                sheet_name = model_name[:31]
                
                # Write dataframe
                standardized_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
                
                # Get worksheet to add header and formatting
                worksheet = writer.sheets[sheet_name]
                
                # Add Report Time header (row 1)
                report_str = report_time.strftime('%Y-%m-%d %H:%M')
                worksheet['A1'] = f"Report Time: {report_str}"
                worksheet['A1'].font = worksheet['A1'].font.copy(bold=True)
                
                # Add timing info if available (row 2)
                if model_name in timings:
                    timing = timings[model_name]
                    worksheet['A2'] = f"Generated in {timing:.2f}s | {len(standardized_df)} matches"
                    worksheet['A2'].font = worksheet['A2'].font.copy(italic=True)
        
        # Create Combined sheet with all results
        all_tables = []
        for df in results.values():
            if len(df) > 0:
                standardized_df = standardize_dataframe(df, measurement_df)
                all_tables.append(standardized_df)
        
        if all_tables:
            combined_df = pd.concat(all_tables, ignore_index=True)
            
            # Sort by Arrival_Output if present, otherwise by Output1
            if 'Arrival_Output' in combined_df.columns:
                combined_df = combined_df.sort_values('Arrival_Output', ascending=False)
            elif 'Output1' in combined_df.columns:
                combined_df = combined_df.sort_values('Output1', ascending=False)
            
            # Write combined sheet
            combined_df.to_excel(writer, sheet_name='Combined', index=False, startrow=2)
            
            # Add headers to combined sheet
            combined_ws = writer.sheets['Combined']
            combined_ws['A1'] = f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')}"
            combined_ws['A1'].font = combined_ws['A1'].font.copy(bold=True)
            combined_ws['A2'] = f"All Models Combined | {len(combined_df)} total matches"
            combined_ws['A2'].font = combined_ws['A2'].font.copy(italic=True)
    
    # Reopen workbook to apply highlighting, freeze panes, and filters
    output.seek(0)
    workbook = load_workbook(output)
    
    # Define fill colors
    yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    blue_fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
    
    # Apply highlighting, freeze panes, and filters to each sheet
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Freeze row 3 (header row with column names)
        # Row 1 = Report Time, Row 2 = Timing info, Row 3 = Column headers
        worksheet.freeze_panes = 'A4'  # Freeze everything above row 4
        
        # Add autofilter to row 3 (column headers)
        max_col = worksheet.max_column
        if max_col > 0:
            last_col_letter = get_column_letter(max_col)
            worksheet.auto_filter.ref = f'A3:{last_col_letter}3'
        
        # Find Arrival_Brackets column for highlighting (data starts at row 4)
        header_row = 3  # Row 3 contains column headers
        
        arrival_brackets_col = None
        for col_idx, cell in enumerate(worksheet[header_row], start=1):
            if cell.value == 'Arrival_Brackets':
                arrival_brackets_col = col_idx
                break
        
        if arrival_brackets_col:
            # Apply highlighting based on Arrival_Brackets values
            for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row):
                if len(row) >= arrival_brackets_col:
                    cell = row[arrival_brackets_col - 1]
                    value = str(cell.value) if cell.value else ''
                    
                    # Check for [0] - yellow
                    if '[0]' in value:
                        cell.fill = yellow_fill
                    # Check for [-1], [-2], [-3] - blue
                    elif any(x in value for x in ['[-1]', '[-2]', '[-3]']):
                        cell.fill = blue_fill
    
    # Save modified workbook
    output_final = io.BytesIO()
    workbook.save(output_final)
    output_final.seek(0)
    
    # Create filename with report datetime
    report_dt_str = report_time.strftime('%Y%m%d_%H%M')
    filename = f"swing_analysis_{report_dt_str}_v23.xlsx"
    
    return output_final.getvalue(), filename


def create_download_button(all_results, report_time, measurement_df=None):
    """
    Create Streamlit download button for Excel export.
    
    Parameters:
    -----------
    all_results : dict
        Dictionary from model_processor
    report_time : datetime
        Report datetime
    measurement_df : DataFrame, optional
        Measurement data for R# extraction
    
    Returns:
    --------
    None (creates Streamlit widget)
    """
    import streamlit as st
    
    try:
        excel_bytes, filename = export_all_models_to_excel(all_results, report_time, measurement_df)
        
        # Count how many models have results
        num_models = sum(1 for df in all_results['results'].values() if len(df) > 0)
        total_matches = sum(len(df) for df in all_results['results'].values())
        
        st.download_button(
            label=f"📥 Download All Model Results (Excel)",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_v23"
        )
        
        st.caption(
            f"Excel file contains {num_models} model sheets + Combined sheet "
            f"({total_matches} total matches) | "
            f"Report Time: {report_time.strftime('%Y-%m-%d %H:%M')} | "
            f"🟡 Yellow = [0] (Today) | 🔵 Blue = [-1][-2][-3] (Recent) | "
            f"Format: 26 standardized columns"
        )
        
    except Exception as e:
        st.error(f"Error creating Excel export: {str(e)}")
        import traceback
        st.code(traceback.format_exc())
