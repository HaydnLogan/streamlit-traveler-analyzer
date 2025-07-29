# Simple Single Line Mega Report - Uses direct imports from existing files
# More reliable approach using the existing model files

import streamlit as st
import pandas as pd
import io
from datetime import datetime
from collections import defaultdict

# Direct imports from existing files
try:
    from models_a_today import detect_A_models_today_only, classify_A_model, find_flexible_descents
    A_MODELS_AVAILABLE = True
except ImportError:
    A_MODELS_AVAILABLE = False

try:
    from fixed_model_c_complete import detect_C_models, classify_c01_sequence, classify_c02_sequence, classify_c04_sequence, find_descending_sequences
    C_MODELS_AVAILABLE = True
except ImportError:
    C_MODELS_AVAILABLE = False

def calculate_comprehensive_metrics(output_rows, report_time):
    """
    Calculate all metrics from the Excel template
    """
    rows = output_rows.sort_values("Arrival").reset_index(drop=True)
    
    # Basic metrics
    total_m_count = len(rows)
    m_numbers = rows["M #"].tolist()
    m_ids = ", ".join([str(m) for m in m_numbers])
    m_start_end = f"{m_numbers[0]} → {m_numbers[-1]}" if len(m_numbers) > 1 else str(m_numbers[0])
    
    # Arrival order
    arrival_order = " → ".join([str(m) for m in m_numbers])
    
    # Strength travelers (M# values of 0, 40, -40, 54, -54)
    strength_values = {0, 40, -40, 54, -54}
    strength_travelers = [m for m in m_numbers if m in strength_values]
    strength_count = len(strength_travelers)
    strength_ids = ", ".join([str(m) for m in strength_travelers])
    
    # Tag B travelers (assuming non-strength travelers)
    tag_b_travelers = [m for m in m_numbers if m not in strength_values]
    tag_b_count = len(tag_b_travelers)
    tag_b_ids = ", ".join([str(m) for m in tag_b_travelers])
    
    # Sequence analysis
    origins = rows["Origin"].tolist()
    families = rows.get("Family", [""] * len(rows)).tolist()
    feeds = rows["Feed"].tolist()
    
    # Sequence characteristics
    same_origin = len(set(origins)) == 1
    same_family = len(set(families)) == 1 if all(f for f in families) else False
    same_feed = len(set(feeds)) == 1
    
    # Time calculations
    latest_arrival = rows["Arrival"].max()
    hours_ago = int((report_time - latest_arrival).total_seconds() / 3600)
    
    # Input values (if available in data)
    input_18 = rows.get("Input @ 18:00", [""]).iloc[-1] if "Input @ 18:00" in rows.columns else ""
    input_arrival = rows.get("Input @ Arrival", [""]).iloc[-1] if "Input @ Arrival" in rows.columns else ""
    input_report = rows.get("Input @ Report", [""]).iloc[-1] if "Input @ Report" in rows.columns else ""
    
    # Differences (if input columns exist)
    diff_18 = ""
    diff_arrival = ""
    diff_report = ""
    
    return {
        "total_m_count": total_m_count,
        "m_ids": m_ids,
        "m_start_end": m_start_end,
        "arrival_order": arrival_order,
        "strength_count": strength_count,
        "strength_ids": strength_ids,
        "tag_b_count": tag_b_count,
        "tag_b_ids": tag_b_ids,
        "same_origin": same_origin,
        "same_family": same_family,
        "same_feed": same_feed,
        "hours_ago": hours_ago,
        "latest_arrival": latest_arrival,
        "input_18": input_18,
        "input_arrival": input_arrival,
        "input_report": input_report,
        "diff_18": diff_18,
        "diff_arrival": diff_arrival,
        "diff_report": diff_report,
        "origins": origins,
        "families": families,
        "feeds": feeds
    }

def analyze_output_patterns_comprehensive(output_rows, report_time):
    """
    Comprehensive pattern analysis with all Excel template metrics
    """
    patterns_found = []
    sequences_found = []
    
    # Sort rows by arrival time
    rows = output_rows.sort_values("Arrival").reset_index(drop=True)
    
    # Check A model patterns if available
    if A_MODELS_AVAILABLE:
        try:
            sequences = find_flexible_descents(rows)
            for seq in sequences:
                if seq.shape[0] >= 2:
                    last = seq.iloc[-1]
                    prior = seq.iloc[:-1] if seq.shape[0] > 1 else pd.DataFrame()
                    
                    # Only process if ends on 'today' ([0])
                    final_day = str(last["Day"]).strip()
                    if final_day == "[0]":
                        model, label = classify_A_model(last, prior)
                        if model:
                            m_path = " → ".join([f"|{row['M #']}|" for _, row in seq.iterrows()])
                            patterns_found.append(f"A:{model}({m_path})")
                            sequences_found.append({
                                "type": "A",
                                "model": model,
                                "sequence": seq,
                                "length": len(seq)
                            })
        except Exception as e:
            st.warning(f"Error in A model detection: {e}")
    
    # Check C model patterns if available
    if C_MODELS_AVAILABLE:
        try:
            # Create mini dataframe for this output
            mini_df = rows.copy()
            mini_df['Output'] = output_rows['Output'].iloc[0]
            
            # Use existing C model detection
            c_results = detect_C_models(mini_df, run_c01=True, run_c02=True, run_c04=True)
            
            for tag, items in c_results.items():
                for item in items:
                    seq = item['sequence']
                    m_path = " → ".join([f"|{row['M #']}|" for _, row in seq.iterrows()])
                    patterns_found.append(f"C:{tag}({m_path})")
                    sequences_found.append({
                        "type": "C",
                        "model": tag,
                        "sequence": seq,
                        "length": len(seq)
                    })
                    
        except Exception as e:
            st.warning(f"Error in C model detection: {e}")
    
    # Calculate additional sequence metrics
    unique_sequences = len(sequences_found)
    largest_seq_length = max([s["length"] for s in sequences_found]) if sequences_found else 0
    
    # Get largest sequence start/end
    largest_seq = None
    if sequences_found:
        largest_seq = max(sequences_found, key=lambda x: x["length"])
        seq_start_end = f"{largest_seq['sequence'].iloc[0]['M #']} → {largest_seq['sequence'].iloc[-1]['M #']}"
    else:
        seq_start_end = ""
    
    return patterns_found, {
        "unique_sequences": unique_sequences,
        "largest_seq_length": largest_seq_length,
        "seq_start_end": seq_start_end,
        "sequences_found": sequences_found
    }

def calculate_star_rating(patterns, largest_seq_length, models_found, strength_ids, top_2_largest, top_5_models, has_top_strength_criteria):
    """
    Calculate star rating based on complex rules
    """
    # 2-star models
    two_star_models = {"A02", "B01a[0]", "C.01.o.[aft0]", "C.02.p1.[L0]", "C.02.p2.[E0]", "C.02.p3.[O0]", "C.04.∀1.[0]"}
    
    # 1-star models  
    one_star_models = {"C.04.∀2.[±1]"}
    
    # Check what models we have
    found_models = set()
    for pattern in patterns:
        # Extract model name from pattern like "A:A01(...)" or "C:C.02.p1.[L0](...)"
        if ":" in pattern and "(" in pattern:
            model_part = pattern.split("(")[0].split(":")[1]
            found_models.add(model_part)
    
    # Count stars from models
    has_two_star = bool(found_models & two_star_models)
    has_one_star = bool(found_models & one_star_models)
    
    # Base stars
    if has_two_star:
        stars = 2
    elif has_one_star:
        stars = 1
    else:
        stars = 0
    
    # Upgrade to 3 stars conditions
    if (has_one_star and has_two_star) or \
       ((has_one_star or has_two_star) and (largest_seq_length in top_2_largest)) or \
       ((has_one_star or has_two_star) and (models_found in top_5_models)):
        stars = 3
    
    # 4 stars conditions
    if (largest_seq_length in top_2_largest or models_found in top_5_models):
        # Check if only 0 or 0.0 in strength travelers
        strength_values = [s.strip() for s in str(strength_ids).split(",") if s.strip()]
        only_zero = all(v in ["0", "0.0"] for v in strength_values if v)
        
        # Check if has two unique strength travelers
        unique_strength = len(set([abs(float(v)) for v in strength_values if v and v not in ["0", "0.0"]])) >= 2
        
        if only_zero or unique_strength:
            stars = 4
    
    # Strength tracker 4 stars (handled separately in highlighting)
    if has_top_strength_criteria:
        stars = 4
        
    return stars

def calculate_highlighting_info(mega_df, input_report_ref):
    """
    Calculate all highlighting information for the dataframe
    """
    # Convert output strings back to float for comparison
    outputs = []
    for output_str in mega_df["Output"]:
        try:
            outputs.append(float(output_str.replace(",", "")))
        except:
            outputs.append(0)
    
    mega_df["_numeric_output"] = outputs
    
    # Get top values for highlighting
    largest_seq_lengths = mega_df["Largest sequence length"].tolist()
    models_found_counts = mega_df["Models found"].tolist()
    
    # Top 2 largest sequence lengths
    top_2_largest = sorted(set(largest_seq_lengths), reverse=True)[:2]
    
    # Top 5 models found 
    top_5_models = sorted(set(models_found_counts), reverse=True)[:5]
    
    highlighting_info = []
    
    for idx, row in mega_df.iterrows():
        info = {
            "output_color": "",
            "largest_seq_color": "",
            "models_found_color": "",
            "strength_id_color": "",
            "row_color": "",
            "strength_tracker_color": ""
        }
        
        output_val = row["_numeric_output"]
        
        # Output highlighting based on Input @ Report
        if input_report_ref:
            try:
                report_ref = float(str(input_report_ref).replace(",", ""))
                if output_val == report_ref:
                    info["output_color"] = "lightgray"
                elif output_val > report_ref:
                    info["output_color"] = "lightcoral"
                elif output_val < report_ref:
                    info["output_color"] = "lightblue"
            except:
                pass
        
        # Largest sequence length highlighting
        if row["Largest sequence length"] in top_2_largest:
            info["largest_seq_color"] = "plum"
        
        # Models found highlighting  
        if row["Models found"] in top_5_models:
            info["models_found_color"] = "plum"
        
        # M# Strength Traveler ID highlighting
        strength_ids = str(row["M # Strength Traveler ID"]).strip()
        if strength_ids and strength_ids != "":
            strength_values = [s.strip() for s in strength_ids.split(",") if s.strip()]
            has_zero = any(v in ["0", "0.0"] for v in strength_values)
            only_zero = all(v in ["0", "0.0"] for v in strength_values if v)
            
            if has_zero and not only_zero:
                info["strength_id_color"] = "lightgray"
            elif only_zero and len(strength_values) >= 1:
                info["strength_id_color"] = "plum"
        
        # Row highlighting for special conditions
        top_criteria = (row["Largest sequence length"] in top_2_largest or 
                       row["Models found"] in top_5_models)
        
        if top_criteria:
            strength_values = [s.strip() for s in str(row["M # Strength Traveler ID"]).split(",") if s.strip()]
            only_zero = all(v in ["0", "0.0"] for v in strength_values if v)
            unique_strength = len(set([abs(float(v)) for v in strength_values if v and v not in ["0", "0.0"]])) >= 2
            
            if only_zero:
                info["row_color"] = "plum"  # Light fuchsia
            elif unique_strength:
                info["row_color"] = "lightgreen"
        
        highlighting_info.append(info)
    
    return highlighting_info, top_2_largest, top_5_models

def apply_strength_tracker_highlighting(mega_df, highlighting_info, input_report_ref):
    """
    Apply strength tracker highlighting rules
    """
    if not input_report_ref:
        return highlighting_info
    
    try:
        report_ref = float(str(input_report_ref).replace(",", ""))
    except:
        return highlighting_info
    
    # Sort by output for strength tracker analysis
    df_sorted = mega_df.sort_values("_numeric_output").reset_index()
    
    for i, (orig_idx, row) in enumerate(df_sorted.iterrows()):
        strength_ids = str(row["M # Strength Traveler ID"]).strip()
        if not strength_ids or strength_ids == "":
            continue
            
        output_val = row["_numeric_output"]
        
        # Check if above report reference
        if output_val > report_ref:
            # Find next higher output with strength traveler
            next_higher = None
            for j in range(i + 1, len(df_sorted)):
                next_row = df_sorted.iloc[j]
                next_strength = str(next_row["M # Strength Traveler ID"]).strip()
                if next_strength and next_strength != "":
                    next_higher = next_row["_numeric_output"]
                    break
            
            if next_higher and abs(next_higher - output_val) > 25:
                highlighting_info[orig_idx]["strength_tracker_color"] = "lightcoral"  # Light red as specified
                highlighting_info[orig_idx]["row_color"] = "lightcoral"
        
        # Check if below report reference  
        elif output_val < report_ref:
            # Find next lower output with strength traveler
            next_lower = None
            for j in range(i - 1, -1, -1):
                next_row = df_sorted.iloc[j]
                next_strength = str(next_row["M # Strength Traveler ID"]).strip()
                if next_strength and next_strength != "":
                    next_lower = next_row["_numeric_output"]
                    break
            
            if next_lower and abs(output_val - next_lower) > 25:
                highlighting_info[orig_idx]["strength_tracker_color"] = "lightblue"
                highlighting_info[orig_idx]["row_color"] = "lightblue"
    
    return highlighting_info

def generate_comprehensive_mega_report(df):
    """
    Generate comprehensive single-line report with all Excel template columns
    """
    st.subheader("🎯 Single Line Mega Report - All Patterns Per Output")

    # Debug info for troubleshooting
    st.write(f"Debug: A_MODELS_AVAILABLE = {A_MODELS_AVAILABLE}")
    st.write(f"Debug: C_MODELS_AVAILABLE = {C_MODELS_AVAILABLE}")
    
    if not A_MODELS_AVAILABLE and not C_MODELS_AVAILABLE:
        st.error("No model detection functions available")
        return
    
    # Get report time and Input @ Report reference
    report_time = df["Arrival"].max()
    report_time_str = report_time.strftime("%Y%m%d_%H%M") if report_time else datetime.now().strftime("%Y%m%d_%H%M")
    input_report_ref = None
    
    # Group by output and analyze patterns
    mega_report_data = []
    outputs = df["Output"].unique()
    
    progress_bar = st.progress(0)
    for i, output in enumerate(outputs):
        output_rows = df[df["Output"] == output]
        
        # Skip if only one row (no patterns possible)
        if len(output_rows) < 2:
            continue
            
        # Get comprehensive analysis
        patterns, sequence_metrics = analyze_output_patterns_comprehensive(output_rows, report_time)
        
        if patterns:  # Only include outputs with found patterns
            # Calculate all metrics from Excel template
            metrics = calculate_comprehensive_metrics(output_rows, report_time)
            
            # Calculate booster score (placeholder - can be customized)
            booster_score = sequence_metrics["largest_seq_length"] * 10
            
            # Fix special characters for Excel compatibility
            pattern_sequences_fixed = " | ".join(patterns).replace("→", ",").replace("⭐", "*")
            m_start_end_fixed = metrics["m_start_end"].replace("→", ",")
            arrival_order_fixed = metrics["arrival_order"].replace("→", ",")
            seq_start_end_fixed = sequence_metrics["seq_start_end"].replace("→", ",")
            
            mega_report_data.append({
                "Output": f"{output:,.3f}",
                "Stars": "",  # Will be calculated later with complex rules
                "Booster Score": booster_score,
                "Input @ 18:00": metrics["input_18"],
                "18:00 Diff": metrics["diff_18"],
                "Input @ Arrival": metrics["input_arrival"], 
                "Arrival Diff": metrics["diff_arrival"],
                "Input @ Report": metrics["input_report"],
                "Report Diff": metrics["diff_report"],
                "Arrival": metrics["latest_arrival"].strftime('%a %y-%m-%d %H:%M'),
                "Day": "[0]" if "[0]" in str(output_rows["Day"].iloc[-1]) else "Other",
                "Hours ago": metrics["hours_ago"],
                "Total M# Count": metrics["total_m_count"],
                "Total M# ID": metrics["m_ids"],
                "Total M#s Start/End": m_start_end_fixed,
                "M # Arrival Order": arrival_order_fixed,
                "M # Strength Traveler Count": metrics["strength_count"],
                "M # Strength Traveler ID": metrics["strength_ids"],
                "Tag B Traveler Count": metrics["tag_b_count"],
                "Tag B Traveler ID": metrics["tag_b_ids"],
                "Pattern Sequences": pattern_sequences_fixed,
                "Unique Sequences": sequence_metrics["unique_sequences"],
                "Largest sequence length": sequence_metrics["largest_seq_length"],
                "Sequence Start/End": seq_start_end_fixed,
                "Sequence Same Origin": "Yes" if metrics["same_origin"] else "No",
                "Sequence Same Family": "Yes" if metrics["same_family"] else "No",
                "Sequence Same Feed": "Yes" if metrics["same_feed"] else "No",
                "Sequence Indigo Families": "",  # Placeholder for future implementation
                "Bucket indigo families": "",    # Placeholder for future implementation
                "Models found": len(patterns),
                # Store raw data for calculations
                "_raw_output": output,
                "_raw_patterns": patterns,
                "_raw_sequences": sequence_metrics["sequences_found"]
            })
        
        progress_bar.progress((i + 1) / len(outputs))
    
    if not mega_report_data:
        st.warning("No patterns found in any outputs")
        return
    
    # Create DataFrame and sort by output descending
    mega_df = pd.DataFrame(mega_report_data)
    mega_df_sorted = mega_df.sort_values(
        by="Output", 
        key=lambda x: x.str.replace(",", "").astype(float), 
        ascending=False
    )
    
    # Get Input @ Report reference for highlighting
    if len(mega_df_sorted) > 0:
        # Try to find a consistent Input @ Report value
        report_values = mega_df_sorted["Input @ Report"].dropna().unique()
        input_report_ref = report_values[0] if len(report_values) > 0 else None
    else:
        input_report_ref = None
    
    # Calculate highlighting information
    highlighting_info, top_2_largest, top_5_models = calculate_highlighting_info(mega_df_sorted, input_report_ref)
    highlighting_info = apply_strength_tracker_highlighting(mega_df_sorted, highlighting_info, input_report_ref)
    
    # Calculate star ratings with complex rules
    for idx, row in mega_df_sorted.iterrows():
        patterns = row["_raw_patterns"]
        largest_seq_length = row["Largest sequence length"]
        models_found = row["Models found"] 
        strength_ids = row["M # Strength Traveler ID"]
        
        # Check if has strength tracker criteria
        has_strength_tracker = highlighting_info[idx]["strength_tracker_color"] != ""
        
        stars = calculate_star_rating(
            patterns, largest_seq_length, models_found, strength_ids,
            top_2_largest, top_5_models, has_strength_tracker
        )
        
        mega_df_sorted.at[idx, "Stars"] = "*" * stars
    
    # Apply highlighting to display
    def highlight_cells(row):
        idx = row.name
        if idx >= len(highlighting_info):
            return [''] * len(row)
        
        info = highlighting_info[idx]
        colors = [''] * len(row)
        
        # Find column indices
        col_indices = {col: i for i, col in enumerate(row.index)}
        
        # Apply specific column highlighting
        if "Output" in col_indices and info["output_color"]:
            colors[col_indices["Output"]] = f'background-color: {info["output_color"]}'
        
        if "Largest sequence length" in col_indices and info["largest_seq_color"]:
            colors[col_indices["Largest sequence length"]] = f'background-color: {info["largest_seq_color"]}'
        
        if "Models found" in col_indices and info["models_found_color"]:
            colors[col_indices["Models found"]] = f'background-color: {info["models_found_color"]}'
        
        if "M # Strength Traveler ID" in col_indices and info["strength_id_color"]:
            colors[col_indices["M # Strength Traveler ID"]] = f'background-color: {info["strength_id_color"]}'
        
        # Apply row highlighting (but don't override specific column highlights)
        # For strength tracker, exclude Output, Largest sequence length, Models found as specified in Mod 3
        if info["row_color"] and info["strength_tracker_color"]:
            for i, col in enumerate(row.index):
                if col not in ["Output", "Largest sequence length", "Models found", "M # Strength Traveler ID"] and colors[i] == '':
                    colors[i] = f'background-color: {info["row_color"]}'
        elif info["row_color"]:
            # For other row highlighting, exclude the already highlighted columns
            for i, col in enumerate(row.index):
                if col not in ["Output", "Largest sequence length", "Models found", "M # Strength Traveler ID"] and colors[i] == '':
                    colors[i] = f'background-color: {info["row_color"]}'
        
        return colors
    
    # Remove helper columns before display
    display_df = mega_df_sorted.drop(columns=["_raw_output", "_raw_patterns", "_raw_sequences", "_numeric_output"], errors='ignore')
    
    st.write(f"Found patterns in {len(display_df)} outputs with comprehensive metrics and highlighting")
    
    # Display with highlighting
    styled_df = display_df.style.apply(highlight_cells, axis=1)
    st.dataframe(styled_df, use_container_width=True)
    
    # Download button with Excel-compatible data and highlighting
    download_df = display_df.copy()
    
    # Create highlighting for Excel export using conditional formatting markers
    for idx, row in download_df.iterrows():
        if idx < len(highlighting_info):
            info = highlighting_info[idx]
            
            # Add highlighting markers to cells for Excel
            if info["output_color"]:
                download_df.at[idx, "Output"] = f"[{info['output_color']}]{download_df.at[idx, 'Output']}"
            
            if info["largest_seq_color"]:
                download_df.at[idx, "Largest sequence length"] = f"[{info['largest_seq_color']}]{download_df.at[idx, 'Largest sequence length']}"
            
            if info["models_found_color"]:
                download_df.at[idx, "Models found"] = f"[{info['models_found_color']}]{download_df.at[idx, 'Models found']}"
            
            if info["strength_id_color"]:
                download_df.at[idx, "M # Strength Traveler ID"] = f"[{info['strength_id_color']}]{download_df.at[idx, 'M # Strength Traveler ID']}"
            
            # Add row highlighting markers to other columns
            if info["row_color"]:
                for col in download_df.columns:
                    if col not in ["Output", "Largest sequence length", "Models found", "M # Strength Traveler ID"]:
                        current_val = download_df.at[idx, col]
                        if not str(current_val).startswith(f"[{info['row_color']}]"):
                            download_df.at[idx, col] = f"[{info['row_color']}]{current_val}"
    
    # CSV download with highlighting markers
    csv_data = download_df.to_csv(index=False)
    st.download_button(
        "Download CSV with Highlighting Markers",
        data=csv_data,
        file_name=f"comprehensive_single_line_mega_report_{report_time_str}.csv",
        mime="text/csv"
    )
    
    # Excel download with actual highlighting
    try:
        # Create Excel file with highlighting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            display_df.to_excel(writer, sheet_name='Single Line Mega Report', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Single Line Mega Report']
            
            # Define color styles
            from openpyxl.styles import PatternFill
            
            lightgray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            lightcoral_fill = PatternFill(start_color='F08080', end_color='F08080', fill_type='solid')
            lightblue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            plum_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')
            lightgreen_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            color_map = {
                'lightgray': lightgray_fill,
                'lightcoral': lightcoral_fill,
                'lightblue': lightblue_fill,
                'plum': plum_fill,
                'lightgreen': lightgreen_fill
            }
            
            # Apply highlighting to cells
            for row_idx, (_, row) in enumerate(display_df.iterrows(), start=2):  # Start at row 2 (1 is header)
                if row_idx - 2 < len(highlighting_info):
                    info = highlighting_info[row_idx - 2]
                    
                    # Column mapping
                    col_map = {col: idx + 1 for idx, col in enumerate(display_df.columns)}
                    
                    # Apply specific column highlighting
                    if info["output_color"] and "Output" in col_map:
                        worksheet.cell(row=row_idx, column=col_map["Output"]).fill = color_map.get(info["output_color"], lightgray_fill)
                    
                    if info["largest_seq_color"] and "Largest sequence length" in col_map:
                        worksheet.cell(row=row_idx, column=col_map["Largest sequence length"]).fill = color_map.get(info["largest_seq_color"], plum_fill)
                    
                    if info["models_found_color"] and "Models found" in col_map:
                        worksheet.cell(row=row_idx, column=col_map["Models found"]).fill = color_map.get(info["models_found_color"], plum_fill)
                    
                    if info["strength_id_color"] and "M # Strength Traveler ID" in col_map:
                        worksheet.cell(row=row_idx, column=col_map["M # Strength Traveler ID"]).fill = color_map.get(info["strength_id_color"], plum_fill)
                    
                    # Apply row highlighting (excluding already highlighted columns)
                    if info["row_color"]:
                        excluded_cols = []
                        if info["strength_tracker_color"]:
                            excluded_cols = ["Output", "Largest sequence length", "Models found", "M # Strength Traveler ID"]
                        else:
                            excluded_cols = ["Output", "Largest sequence length", "Models found", "M # Strength Traveler ID"]
                        
                        for col_name, col_idx in col_map.items():
                            if col_name not in excluded_cols:
                                current_fill = worksheet.cell(row=row_idx, column=col_idx).fill
                                if current_fill.start_color.rgb == '00000000':  # No existing fill
                                    worksheet.cell(row=row_idx, column=col_idx).fill = color_map.get(info["row_color"], lightgray_fill)
        
        excel_data = output.getvalue()
        st.download_button(
            "📥 Download Excel with Actual Highlighting",
            data=excel_data,
            file_name=f"comprehensive_single_line_mega_report_{report_time_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        st.warning(f"Excel export with highlighting not available: {e}")
        # Fallback to simple Excel without highlighting
        simple_excel = display_df.to_excel(index=False)
        st.download_button(
            "Download Excel (Simple)",
            data=simple_excel,
            file_name=f"comprehensive_single_line_mega_report_{report_time_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    return display_df

def run_simple_single_line_analysis(df):
    """Main function to run comprehensive single line mega report analysis"""
    try:
        return generate_comprehensive_mega_report(df)
    except Exception as e:
        st.error(f"Error generating mega report: {e}")
        st.write("Traceback:", str(e))
        return None
