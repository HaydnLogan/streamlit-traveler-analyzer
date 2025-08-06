import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from collections import defaultdict
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Master M# list - all 118 values
MASTER_M_VALUES = [
    111, 107, 103, 101.0, 98.2, 99.1, 98.2, 98.1, 97.2, 97.1, 96.1, 96, 95.5, 95, 93.5, 92, 90.5, 89, 87, 86.5, 85, 83.00, 80, 78.01, 77, 74, 71, 68, 67, 62, 60, 57, 55, 54, 50, 47, 45, 43, 41, 40, 39, 38, 36, 33, 30, 27, 24, 22, 15, 14, 12, 10, 6, 5, 3, 2, 1, 0, -1, -2, -3, -5, -6, -10, -12, -14, -15, -22, -24, -27, -30, -33, -36, -38, -39, -40, -41, -43, -45, -47, -50, -54, -55, -57, -60, -62, -67, -68, -71, -74, -77, -78.01, -80, -83.00, -85, -86.5, -87, -89, -90.5, -92, -93.5, -95, -95.5, -96, -96.1, -97.1, -97.2, -98.1, -98.2, -99.1, -98.2, -101.0, -103, -107, -111
]

# Group definitions
GROUP_1A_VALUES = [111, 107, 103, 96, 87, 77, 68, 60, 50, -50, -60, -68, -77, -87, -96, -103, -107, -111]

GROUP_1B_VALUES = [111, 107, 103, 96, 87, 77, 68, 60, 55, 50, 43, 41, 40, 39, 36, 30, 22, 14, 10, 6, 5, 3, 2, 1, 0, -1, -2, -3, -5, -6, -10, -14, -22, -30, -36, -39, -40, -41, -43, -50, -55, -60, -68, -77, -87, -96, -103, -107, -111]

GROUP_2A_VALUES = [101.0, 98.2, 99.1, 98.2, 98.1, 97.2, 97.1, 96.1, 95.5, 95, 93.5, 92, 90.5, 89, 86.5, 85, 83.00, 80, 78.01, 74, 71, 67, 62, 54, 40, 0, -40, -54, -62, -67, -71, -74, -78.01, -80, -83.00, -85, -86.5, -89, -90.5, -92, -93.5, -95, -95.5, -96.1, -97.1, -97.2, -98.1, -98.2, -99.1, -98.2, -101.0]

GROUP_2B_VALUES = [101.0, 98.2, 99.1, 98.2, 98.1, 97.2, 97.1, 96.1, 95.5, 95, 93.5, 92, 90.5, 89, 86.5, 85, 83.00, 80, 78.01, 74, 71, 67, 62, 57, 54, 47, 45, 40, 38, 33, 27, 24, 15, 12, 0, -12, -15, -24, -27, -33, -38, -40, -45, -47, -54, -57, -62, -67, -71, -74, -78.01, -80, -83.00, -85, -86.5, -89, -90.5, -92, -93.5, -95, -95.5, -96.1, -97.1, -97.2, -98.1, -98.2, -99.1, -98.2, -101.0]

def get_measurement_value(row):
    """Extract measurement value from row using flexible column naming"""
    for col in ['m #', 'M #', 'M value', 'm value', 'measurement']:
        if col in row and pd.notna(row[col]):
            return float(row[col])
    return 0

def calculate_pivot(H, L, C, m_value):
    """Calculate output using pivot formula"""
    try:
        H, L, C, m_value = float(H), float(L), float(C), float(m_value)
        return ((H + L + C) / 3) + (m_value / 1000)
    except:
        return 0

def get_day_index(arrival_time, report_time, start_hour):
    """Calculate day index based on arrival and report times"""
    try:
        # Convert to datetime if needed
        if isinstance(arrival_time, str):
            arrival_time = pd.to_datetime(arrival_time)
        if isinstance(report_time, str):
            report_time = pd.to_datetime(report_time)
        
        # Calculate day difference
        day_diff = (report_time.date() - arrival_time.date()).days
        return f"[{day_diff}]"
    except:
        return "[0]"

def get_input_at_time(small_df, target_time):
    """Get input value at specific time"""
    try:
        if target_time is None or small_df.empty:
            return None
        
        # Find closest time
        small_df['time_diff'] = abs(small_df['time'] - target_time)
        closest_idx = small_df['time_diff'].idxmin()
        return small_df.loc[closest_idx, 'input']
    except:
        return None

def generate_master_traveler_list(data, report_time, start_hour=17, fast_mode=True):
    """Generate master traveler list with all M# values from master list"""
    
    # Ensure time column exists and is datetime
    if 'time' not in data.columns:
        return pd.DataFrame()
    
    data['time'] = pd.to_datetime(data['time'])
    
    # Define measurement columns
    price_cols = ['high', 'low', 'close']
    if not all(col in data.columns for col in price_cols):
        return pd.DataFrame()
    
    # Get unique origins
    origins = data['origin'].unique() if 'origin' in data.columns else ['default']
    
    # Process each origin
    all_traveler_data = []
    
    for origin in origins:
        origin_data = data[data['origin'] == origin] if 'origin' in data.columns else data
        
        if len(origin_data) < 2:
            continue
            
        # Sort by time
        origin_data = origin_data.sort_values('time')
        
        # Find price changes
        for i in range(len(origin_data) - 1):
            current = origin_data.iloc[i]
            next_row = origin_data.iloc[i + 1]
            
            # Check if any price changed
            price_changed = any(current[col] != next_row[col] for col in price_cols)
            
            if price_changed:
                arrival_time = current['time']
                H, L, C = current['high'], current['low'], current['close']
                
                # Calculate input values
                input_at_arrival = get_input_at_time(data, arrival_time)
                input_at_report = get_input_at_time(data, report_time)
                input_at_start = get_input_at_time(data, report_time.replace(hour=start_hour, minute=0, second=0))
                
                # Generate entries for ALL master M# values
                for m_value in MASTER_M_VALUES:
                    output = calculate_pivot(H, L, C, m_value)
                    day = get_day_index(arrival_time, report_time, start_hour)
                    
                    # Format arrival time
                    try:
                        day_abbrev = arrival_time.strftime('%a')
                        arrival_excel = arrival_time.strftime('%d-%b-%Y %H:%M')
                    except:
                        day_abbrev = ""
                        arrival_excel = str(arrival_time)
                    
                    # Classify group
                    if m_value in GROUP_1A_VALUES:
                        group = "Grp 1a"
                    elif m_value in GROUP_1B_VALUES:
                        group = "Grp 1b"
                    elif m_value in GROUP_2A_VALUES:
                        group = "Grp 2a"
                    elif m_value in GROUP_2B_VALUES:
                        group = "Grp 2b"
                    else:
                        group = "Unclassified"
                    
                    traveler_entry = {
                        "Feed": "auto",
                        "ddd": day_abbrev,
                        "Arrival": arrival_excel,
                        "Arrival_datetime": arrival_time,
                        "Day": day,
                        "Origin": origin,
                        "M Name": f"M{m_value}",
                        "M #": m_value,
                        "R #": "",
                        "Tag": "",
                        "Family": "",
                        "Group": group,
                        f"Input @ {start_hour:02d}:00": input_at_start,
                        f"Diff @ {start_hour:02d}:00": output - input_at_start if input_at_start is not None else None,
                        "Input @ Arrival": input_at_arrival,
                        "Diff @ Arrival": output - input_at_arrival if input_at_arrival is not None else None,
                        "Input @ Report": input_at_report,
                        "Diff @ Report": output - input_at_report if input_at_report is not None else None,
                        "Output": output
                    }
                    
                    all_traveler_data.append(traveler_entry)
    
    # Convert to DataFrame
    master_df = pd.DataFrame(all_traveler_data)
    
    # Remove Arrival_datetime for final output
    if 'Arrival_datetime' in master_df.columns:
        master_df = master_df.drop('Arrival_datetime', axis=1)
    
    return master_df

def filter_master_list_by_group(master_df, group_m_values):
    """Filter master traveler list by specific M# values for a group"""
    if master_df.empty:
        return pd.DataFrame()
    
    # Filter by M# values
    filtered_df = master_df[master_df['M #'].isin(group_m_values)].copy()
    
    return filtered_df

def create_excel_with_highlighting(reports_dict, filename):
    """Create Excel file with proper highlighting for all 4 reports"""
    
    # Create Excel buffer
    excel_buffer = io.BytesIO()
    
    # Create workbook manually to avoid BytesIO issues
    workbook = Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    for sheet_name, df in reports_dict.items():
        if not df.empty:
            # Create worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(row)
            
            # Apply formatting
            apply_basic_formatting(worksheet, df)
    
    # Save to buffer
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

def apply_basic_formatting(worksheet, df):
    """Apply basic formatting to Excel worksheet"""
    
    # Define colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Format headers
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
