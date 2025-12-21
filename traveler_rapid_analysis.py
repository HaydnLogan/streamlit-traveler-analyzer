"""
Traveler Rapid Analysis Dashboard
Processes data grab files and identifies high-priority trading zones
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import sys

# Traveler Classifications
EPIC_ORIGINS = {"Trinidad", "Tobago", "WASP-12b", "Macedonia"}
ANCHOR_ORIGINS = {"Spain", "Saturn", "Jupiter", "Kepler-62", "Kepler-44"}

# FOGZ (Friends of Green Zero)
FOGZ_TRAVELERS = {0, 40, -40}

# Family classifications for pattern matching
GREEN_FAMILY = {
    2, -2, 10, -10, 21, -21, 22, -22, 30, -30, 36, -36, 39, -39, 41, -41, 43, -43, 50, -50, 60, -60, 77, -77, 107, -107,
    5, -5, 14, -14, 26, -26, 55, -55, 56, -56, 68, -68, 91, -91, 96, -96,
    6, -6, 25, -25, 49, -49, 79, -79, 87, -87,
    3, -3, 37, -37, 63, -63, 99, -99.2, 103, -103,
    1, -1, 73, -73, 111, -111,
    52, -52, 59, -59, 65, -65, 70, -70, 76, -76, 82, -82, 86, -86, 88, -88, 97, -97
}

INDIGO_FAMILY = {
    0, 40, -40,  # Wild
    15, -15, 27, -27, 33, -33, 38, -38, 42, -42, 45, -45, 54, -54, 67, -67, 74, -74, 80, -80, 85, -85, 89, -89, 92, -92, 95, -95, 96.1, -96.1, 97.2, -97.2, 98.2, -98.2, 99.3, -99.3,
    4, -4, 12, -12, 24, -24, 31, -31, 47, -47, 57, -57, 71, -71, 93.5, -93.5, 101, -101,
    62, -62, 78.01, -78.01, 83, -83, 86.5, -86.5, 90.5, -90.5, 95.5, -95.5, 97.1, -97.1, 98.1, -98.1, 99.1, -99.1
}

X0_TAGS = {"X01p", "X02p", "X03p", "X04p", "X05p", "X06p", "X07p", "X08p", "X09p",
           "X01d", "X02d", "X03d", "X04d", "X05d", "X06d", "X07d", "X08d", "X09d"}
XD0_TAGS = {"XD01p", "XD02p", "XD03p", "XD04p", "XD05p", "XD06p", "XD07p", "XD08p", "XD09p", "XD010p",
            "XD01d", "XD02d"}

def load_data_grab(filepath):
    """Load the data grab Excel file - handles both raw and processed formats"""
    # Try loading as processed swing analysis format first (with header at row 2)
    try:
        df = pd.read_excel(filepath, sheet_name=0, header=2)
        # Check if it has the expected columns
        if 'M_#s' in df.columns and 'Outputs' in df.columns:
            return df, 'processed'
    except:
        pass
    
    # Fall back to raw data grab format
    df = pd.read_excel(filepath, sheet_name=0, header=0)
    return df, 'raw'

def calculate_proximity(outputs):
    """Calculate proximity between two outputs"""
    if pd.isna(outputs) or ',' not in str(outputs):
        return None
    parts = str(outputs).split(',')
    if len(parts) == 2:
        try:
            return abs(float(parts[0].strip()) - float(parts[1].strip()))
        except:
            return None
    return None

def parse_m_numbers(m_str):
    """Parse M# string into tuple of numbers"""
    if pd.isna(m_str):
        return None
    parts = str(m_str).replace(' ', '').split(',')
    if len(parts) == 2:
        try:
            return (float(parts[0]), float(parts[1]))
        except:
            return None
    return None

def parse_origins(origin_str):
    """Parse origin string into tuple"""
    if pd.isna(origin_str):
        return None
    parts = str(origin_str).split(',')
    if len(parts) == 2:
        return (parts[0].strip(), parts[1].strip())
    return None

def score_zone(row, feed_counts):
    """
    Score a zone based on multiple factors:
    - M# 0 presence (highest priority)
    - Anchor origins
    - Epic origins
    - Same family pairs
    - X0/XD0 tags
    - Multi-feed confirmation
    - Proximity (tighter is better)
    """
    score = 0
    reasons = []
    
    m_nums = parse_m_numbers(row.get('M_#s', ''))
    origins = parse_origins(row.get('Origins', ''))
    
    # M# 0 presence (100 points)
    if m_nums and 0 in m_nums:
        score += 100
        reasons.append("M# 0 (FOGZ)")
    
    # FOGZ presence (80 points)
    if m_nums and any(m in FOGZ_TRAVELERS for m in m_nums):
        score += 80
        reasons.append("FOGZ traveler")
    
    # Anchor origins (50 points per anchor)
    if origins:
        anchor_count = sum(1 for o in origins if o in ANCHOR_ORIGINS)
        if anchor_count > 0:
            score += anchor_count * 50
            reasons.append(f"{anchor_count} Anchor origin(s)")
    
    # Epic origins (30 points per epic)
    if origins:
        epic_count = sum(1 for o in origins if o in EPIC_ORIGINS)
        if epic_count > 0:
            score += epic_count * 30
            reasons.append(f"{epic_count} Epic origin(s)")
    
    # Same family pair (40 points)
    if m_nums:
        both_green = all(m in GREEN_FAMILY for m in m_nums)
        both_indigo = all(m in INDIGO_FAMILY for m in m_nums)
        if both_green or both_indigo:
            score += 40
            family = "Green" if both_green else "Indigo"
            reasons.append(f"Same {family} family")
    
    # X0 or XD0 tags (30 points)
    tag = row.get('Tag', '')
    if tag in X0_TAGS or tag in XD0_TAGS:
        score += 30
        reasons.append(f"{tag} tag")
    
    # Multi-feed confirmation (60 points)
    outputs = row.get('Output', '')
    if pd.notna(outputs) and outputs in feed_counts and feed_counts[outputs] > 1:
        score += 60
        reasons.append("Multi-feed match")
    
    # Proximity bonus (tighter = better, max 20 points)
    prox = calculate_proximity(row.get('Outputs', ''))
    if prox is not None and prox < 10:
        prox_score = int((10 - prox) * 2)
        score += prox_score
        reasons.append(f"Tight prox ({prox:.2f})")
    
    return score, "; ".join(reasons)

def identify_gandalf_zones(df):
    """
    Identify Gandalf zones: (0, +/-40) combinations that appear on only ONE feed
    These are "You shall not pass" zones
    """
    # Filter for (0, 40) or (0, -40) combinations
    fogz_matches = df[df['M_#s'].str.contains('0, 40|0, -40', na=False, regex=True)].copy()
    
    if len(fogz_matches) == 0:
        return []
    
    # Parse outputs into individual values
    output_data = []
    for _, row in fogz_matches.iterrows():
        outputs_str = str(row.get('Outputs', ''))
        if ',' in outputs_str:
            outputs = [float(x.strip()) for x in outputs_str.split(',')]
            for out in outputs:
                output_data.append({
                    'Output': out,
                    'Feed': row['Feed'],
                    'M_#s': row['M_#s'],
                    'Origins': row.get('Origins', ''),
                    'Full_Row': row
                })
    
    output_df = pd.DataFrame(output_data)
    
    # Count feeds per output
    feed_counts = output_df.groupby('Output')['Feed'].nunique()
    
    # Gandalf zones appear on only ONE feed
    gandalf_outputs = feed_counts[feed_counts == 1].index.tolist()
    
    gandalf_zones = []
    for out in gandalf_outputs:
        zone_data = output_df[output_df['Output'] == out].iloc[0]
        gandalf_zones.append({
            'Output': out,
            'Feed': zone_data['Feed'],
            'M_#s': zone_data['M_#s'],
            'Origins': zone_data['Origins'],
            'Type': 'GANDALF ZONE',
            'Description': f'{zone_data["Feed"]} feed only - Price should not stay past this level'
        })
    
    return gandalf_zones

def create_dashboard(df, report_time, output_file):
    """Create the rapid analysis dashboard"""
    
    # Extract individual outputs for multi-feed counting
    # From the Outputs column which contains pairs like "25820.10, 25820.67"
    output_feed_map = {}
    for _, row in df.iterrows():
        outputs_str = str(row.get('Outputs', ''))
        feed = row.get('Feed', '')
        if ',' in outputs_str:
            outputs = [float(x.strip()) for x in outputs_str.split(',')]
            for out in outputs:
                if out not in output_feed_map:
                    output_feed_map[out] = set()
                output_feed_map[out].add(feed)
    
    # Count unique feeds per output
    output_feed_counts = {out: len(feeds) for out, feeds in output_feed_map.items()}
    
    # Use Arrival_Output as the primary output for sorting/display
    df['Display_Output'] = df.get('Arrival_Output', df.get('Output', ''))
    
    # Score all zones
    df['Score'], df['Reasons'] = zip(*df.apply(lambda row: score_zone(row, output_feed_counts), axis=1))
    
    # Sort by score
    df_sorted = df.sort_values('Score', ascending=False)
    
    # Get top zones
    top_zones = df_sorted.head(50).copy()
    
    # Identify Gandalf zones
    gandalf_zones = identify_gandalf_zones(df)
    
    # Create workbook
    wb = Workbook()
    
    # === PRIORITY ZONES SHEET ===
    ws_priority = wb.active
    ws_priority.title = "Priority Zones"
    
    # Header
    ws_priority['A1'] = 'TRAVELER RAPID ANALYSIS DASHBOARD'
    ws_priority['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_priority['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws_priority.merge_cells('A1:K1')
    
    ws_priority['A2'] = f'Report Time: {report_time}'
    ws_priority['A2'].font = Font(size=11, italic=True)
    ws_priority.merge_cells('A2:K2')
    
    ws_priority['A3'] = f'Total Matches: {len(df)} | Top {len(top_zones)} High-Priority Zones'
    ws_priority['A3'].font = Font(size=10)
    ws_priority.merge_cells('A3:K3')
    
    # Column headers
    headers = ['Rank', 'Score', 'Output', 'M_#s', 'Origins', 'Feed', 'Tag', 'Proximity', 'Arrival', 'Day', 'Priority Reasons']
    for col, header in enumerate(headers, 1):
        cell = ws_priority.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for idx, (_, row) in enumerate(top_zones.iterrows(), 1):
        r = idx + 5
        ws_priority.cell(r, 1, idx)
        ws_priority.cell(r, 2, row['Score'])
        ws_priority.cell(r, 3, row.get('Display_Output', ''))
        ws_priority.cell(r, 4, row.get('M_#s', ''))
        ws_priority.cell(r, 5, row.get('Origins', ''))
        ws_priority.cell(r, 6, row.get('Feed', ''))
        ws_priority.cell(r, 7, row.get('Tags', row.get('Tag', '')))
        
        prox = calculate_proximity(row.get('Outputs', ''))
        ws_priority.cell(r, 8, f"{prox:.2f}" if prox else "")
        
        arrival = row.get('Arrival_DateTime', row.get('Arrival', ''))
        ws_priority.cell(r, 9, str(arrival) if pd.notna(arrival) else "")
        ws_priority.cell(r, 10, row.get('Day', ''))
        ws_priority.cell(r, 11, row.get('Reasons', ''))
        
        # Color coding by score
        if row['Score'] >= 200:
            fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        elif row['Score'] >= 150:
            fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        else:
            fill = None
        
        if fill:
            for col in range(1, 12):
                ws_priority.cell(r, col).fill = fill
    
    # Set column widths
    ws_priority.column_dimensions['A'].width = 8
    ws_priority.column_dimensions['B'].width = 10
    ws_priority.column_dimensions['C'].width = 12
    ws_priority.column_dimensions['D'].width = 12
    ws_priority.column_dimensions['E'].width = 25
    ws_priority.column_dimensions['F'].width = 10
    ws_priority.column_dimensions['G'].width = 10
    ws_priority.column_dimensions['H'].width = 10
    ws_priority.column_dimensions['I'].width = 20
    ws_priority.column_dimensions['J'].width = 10
    ws_priority.column_dimensions['K'].width = 50
    
    # === GANDALF ZONES SHEET ===
    if gandalf_zones:
        ws_gandalf = wb.create_sheet("Gandalf Zones")
        
        ws_gandalf['A1'] = 'GANDALF ZONES - "You Shall Not Pass"'
        ws_gandalf['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_gandalf['A1'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        ws_gandalf.merge_cells('A1:F1')
        
        ws_gandalf['A2'] = 'Single-feed (0, +/-40) matches - Price should not stay past these levels'
        ws_gandalf['A2'].font = Font(size=10, italic=True)
        ws_gandalf.merge_cells('A2:F2')
        
        headers_g = ['Output', 'Feed', 'M_#s', 'Origins', 'Type', 'Description']
        for col, header in enumerate(headers_g, 1):
            cell = ws_gandalf.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        for idx, zone in enumerate(gandalf_zones, 1):
            r = idx + 4
            ws_gandalf.cell(r, 1, zone['Output'])
            ws_gandalf.cell(r, 2, zone['Feed'])
            ws_gandalf.cell(r, 3, zone['M_#s'])
            ws_gandalf.cell(r, 4, zone['Origins'])
            ws_gandalf.cell(r, 5, zone['Type'])
            ws_gandalf.cell(r, 6, zone['Description'])
            
            # Highlight in red
            for col in range(1, 7):
                ws_gandalf.cell(r, col).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        ws_gandalf.column_dimensions['A'].width = 12
        ws_gandalf.column_dimensions['B'].width = 10
        ws_gandalf.column_dimensions['C'].width = 12
        ws_gandalf.column_dimensions['D'].width = 25
        ws_gandalf.column_dimensions['E'].width = 15
        ws_gandalf.column_dimensions['F'].width = 60
    
    # === FOGZ ZONES SHEET ===
    fogz_data = df[df['M_#s'].str.contains('0, 40|0, -40', na=False, regex=True)].copy()
    if len(fogz_data) > 0:
        ws_fogz = wb.create_sheet("FOGZ (0,±40)")
        
        ws_fogz['A1'] = 'FOGZ M# 0 with ±40 Combinations'
        ws_fogz['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_fogz['A1'].fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        ws_fogz.merge_cells('A1:H1')
        
        headers_f = ['Arrival Output', 'M_#s', 'Outputs', 'Origins', 'Feed', 'Tags', 'Arrival DateTime', 'Day']
        for col, header in enumerate(headers_f, 1):
            cell = ws_fogz.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        for idx, (_, row) in enumerate(fogz_data.iterrows(), 1):
            r = idx + 3
            ws_fogz.cell(r, 1, row.get('Display_Output', ''))
            ws_fogz.cell(r, 2, row.get('M_#s', ''))
            ws_fogz.cell(r, 3, row.get('Outputs', ''))
            ws_fogz.cell(r, 4, row.get('Origins', ''))
            ws_fogz.cell(r, 5, row.get('Feed', ''))
            ws_fogz.cell(r, 6, row.get('Tags', row.get('Tag', '')))
            arrival = row.get('Arrival_DateTime', row.get('Arrival', ''))
            ws_fogz.cell(r, 7, str(arrival) if pd.notna(arrival) else "")
            ws_fogz.cell(r, 8, row.get('Day', ''))
        
        for col in 'ABCDEFGH':
            ws_fogz.column_dimensions[col].width = 15
    
    # Save
    wb.save(output_file)
    return len(top_zones), len(gandalf_zones)

def main(data_grab_file, output_file=None):
    """Main analysis function"""
    print(f"Loading data from: {data_grab_file}")
    df, file_type = load_data_grab(data_grab_file)
    print(f"Detected file type: {file_type}")
    
    # Extract report time from filename or data
    try:
        xl = pd.ExcelFile(data_grab_file)
        sheet_name = xl.sheet_names[0]
        report_time = sheet_name
    except:
        report_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_file = f"/mnt/user-data/outputs/traveler_analysis_{timestamp}.xlsx"
    
    print(f"Processing {len(df)} matches...")
    
    top_count, gandalf_count = create_dashboard(df, report_time, output_file)
    
    print(f"\nAnalysis complete!")
    print(f"Top zones identified: {top_count}")
    print(f"Gandalf zones found: {gandalf_count}")
    print(f"Dashboard saved to: {output_file}")
    
    return output_file

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python traveler_rapid_analysis.py <data_grab_file> [output_file]")
        sys.exit(1)
    
    data_file = sys.argv[1]
    output = sys.argv[2] if len(sys.argv) > 2 else None
    main(data_file, output)
