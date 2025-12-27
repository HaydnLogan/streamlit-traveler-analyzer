# v33 - 12.27.25 ADDED Full Range (multiple) processing mode + Origin Filtering
# Major changes:
# - Renamed "Full Range" → "Full Range (single)" 
# - Added "Full Range (multiple)" mode for multi-day batch processing (up to 30 days)
# - Integrated Origin Filtering from Swing Analysis Tool
# - Multi-day mode produces one Excel file with 4 tabs per day (1800, 0315, 0900, 1230)
# - Each tab combines Grp_1a + Grp_1b, sorted by Output (desc) then Arrival (asc)
# - Auto-filtered headers with freeze panes on row 1
# - Filename format: asset_traveler_reports_for_trading_day

# v31i - 11.8.25 REMOVED time-based proximity, added global output spread filter, optimized G.11 for same-feed matching
# Major changes:
# - Removed "Proximity Threshold (hrs)" slider - was originally intended for output spread, not time
# - Added global "Output Spread Filter (pts)" for ALL G models  
# - G.11 now groups by feed FIRST, then looks for same-feed pairs (67% fewer wasted comparisons)
# - Removed all time-based group_by_proximity() calls
# - Neighbor detection now based on output proximity within same feed, not time proximity

# 11.6.25 Model G.11 added.

# v31f - Process Timeframes Separately" Option add.  with one high and one low range, current process take 200+ seconds.  
# this mod should bring it down to under 90 seconds.

# v31e - Processing time optimization to remove redundant data points.  Zone Sorting Fix #2

# v31d - Asset ID in Filename; for Custom Ranges path, Changed sort order from ['Group', 'Output', 'Arrival'] to ['Range', 'Group', 'Output', 'Arrival']

# v31c - "Most Current" mode: Now uses dt.datetime.now() as the report time (current timestamp)
# This ensures that report_time always has a valid datetime value when processing, preventing the NaN conversion error.

# v31b - Updated file uploads (7 files), added asset selector, modified custom ranges to single tab output
# Previous: v30b - Added HOD/LOD report mode with multi-day processing

import streamlit as st
import pandas as pd
import datetime as dt
import io
from typing import Optional
from pandas import ExcelWriter
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from custom_range_calculator_0813 import apply_custom_ranges_advanced, apply_full_range_advanced
from hod_lod_processor import (
    process_hod_lod_mode, 
    is_complete_trading_day, 
    get_trading_day_bounds
)

# Configure pandas to handle large datasets
pd.set_option("styler.render.max_elements", 2000000)

# Import functions
from a_helpers import (
    clean_timestamp, process_feed, get_input_at_day_start, apply_excel_highlighting,
    GROUP_1A_TRAVELERS, GROUP_1B_TRAVELERS, GROUP_2A_TRAVELERS, GROUP_2B_TRAVELERS,
    get_input_value, highlight_traveler_report, get_input_at_time, highlight_custom_traveler_report,
    generate_master_traveler_list, EPIC_ORIGINS, ANCHOR_ORIGINS   
)

# Model imports
from model_g_manager import run_model_g_detection

try:
    from models.models_a_today import run_a_model_detection_today
except ImportError:
    def run_a_model_detection_today(df):
        st.warning("Model A detection not available in this environment")

try:
    from models.mod_b_05pg1 import run_b_model_detection
except ImportError:
    def run_b_model_detection(df):
        st.warning("Model B detection not available in this environment")

try:
    from models.mod_c_04gpr3 import run_c_model_detection
except ImportError:
    def run_c_model_detection(df, run_c01=True, run_c02=True, run_c04=True):
        st.warning("Model C detection not available in this environment")

try:
    from models.mod_x_03g import run_x_model_detection
except ImportError:
    def run_x_model_detection(df):
        st.warning("Model X detection not available in this environment")

try:
    from models.simple_mega_report2 import run_simple_single_line_analysis
except ImportError:
    def run_simple_single_line_analysis(df):
        st.warning("Single Line Mega Report not available in this environment")


# === Helper Functions for Full Range (multiple) ===

def get_trading_day_from_datetime(dt_obj):
    """
    Given a datetime (e.g., 1800 on 7-Dec), return the trading day date.
    The 1800 open on 7-Dec is the start of 8-Dec trading day.
    If time is before 1800 (6PM), it's the same day. If >= 1800, it's the next day.
    """
    if dt_obj.hour >= 18:
        return (dt_obj + dt.timedelta(days=1)).date()
    else:
        return dt_obj.date()

def format_tab_name(dt_obj):
    """
    Format datetime to tab name like '07Dec 1800' or '08Dec 0315'
    """
    return dt_obj.strftime("%d%b %H%M")

def combine_and_format_groups(grp_1a_df, grp_1b_df):
    """
    Combine Grp_1a and Grp_1b DataFrames, sort by Output (desc) then Arrival (asc),
    and add Star and Reason columns.
    """
    # Combine the groups
    combined = pd.concat([grp_1a_df, grp_1b_df], ignore_index=True)
    
    # Sort by Output (desc) then Arrival (asc)
    if not combined.empty and 'Output' in combined.columns and 'Arrival' in combined.columns:
        combined = combined.sort_values(['Output', 'Arrival'], ascending=[False, True])
    elif not combined.empty and 'Output' in combined.columns:
        combined = combined.sort_values(['Output'], ascending=[False])
    
    # Add Star and Reason columns if they don't exist
    if 'Star' not in combined.columns:
        combined['Star'] = pd.NA
    if 'Reason' not in combined.columns:
        combined['Reason'] = pd.NA
    
    return combined

def create_multi_day_excel(all_day_reports, asset_id, trading_day_base):
    """
    Create Excel file with multiple days of reports.
    Each day has 4 tabs (1800, 0315, 0900, 1230).
    
    Args:
        all_day_reports: List of tuples (tab_name, dataframe)
        asset_id: Asset identifier (e.g., 'nq', 'es')
        trading_day_base: The first trading day for filename
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for tab_name, df in all_day_reports:
            df.to_excel(writer, sheet_name=tab_name, index=False)
    
    # Apply auto-filter and freeze panes
    output.seek(0)
    wb = load_workbook(output)
    
    for sheet in wb.worksheets:
        if sheet.max_row > 0:
            # Auto-filter on header row
            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}1"
            
            # Freeze header row (row 1)
            sheet.freeze_panes = 'A2'
            
            # Make header bold
            for cell in sheet[1]:
                cell.font = Font(bold=True)
    
    output_final = io.BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    
    return output_final


# === Unified Export Helper ===
def render_unified_export(traveler_reports, report_time, asset_id=""):
    if not traveler_reports:
        return

    st.markdown("---")
    st.markdown("### 📥 Unified Excel Download")

    # Add asset_id to filename if provided
    asset_prefix = f"{asset_id.lower()}_" if asset_id else ""
    report_datetime_str = report_time.strftime("%d-%b-%y_%H-%M")

    def _coerce_arrival_datetime(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        if "Arrival_datetime" in df.columns:
            df["Arrival"] = pd.to_datetime(df["Arrival_datetime"], errors="coerce")
        elif "Arrival" in df.columns:
            df["Arrival"] = pd.to_datetime(df["Arrival"], errors="coerce", infer_datetime_format=True)
        return df

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for tab_name, df_data in traveler_reports.items():
            if df_data is None or df_data.empty:
                continue
            df_coerced = _coerce_arrival_datetime(df_data)
            df_coerced.to_excel(writer, sheet_name=tab_name, index=False)

    output.seek(0)
    filename = f"{asset_prefix}traveler_report_{report_datetime_str}.xlsx"
    st.download_button(
        label="📥 Download Unified Excel Report",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Streamlit app
st.title("🌌 Traveler Report Generator")

# Sidebar: Asset Selector
st.sidebar.markdown("### Asset Selection")
asset_id = st.sidebar.selectbox(
    "Select Asset",
    ["NQ", "ES", "YM", "RTY"],
    index=0,
    help="Choose the asset for this analysis"
)

st.sidebar.markdown("---")

# File uploads
st.markdown("### 📁 Upload Files")

# Row 1: Small feeds (3m, 5m, 15m)
col1, col2, col3 = st.columns(3)
with col1:
    small_3m_file = st.file_uploader("Upload 3m small", type="csv", key="small_3m")
with col2:
    small_5m_file = st.file_uploader("Upload 5m small", type="csv", key="small_5m")
with col3:
    small_15m_file = st.file_uploader("Upload 15m small", type="csv", key="small_15m")

# Row 2: Big feeds (3m, 5m, 15m)
col4, col5, col6 = st.columns(3)
with col4:
    big_3m_file = st.file_uploader("Upload 3m big", type="csv", key="big_3m")
with col5:
    big_5m_file = st.file_uploader("Upload 5m big", type="csv", key="big_5m")
with col6:
    big_15m_file = st.file_uploader("Upload 15m big", type="csv", key="big_15m")

# Row 3: Measurement file
measurement_file = st.file_uploader("Upload measurement file", type=["xlsx", "xls"])

# Optional: Upload Final Traveler Report (bypass feed upload)
st.markdown("---")
st.markdown("### 📂 Optional: Upload Final Traveler Report (bypass feed upload)")
bypass_traveler_file = st.file_uploader(
    "Upload Final Traveler Report",
    type=['xlsx', 'csv'],
    help="Skip feed processing and upload traveler report directly"
)

# === Origin Filtering Section ===
st.markdown("---")
st.markdown("### 🌍 Origin Filtering")

# Extract all origins from uploaded CSVs
all_origins = set()
files_checked = 0
files_with_origins = 0

for file_obj, name in [(small_3m_file, '3m small'), (small_5m_file, '5m small'), (small_15m_file, '15m small'), 
                        (big_3m_file, '3m big'), (big_5m_file, '5m big'), (big_15m_file, '15m big')]:
    if file_obj is not None:
        files_checked += 1
        try:
            temp_df = pd.read_csv(file_obj)
            
            # Check for origin column (case-insensitive)
            origin_col = None
            for col in temp_df.columns:
                if col.lower() == 'origin':
                    origin_col = col
                    break
            
            if origin_col:
                origins_in_file = temp_df[origin_col].dropna().unique().tolist()
                if len(origins_in_file) > 0:
                    all_origins.update(origins_in_file)
                    files_with_origins += 1
            
            file_obj.seek(0)  # Reset file pointer
        except Exception as e:
            st.warning(f"Could not read origins from {name}: {e}")

# Categorize origins (case-insensitive matching)
EPIC_NAMES = {name.lower() for name in EPIC_ORIGINS}
ANCHOR_NAMES = {name.lower() for name in ANCHOR_ORIGINS}

epic_origins = [o for o in all_origins if o.lower() in EPIC_NAMES]
anchor_origins = [o for o in all_origins if o.lower() in ANCHOR_NAMES]
other_origins = [o for o in all_origins if o.lower() not in EPIC_NAMES and o.lower() not in ANCHOR_NAMES]

# Sort alphabetically
epic_origins.sort()
anchor_origins.sort()
other_origins.sort()

if all_origins:
    st.success(f"✅ Detected {len(all_origins)} origins from {files_with_origins} file(s): {len(epic_origins)} Epic, {len(anchor_origins)} Anchor, {len(other_origins)} Other")
    
    # Debug expander to show detected origins
    with st.expander("🔍 Show detected origins", expanded=False):
        if epic_origins:
            st.write(f"**Epic Origins ({len(epic_origins)}):** {', '.join(epic_origins)}")
        if anchor_origins:
            st.write(f"**Anchor Origins ({len(anchor_origins)}):** {', '.join(anchor_origins)}")
        if other_origins:
            st.write(f"**Other Origins ({len(other_origins)}):** {', '.join(other_origins)}")
            
elif files_checked > 0:
    st.warning(f"⚠️ Checked {files_checked} file(s) but found no 'origin' column or no origin data. Origin filtering will be disabled.")
    
    # Debug expander
    with st.expander("🔍 Troubleshooting - Show file columns", expanded=False):
        st.write("Checking for 'origin' column in uploaded files...")
        for file_obj, name in [(small_3m_file, '3m small'), (small_5m_file, '5m small'), (small_15m_file, '15m small'), 
                                (big_3m_file, '3m big'), (big_5m_file, '5m big'), (big_15m_file, '15m big')]:
            if file_obj is not None:
                try:
                    temp_df = pd.read_csv(file_obj)
                    st.write(f"**{name}:** Columns found: {list(temp_df.columns)}")
                    file_obj.seek(0)
                except Exception as e:
                    st.write(f"**{name}:** Error reading file: {e}")
else:
    st.info("Upload feed files to enable origin filtering")

col_filter1, col_filter2 = st.columns([2, 1])
with col_filter1:
    filter_origins = st.checkbox(
        "Filter Origins (Faster processing, smaller results)", 
        value=False, 
        key="filter_origins_main",
        disabled=(len(all_origins) == 0)
    )
    if len(all_origins) > 0:
        st.caption("Process only selected origins. All priority origins (Epic + Anchor) selected by default.")
    else:
        st.caption("Origin filtering unavailable - no origins detected in uploaded files.")

allowed_origins = None
if filter_origins and all_origins:
    # Epic Origins
    if epic_origins:
        st.markdown("**Epic Origins (Priority):**")
        epic_selections = {}
        cols = st.columns(min(len(epic_origins), 4))
        for idx, origin in enumerate(epic_origins):
            with cols[idx % len(cols)]:
                epic_selections[origin] = st.checkbox(origin, value=True, key=f"epic_main_{origin}")
    
    # Anchor Origins
    if anchor_origins:
        st.markdown("**Anchor Origins (Priority):**")
        anchor_selections = {}
        cols = st.columns(min(len(anchor_origins), 5))
        for idx, origin in enumerate(anchor_origins):
            with cols[idx % len(cols)]:
                anchor_selections[origin] = st.checkbox(origin, value=True, key=f"anchor_main_{origin}")
    
    # Other Origins
    if other_origins:
        st.markdown("**Other Origins (Optional):**")
        other_selections = {}
        cols = st.columns(min(len(other_origins), 4))
        for idx, origin in enumerate(other_origins):
            with cols[idx % len(cols)]:
                other_selections[origin] = st.checkbox(origin, value=False, key=f"other_main_{origin}")
    
    # Build allowed origins set (use exact names from CSV, case-sensitive)
    allowed_origins = set()
    if epic_origins:
        allowed_origins.update([o for o, selected in epic_selections.items() if selected])
    if anchor_origins:
        allowed_origins.update([o for o, selected in anchor_selections.items() if selected])
    if other_origins:
        allowed_origins.update([o for o, selected in other_selections.items() if selected])
    
    epic_count = sum(1 for o, s in (epic_selections.items() if epic_origins else []) if s)
    anchor_count = sum(1 for o, s in (anchor_selections.items() if anchor_origins else []) if s)
    other_count = sum(1 for o, s in (other_selections.items() if other_origins else []) if s)
    
    st.info(f"Processing {len(allowed_origins)} origins ({epic_count} Epic + {anchor_count} Anchor + {other_count} Other)")
elif filter_origins and not all_origins:
    st.warning("Origin filtering enabled but no origins detected. Processing will continue with all data.")
else:
    if all_origins:
        st.info(f"Processing ALL {len(all_origins)} origins (may take longer with large results)")
    else:
        st.info("Processing all data (no origin filtering available)")

st.markdown("---")

# Report Time UI
report_mode = st.radio("Select Report Time & Date", ["Most Current", "Choose a time"])
if report_mode == "Choose a time":
    selected_date = st.date_input("Select Report Date", value=dt.date.today())
    selected_time = st.time_input("Select Report Time", value=dt.time(18, 0))
    report_time = dt.datetime.combine(selected_date, selected_time)
else:
    # Use current time as default for "Most Current" mode
    report_time = dt.datetime.now()

# Toggles
run_g_models = st.sidebar.checkbox("🟢 Run Model G Detection", value=False)
if run_g_models:
    st.sidebar.markdown("**G Model Controls:**")
    
    # Global output spread filter - applies to ALL G models
    output_spread_filter = st.sidebar.slider(
        "Output Spread Filter (pts)", 
        min_value=0.1, 
        max_value=20.0, 
        value=3.0, 
        step=0.1,
        help="Maximum output spread (max - min) for all G models. Lower values = tighter clusters."
    )
    
    # Show Rejected Groups - applies to all G models
    show_rejected_groups = st.sidebar.checkbox("Show Rejected Groups", value=False,
        help="Display groups/pairs that were rejected during detection")
    
    st.sidebar.markdown("---")
    
    run_g05_g06 = st.sidebar.checkbox("   • G.05 & G.06 (Proximity Groups)", value=False)
    run_g08 = st.sidebar.checkbox("   • G.08 (x0Pd.w Patterns)", value=False)
    run_g09 = st.sidebar.checkbox("   • G.09 (Flip Endings)", value=False)
    run_g10 = st.sidebar.checkbox("   • G.10 (Pair Detection)", value=False)
    
    if run_g10:
        st.sidebar.markdown("**G.10 Group Controls:**")
        g10_group_0 = st.sidebar.checkbox("      ○ Group 0", value=True)
        g10_group_1 = st.sidebar.checkbox("      ○ Group 1", value=True)
        g10_group_2 = st.sidebar.checkbox("      ○ Group 2", value=True)
        g10_group_3 = st.sidebar.checkbox("      ○ Group 3", value=False)
        g10_group_4 = st.sidebar.checkbox("      ○ Group 4", value=False)
    else:
        g10_group_0 = g10_group_1 = g10_group_2 = g10_group_3 = g10_group_4 = False
    
    run_g11 = st.sidebar.checkbox("   • G.11 (Pair Detection SF - Same Feed)", value=False)
    
    if run_g11:
        st.sidebar.markdown("**G.11 Group Controls:**")
        g11_group_0 = st.sidebar.checkbox("      ○ Grp 0 TA", value=True)
        g11_group_1 = st.sidebar.checkbox("      ○ Grp 1 sAA", value=True)
        g11_group_2 = st.sidebar.checkbox("      ○ Grp 2 AA", value=True)
        g11_group_3 = st.sidebar.checkbox("      ○ Grp 3 oA", value=True)
        g11_group_4 = st.sidebar.checkbox("      ○ Grp 4 Ao", value=True)
        st.sidebar.markdown("**G.11 Display Filters:**")
        g11_display_recipes = st.sidebar.checkbox("      Display Recips", value=True)
        g11_display_others = st.sidebar.checkbox("      Display others", value=True)
    else:
        g11_group_0 = g11_group_1 = g11_group_2 = g11_group_3 = g11_group_4 = True
        g11_display_recipes = g11_display_others = True
    
    debug_g08 = st.sidebar.checkbox("Debug G.08 Detection", value=False)
    if debug_g08:
        st.session_state['debug_g08'] = True
    else:
        st.session_state['debug_g08'] = False
    
    debug_g_models = st.sidebar.checkbox("Debug G Models (show DataFrame info)", value=False)
    if debug_g_models:
        st.session_state['debug_g_models'] = True
    else:
        st.session_state['debug_g_models'] = False
else:
    run_g05_g06 = False
    run_g08 = False
    run_g09 = False
    run_g10 = False
    run_g11 = False
    output_spread_filter = 3.0
    show_rejected_groups = False
    g10_group_0 = g10_group_1 = g10_group_2 = g10_group_3 = g10_group_4 = False
    g11_group_0 = g11_group_1 = g11_group_2 = g11_group_3 = g11_group_4 = True
    g11_display_recipes = g11_display_others = True
    debug_g08 = False
    debug_g_models = False

run_a_models = st.sidebar.checkbox("🔵 Run Model A Detection", value=False)
run_b_models = st.sidebar.checkbox("🟡 Run Model B Detection", value=False)
run_c_models = st.sidebar.checkbox("🟠 Run Model C Detection", value=False)
if run_c_models:
    st.sidebar.markdown("**C Model Sub-detections:**")
    run_c01 = st.sidebar.checkbox("   • C.01", value=True)
    run_c02 = st.sidebar.checkbox("   • C.02", value=True)
    run_c04 = st.sidebar.checkbox("   • C.04", value=True)
else:
    run_c01 = run_c02 = run_c04 = False

run_x_models = st.sidebar.checkbox("🟣 Run Model X Detection", value=False)
run_single_line = st.sidebar.checkbox("📊 Simple Single Line Mega Report", value=False)

# Performance toggles
fast_mode = st.sidebar.checkbox("⚡ Fast Mode (skip dataframe displays)", value=False)
minimal_display = st.sidebar.checkbox("📉 Minimal Display (summaries only)", value=False)
filter_future_data = st.sidebar.checkbox("🔮 Filter Future Data", value=False, 
    help="Remove entries with arrival times after report time")

st.sidebar.markdown("---")
st.sidebar.markdown("### ⚙️ Processing Mode")
processing_mode = st.sidebar.radio(
    "Select Processing Mode",
    ["Full Range (single)", "Full Range (multiple)", "Custom Ranges", "HOD/LOD Mode"],
    help="Choose how to process the data"
)

# Full Range (single) options
if processing_mode == "Full Range (single)":
    st.sidebar.markdown("**Full Range (single) Settings:**")
    window_radius = st.sidebar.number_input("Window Radius", value=60.0, min_value=1.0, step=1.0,
        help="Full range window radius around center value")
    input_value_at_start = st.sidebar.number_input("Input Value @ Start (optional)", value=0.0,
        help="Leave at 0 to auto-calculate from feed data")
    run_g_on_full = st.sidebar.checkbox("Run Model G on Full Range", value=False)

# Full Range (multiple) options
if processing_mode == "Full Range (multiple)":
    st.sidebar.markdown("**Full Range (multiple) Settings:**")
    st.sidebar.info("Processes up to 30 days with 4 grab times per day (1800, 0315, 0900, 1230)")
    
    num_days_multi = st.sidebar.number_input(
        "Number of Days", 
        value=1, 
        min_value=1, 
        max_value=30,
        help="Number of trading days to process"
    )
    
    multi_start_date = st.sidebar.date_input(
        "Starting Date (1800 open)", 
        value=dt.date.today() - dt.timedelta(days=1),
        help="The date of the first 1800 open (e.g., 7-Dec for 8-Dec trading day)"
    )
    
    window_radius_multi = st.sidebar.number_input(
        "Window Radius", 
        value=60.0, 
        min_value=1.0, 
        step=1.0,
        help="Full range window radius around center value"
    )
    
    input_value_at_start_multi = st.sidebar.number_input(
        "Input Value @ Start (optional)", 
        value=0.0,
        help="Leave at 0 to auto-calculate from feed data"
    )

# Custom Ranges options
if processing_mode == "Custom Ranges":
    st.sidebar.markdown("**Custom Range Settings:**")
    st.sidebar.info("Define up to 4 custom ranges (2 High, 2 Low)")
    
    use_high1 = st.sidebar.checkbox("Enable High Range 1", value=False)
    high1 = st.sidebar.number_input("High Range 1 Center", value=0.0, step=0.1,
        help="Range will be [value-24, value]") if use_high1 else 0.0
    
    use_high2 = st.sidebar.checkbox("Enable High Range 2", value=False)
    high2 = st.sidebar.number_input("High Range 2 Center", value=0.0, step=0.1,
        help="Range will be [value-24, value]") if use_high2 else 0.0
    
    use_low1 = st.sidebar.checkbox("Enable Low Range 1", value=False)
    low1 = st.sidebar.number_input("Low Range 1 Center", value=0.0, step=0.1,
        help="Range will be [value, value+24]") if use_low1 else 0.0
    
    use_low2 = st.sidebar.checkbox("Enable Low Range 2", value=False)
    low2 = st.sidebar.number_input("Low Range 2 Center", value=0.0, step=0.1,
        help="Range will be [value, value+24]") if use_low2 else 0.0
    
    run_g_on_custom = st.sidebar.checkbox("Run Model G on Custom Ranges", value=False)

# HOD/LOD options
if processing_mode == "HOD/LOD Mode":
    st.sidebar.markdown("**HOD/LOD Settings:**")
    hod_lod_num_days = st.sidebar.number_input("Number of Days to Analyze", value=1, min_value=1, max_value=30,
        help="Number of complete trading days to analyze")
    include_partial_day = st.sidebar.checkbox("Include Partial Current Day", value=True,
        help="Include today's session even if incomplete")
    run_g_on_hod_lod = st.sidebar.checkbox("Run Model G on HOD/LOD", value=False)

# Processing Optimization
st.sidebar.markdown("---")
st.sidebar.markdown("### ⚡ Processing Optimization")
separate_timeframes = st.sidebar.checkbox(
    "Process Timeframes Separately",
    value=True,
    help="Process each timeframe independently for better performance (recommended for Custom Ranges and Full Range)"
)

# Day start time configuration
st.sidebar.markdown("---")
st.sidebar.markdown("### 🕐 Day Start Configuration")
day_start_hour = st.sidebar.slider("Day Start Hour", 0, 23, 17, help="Hour when trading day starts (default: 17 = 5 PM)")

# Run button
if st.button("🚀 Process Data"):
    # Check required files
    if not bypass_traveler_file:
        has_small = any([small_3m_file, small_5m_file, small_15m_file])
        has_big = any([big_3m_file, big_5m_file, big_15m_file])
        has_measurement = measurement_file is not None
        
        if not (has_small or has_big) or not has_measurement:
            st.error("⚠️ Please upload at least one feed file (small or big) and a measurement file, or upload a Final Traveler Report")
            st.stop()
    
    # === Load bypass traveler file if provided ===
    bypass_df = None
    if bypass_traveler_file:
        try:
            if bypass_traveler_file.name.endswith('.csv'):
                bypass_df = pd.read_csv(bypass_traveler_file)
            else:
                bypass_df = pd.read_excel(bypass_traveler_file)
            st.success(f"✅ Loaded bypass traveler report with {len(bypass_df)} entries")
        except Exception as e:
            st.error(f"❌ Error loading bypass file: {e}")
            st.stop()
    
    # === Load and process feeds ===
    small_feeds_dict = {}
    big_feeds_dict = {}
    
    if not bypass_traveler_file:
        # Process small feeds
        for file, timeframe in [(small_3m_file, '3m'), (small_5m_file, '5m'), (small_15m_file, '15m')]:
            if file:
                try:
                    df = pd.read_csv(file)
                    
                    # Apply origin filtering
                    if allowed_origins is not None and 'origin' in df.columns:
                        df = df[df['origin'].isin(allowed_origins)].copy()
                        if df.empty:
                            st.warning(f"⚠️ No data for selected origins in {timeframe} small feed")
                            continue
                    
                    processed = process_feed(df, is_small=True)
                    if processed is not None and not processed.empty:
                        small_feeds_dict[timeframe] = processed
                        st.success(f"✅ Loaded {timeframe} small feed ({len(processed)} rows)")
                except Exception as e:
                    st.error(f"❌ Error loading {timeframe} small: {e}")
        
        # Process big feeds
        for file, timeframe in [(big_3m_file, '3m'), (big_5m_file, '5m'), (big_15m_file, '15m')]:
            if file:
                try:
                    df = pd.read_csv(file)
                    
                    # Apply origin filtering
                    if allowed_origins is not None and 'origin' in df.columns:
                        df = df[df['origin'].isin(allowed_origins)].copy()
                        if df.empty:
                            st.warning(f"⚠️ No data for selected origins in {timeframe} big feed")
                            continue
                    
                    processed = process_feed(df, is_small=False)
                    if processed is not None and not processed.empty:
                        big_feeds_dict[timeframe] = processed
                        st.success(f"✅ Loaded {timeframe} big feed ({len(processed)} rows)")
                except Exception as e:
                    st.error(f"❌ Error loading {timeframe} big: {e}")
        
        # Combine feeds if not processing separately
        if not separate_timeframes:
            if small_feeds_dict:
                small_df = pd.concat(small_feeds_dict.values(), ignore_index=True)
                st.success(f"✅ Combined small feeds: {len(small_df)} rows")
            else:
                small_df = None
            
            if big_feeds_dict:
                big_df = pd.concat(big_feeds_dict.values(), ignore_index=True)
                st.success(f"✅ Combined big feeds: {len(big_df)} rows")
            else:
                big_df = None
        else:
            small_df = None
            big_df = None
        
        # Load measurement file
        try:
            measurements_df = pd.read_excel(measurement_file)
            st.success(f"✅ Loaded measurement file with {len(measurements_df)} rows")
        except Exception as e:
            st.error(f"❌ Error loading measurement file: {e}")
            st.stop()
    
    # === PROCESSING MODE LOGIC ===
    
    use_full_range_single = (processing_mode == "Full Range (single)")
    use_full_range_multi = (processing_mode == "Full Range (multiple)")
    use_custom_ranges = (processing_mode == "Custom Ranges")
    use_hod_lod = (processing_mode == "HOD/LOD Mode")
    
    traveler_reports = {}
    
    if bypass_df is not None:
        st.markdown("---")
        st.markdown("### 📊 Bypass Mode: Using Uploaded Traveler Report")
        
        # Split into groups for display
        traveler_reports["Grp 1a"] = bypass_df[bypass_df['M #'].isin(GROUP_1A_TRAVELERS)].copy()
        traveler_reports["Grp 1b"] = bypass_df[bypass_df['M #'].isin(GROUP_1B_TRAVELERS)].copy()
        traveler_reports["Grp 2a"] = bypass_df[bypass_df['M #'].isin(GROUP_2A_TRAVELERS)].copy()
        traveler_reports["Grp 2b"] = bypass_df[bypass_df['M #'].isin(GROUP_2B_TRAVELERS)].copy()
        
        for gname, gdf in traveler_reports.items():
            if not gdf.empty:
                st.markdown(f"#### {gname}")
                st.info(f"{len(gdf)} entries")
                if not fast_mode:
                    st.dataframe(gdf, use_container_width=True)
    
    else:
        # ---------- 1) FULL RANGE (SINGLE) MODE ----------
        if use_full_range_single:
            st.markdown("---")
            st.markdown("### Full Range (single) Processing Mode")
            
            # Determine input value
            input_val = input_value_at_start if input_value_at_start != 0.0 else None
            
            if separate_timeframes and (small_feeds_dict or big_feeds_dict):
                # Process each timeframe separately and combine results
                st.info("Processing timeframes separately for optimal performance...")
                all_timeframe_results = []
                
                timeframes = sorted(set(list(small_feeds_dict.keys()) + list(big_feeds_dict.keys())))
                
                for tf in timeframes:
                    st.text(f"Processing {tf} timeframe...")
                    tf_small = small_feeds_dict.get(tf, pd.DataFrame())
                    tf_big = big_feeds_dict.get(tf, pd.DataFrame())
                    
                    if tf_small.empty and tf_big.empty:
                        continue
                    
                    # Process this timeframe
                    tf_result = apply_full_range_advanced(
                        measurements_df,
                        tf_small if not tf_small.empty else pd.DataFrame(),
                        report_time,
                        window_radius,
                        day_start_hour=day_start_hour,
                        input_value_at_start=input_val,
                        big_df=tf_big if not tf_big.empty else pd.DataFrame(),
                        run_model_g=False  # Run G models only on combined data
                    )
                    
                    if tf_result is not None and not tf_result.empty:
                        all_timeframe_results.append(tf_result)
                
                # Combine all timeframe results
                if all_timeframe_results:
                    final_df_filtered = pd.concat(all_timeframe_results, ignore_index=True)
                else:
                    final_df_filtered = pd.DataFrame()
            else:
                # Combined processing (original logic)
                if small_df is None and big_df is None:
                    st.error("No feed data available for combined processing. Please disable 'Process Timeframes Separately' or ensure files are uploaded.")
                    final_df_filtered = pd.DataFrame()
                else:
                    final_df_filtered = apply_full_range_advanced(
                        measurements_df,
                        small_df if small_df is not None else pd.DataFrame(),
                        report_time,
                        window_radius,
                        day_start_hour=day_start_hour,
                        input_value_at_start=input_val,
                        big_df=big_df if big_df is not None else pd.DataFrame(),
                        run_model_g=run_g_on_full
                    )
            
            if final_df_filtered is None or final_df_filtered.empty:
                st.warning("No entries found in full range processing")
                traveler_reports = {}
            else:
                # Build 4 groups
                traveler_reports = {}
                traveler_reports["Grp 1a"] = final_df_filtered[final_df_filtered['M #'].isin(GROUP_1A_TRAVELERS)].copy()
                traveler_reports["Grp 1b"] = final_df_filtered[final_df_filtered['M #'].isin(GROUP_1B_TRAVELERS)].copy()
                traveler_reports["Grp 2a"] = final_df_filtered[final_df_filtered['M #'].isin(GROUP_2A_TRAVELERS)].copy()
                traveler_reports["Grp 2b"] = final_df_filtered[final_df_filtered['M #'].isin(GROUP_2B_TRAVELERS)].copy()
                
                # Sort each group
                for gname, gdf in traveler_reports.items():
                    if not gdf.empty and 'Output' in gdf.columns and 'Arrival' in gdf.columns:
                        traveler_reports[gname] = gdf.sort_values(['Output', 'Arrival'], ascending=[False, True])
                    elif not gdf.empty and 'Output' in gdf.columns:
                        traveler_reports[gname] = gdf.sort_values(['Output'], ascending=[False])
                
                # Display compact summaries
                if not minimal_display:
                    for gname, gdf in traveler_reports.items():
                        if not gdf.empty:
                            st.markdown(f"#### {gname}")
                            st.info(f"{len(gdf)} entries")
                            if not fast_mode:
                                st.dataframe(gdf, use_container_width=True)
                
                st.success("Full range (single) processing complete.")
        
        # ---------- 2) FULL RANGE (MULTIPLE) MODE ----------
        elif use_full_range_multi:
            st.markdown("---")
            st.markdown("### Full Range (multiple) Processing Mode")
            st.info(f"Processing {num_days_multi} days with 4 grab times per day (1800, 0315, 0900, 1230)")
            
            # Determine input value
            input_val = input_value_at_start_multi if input_value_at_start_multi != 0.0 else None
            
            # Generate list of all datetimes to process
            grab_times = [(18, 0), (3, 15), (9, 0), (12, 30)]  # (hour, minute) tuples
            all_datetimes = []
            
            for day_offset in range(num_days_multi):
                current_date = multi_start_date + dt.timedelta(days=day_offset)
                
                for hour, minute in grab_times:
                    grab_dt = dt.datetime.combine(current_date, dt.time(hour, minute))
                    all_datetimes.append(grab_dt)
            
            # Process each datetime
            all_day_reports = []
            
            progress_bar = st.progress(0)
            progress_text = st.empty()
            
            for idx, grab_dt in enumerate(all_datetimes):
                progress_text.text(f"Processing {format_tab_name(grab_dt)}...")
                
                # Process this grab time
                if separate_timeframes and (small_feeds_dict or big_feeds_dict):
                    all_timeframe_results = []
                    timeframes = sorted(set(list(small_feeds_dict.keys()) + list(big_feeds_dict.keys())))
                    
                    for tf in timeframes:
                        tf_small = small_feeds_dict.get(tf, pd.DataFrame())
                        tf_big = big_feeds_dict.get(tf, pd.DataFrame())
                        
                        if tf_small.empty and tf_big.empty:
                            continue
                        
                        tf_result = apply_full_range_advanced(
                            measurements_df,
                            tf_small if not tf_small.empty else pd.DataFrame(),
                            grab_dt,
                            window_radius_multi,
                            day_start_hour=day_start_hour,
                            input_value_at_start=input_val,
                            big_df=tf_big if not tf_big.empty else pd.DataFrame(),
                            run_model_g=False
                        )
                        
                        if tf_result is not None and not tf_result.empty:
                            all_timeframe_results.append(tf_result)
                    
                    if all_timeframe_results:
                        final_df = pd.concat(all_timeframe_results, ignore_index=True)
                    else:
                        final_df = pd.DataFrame()
                else:
                    if small_df is None and big_df is None:
                        st.error("No feed data available")
                        final_df = pd.DataFrame()
                    else:
                        final_df = apply_full_range_advanced(
                            measurements_df,
                            small_df if small_df is not None else pd.DataFrame(),
                            grab_dt,
                            window_radius_multi,
                            day_start_hour=day_start_hour,
                            input_value_at_start=input_val,
                            big_df=big_df if big_df is not None else pd.DataFrame(),
                            run_model_g=False
                        )
                
                # Split into Grp_1a and Grp_1b, combine and format
                if final_df is not None and not final_df.empty:
                    grp_1a = final_df[final_df['M #'].isin(GROUP_1A_TRAVELERS)].copy()
                    grp_1b = final_df[final_df['M #'].isin(GROUP_1B_TRAVELERS)].copy()
                    
                    combined = combine_and_format_groups(grp_1a, grp_1b)
                    
                    tab_name = format_tab_name(grab_dt)
                    all_day_reports.append((tab_name, combined))
                
                # Update progress
                progress_bar.progress((idx + 1) / len(all_datetimes))
            
            progress_text.text("Processing complete!")
            
            # Create Excel file with all tabs
            if all_day_reports:
                st.success(f"✅ Generated {len(all_day_reports)} report tabs")
                
                # Determine trading day for filename
                trading_day = get_trading_day_from_datetime(all_datetimes[0])
                trading_day_str = trading_day.strftime("%d-%b")
                
                # Create Excel file
                excel_buffer = create_multi_day_excel(
                    all_day_reports, 
                    asset_id, 
                    trading_day
                )
                
                filename = f"{asset_id.lower()}_traveler_reports_for_{trading_day_str}.xlsx"
                
                st.download_button(
                    label="📥 Download Multi-Day Report",
                    data=excel_buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Display summary
                st.markdown("### Report Summary")
                for tab_name, df in all_day_reports:
                    st.text(f"{tab_name}: {len(df)} entries")
            else:
                st.warning("No data generated for the selected date range")
        
        # ---------- 3) CUSTOM RANGES MODE ----------
        elif use_custom_ranges:
            st.markdown("---")
            st.markdown("### Custom Ranges Processing Mode")
            
            if separate_timeframes and (small_feeds_dict or big_feeds_dict):
                # Process each timeframe separately and combine results
                st.info("Processing timeframes separately for optimal performance...")
                all_timeframe_results = []
                
                timeframes = sorted(set(list(small_feeds_dict.keys()) + list(big_feeds_dict.keys())))
                
                for tf in timeframes:
                    st.text(f"Processing {tf} timeframe...")
                    tf_small = small_feeds_dict.get(tf, pd.DataFrame())
                    tf_big = big_feeds_dict.get(tf, pd.DataFrame())
                    
                    if tf_small.empty and tf_big.empty:
                        continue
                    
                    # Process this timeframe
                    tf_result = apply_custom_ranges_advanced(
                        measurements_df,
                        tf_small if not tf_small.empty else None,
                        report_time,
                        high1, high2, low1, low2,
                        use_high1, use_high2, use_low1, use_low2,
                        big_df=tf_big if not tf_big.empty else None,
                        run_model_g=False,  # Run G models only on combined data
                        day_start_hour=day_start_hour
                    )
                    
                    if tf_result is not None and not tf_result.empty:
                        all_timeframe_results.append(tf_result)
                
                # Combine all timeframe results
                if all_timeframe_results:
                    final_df_filtered = pd.concat(all_timeframe_results, ignore_index=True)
                else:
                    final_df_filtered = pd.DataFrame()
            else:
                # Combined processing (original logic)
                if small_df is None and big_df is None:
                    st.error("No feed data available for combined processing. Please disable 'Process Timeframes Separately' or ensure files are uploaded.")
                    final_df_filtered = pd.DataFrame()
                else:
                    final_df_filtered = apply_custom_ranges_advanced(
                        measurements_df,
                        small_df if small_df is not None else pd.DataFrame(),
                        report_time,
                        high1, high2, low1, low2,
                        use_high1, use_high2, use_low1, use_low2,
                        big_df=big_df if big_df is not None else pd.DataFrame(),
                        run_model_g=run_g_on_custom,
                        day_start_hour=day_start_hour
                    )
            
            if final_df_filtered is None or final_df_filtered.empty:
                st.warning("No entries found in custom range processing")
                traveler_reports = {}
            else:
                # Single-tab output sorted by Range, Group, Output, Arrival
                sort_cols = []
                if 'Range' in final_df_filtered.columns:
                    sort_cols.append('Range')
                if 'Group' in final_df_filtered.columns:
                    sort_cols.append('Group')
                if 'Output' in final_df_filtered.columns:
                    sort_cols.append('Output')
                if 'Arrival' in final_df_filtered.columns:
                    sort_cols.append('Arrival')
                
                if sort_cols:
                    # Define sort order: Range and Group ascending, Output descending, Arrival ascending
                    ascending_list = []
                    for col in sort_cols:
                        if col in ['Range', 'Group']:
                            ascending_list.append(True)
                        elif col == 'Output':
                            ascending_list.append(False)
                        else:  # Arrival
                            ascending_list.append(True)
                    
                    final_df_filtered = final_df_filtered.sort_values(sort_cols, ascending=ascending_list)
                
                # Assign group labels
                final_df_filtered['Group'] = final_df_filtered['M #'].apply(
                    lambda m_num: 
                        '1a' if m_num in GROUP_1A_TRAVELERS else
                        '1b' if m_num in GROUP_1B_TRAVELERS else
                        '2a' if m_num in GROUP_2A_TRAVELERS else
                        '2b' if m_num in GROUP_2B_TRAVELERS else
                        'Other'
                )
                
                traveler_reports = {"Custom Ranges": final_df_filtered}
                
                if not minimal_display:
                    st.markdown("#### Custom Ranges Report")
                    st.info(f"{len(final_df_filtered)} total entries")
                    if not fast_mode:
                        st.dataframe(final_df_filtered, use_container_width=True)
                
                st.success("Custom ranges processing complete.")
        
        # ---------- 4) HOD/LOD MODE ----------
        elif use_hod_lod:
            st.markdown("---")
            st.markdown("### HOD/LOD Processing Mode")
            
            if small_df is None:
                st.error("HOD/LOD mode requires at least one small feed file")
            else:
                # Process HOD/LOD
                results = process_hod_lod_mode(
                    small_df,
                    measurements_df,
                    hod_lod_num_days,
                    include_partial_day,
                    day_start_hour
                )
                
                if results:
                    st.success(f"✅ Processed {len(results)} trading days")
                    
                    # Build traveler reports from HOD/LOD results
                    for day_label, day_data in results.items():
                        st.markdown(f"#### {day_label}")
                        
                        if 'hod_travelers' in day_data and not day_data['hod_travelers'].empty:
                            st.markdown("**HOD Travelers:**")
                            st.dataframe(day_data['hod_travelers'], use_container_width=True)
                        
                        if 'lod_travelers' in day_data and not day_data['lod_travelers'].empty:
                            st.markdown("**LOD Travelers:**")
                            st.dataframe(day_data['lod_travelers'], use_container_width=True)
                    
                    # Create download for HOD/LOD results
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        for day_label, day_data in results.items():
                            if 'hod_travelers' in day_data and not day_data['hod_travelers'].empty:
                                day_data['hod_travelers'].to_excel(writer, sheet_name=f"{day_label}_HOD", index=False)
                            if 'lod_travelers' in day_data and not day_data['lod_travelers'].empty:
                                day_data['lod_travelers'].to_excel(writer, sheet_name=f"{day_label}_LOD", index=False)
                    
                    output.seek(0)
                    st.download_button(
                        label="📥 Download HOD/LOD Report",
                        data=output,
                        file_name=f"{asset_id.lower()}_hod_lod_report_{dt.datetime.now().strftime('%d-%b-%y_%H-%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("No HOD/LOD data generated")
    
    # === MODEL DETECTION (runs on traveler_reports) ===
    if traveler_reports and not bypass_df:
        st.markdown("---")
        st.markdown("### 🔍 Model Detection")
        
        # Combine all reports for model detection
        combined_for_models = pd.concat([df for df in traveler_reports.values() if not df.empty], ignore_index=True)
        
        if filter_future_data:
            if 'Arrival' in combined_for_models.columns:
                combined_for_models['Arrival'] = pd.to_datetime(combined_for_models['Arrival'], errors='coerce')
                before_filter = len(combined_for_models)
                combined_for_models = combined_for_models[combined_for_models['Arrival'] <= report_time]
                after_filter = len(combined_for_models)
                st.info(f"🔮 Filtered future data: {before_filter} → {after_filter} entries ({before_filter - after_filter} removed)")
        
        # Run Model G
        if run_g_models and not combined_for_models.empty:
            st.markdown("#### 🟢 Model G Detection")
            
            g_results = run_model_g_detection(
                combined_for_models,
                run_g05_g06=run_g05_g06,
                run_g08=run_g08,
                run_g09=run_g09,
                run_g10=run_g10,
                run_g11=run_g11,
                output_spread_filter=output_spread_filter,
                show_rejected_groups=show_rejected_groups,
                g10_group_0=g10_group_0,
                g10_group_1=g10_group_1,
                g10_group_2=g10_group_2,
                g10_group_3=g10_group_3,
                g10_group_4=g10_group_4,
                g11_group_0=g11_group_0,
                g11_group_1=g11_group_1,
                g11_group_2=g11_group_2,
                g11_group_3=g11_group_3,
                g11_group_4=g11_group_4,
                g11_display_recipes=g11_display_recipes,
                g11_display_others=g11_display_others
            )
            
            if g_results:
                for model_name, model_df in g_results.items():
                    if not model_df.empty:
                        st.markdown(f"**{model_name}:** {len(model_df)} detections")
                        if not fast_mode:
                            st.dataframe(model_df, use_container_width=True)
        
        # Run Model A
        if run_a_models and not combined_for_models.empty:
            st.markdown("#### 🔵 Model A Detection")
            a_result = run_a_model_detection_today(combined_for_models)
            if a_result is not None and not a_result.empty:
                st.dataframe(a_result, use_container_width=True)
        
        # Run Model B
        if run_b_models and not combined_for_models.empty:
            st.markdown("#### 🟡 Model B Detection")
            b_result = run_b_model_detection(combined_for_models)
            if b_result is not None and not b_result.empty:
                st.dataframe(b_result, use_container_width=True)
        
        # Run Model C
        if run_c_models and not combined_for_models.empty:
            st.markdown("#### 🟠 Model C Detection")
            c_result = run_c_model_detection(combined_for_models, run_c01, run_c02, run_c04)
            if c_result is not None and not c_result.empty:
                st.dataframe(c_result, use_container_width=True)
        
        # Run Model X
        if run_x_models and not combined_for_models.empty:
            st.markdown("#### 🟣 Model X Detection")
            x_result = run_x_model_detection(combined_for_models)
            if x_result is not None and not x_result.empty:
                st.dataframe(x_result, use_container_width=True)
        
        # Run Simple Single Line Mega Report
        if run_single_line and not combined_for_models.empty:
            st.markdown("#### 📊 Simple Single Line Mega Report")
            single_line_result = run_simple_single_line_analysis(combined_for_models)
            if single_line_result is not None and not single_line_result.empty:
                st.dataframe(single_line_result, use_container_width=True)
    
    # === EXPORT (only for non-multi modes) ===
    if traveler_reports and not use_full_range_multi:
        render_unified_export(traveler_reports, report_time, asset_id)

st.markdown("---")
st.caption("🌌 Traveler Report Generator v33 | Full Range (single), Full Range (multiple), Custom Ranges, HOD/LOD Mode with Origin Filtering")
